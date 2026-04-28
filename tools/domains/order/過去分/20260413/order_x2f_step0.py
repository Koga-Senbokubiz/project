# -*- coding: utf-8 -*-
"""
order_x2f_step0.py

Step0:
Schema_Order_20090901.xlsx と 顧客データxml をもとに、
Order / v1_3 用の項目辞書 order_field_dictionary.xlsx を生成・補完する。

新方針:
- Step0は「辞書の初期作成」と「不足補完」のみを担当する
- 辞書がすでに存在し、必要列・必要行がそろっていれば何もしない
- Step2でアンマッチが出た場合は、Step0辞書を手修正し、Step2を再実行する
- order_x2f_step2.xlsx は作業用ファイル、項目辞書がマスター
- 将来のリポジトリ化を見据え、message_type / version を列として持つ
"""

import os
import re
import sys
import argparse
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET


# =========================================================
# 共通
# =========================================================

def log(msg: str) -> None:
    print(msg, flush=True)


def nz(value, default: str = "") -> str:
    if value is None:
        return default
    return str(value).strip()


def ensure_dir(path: str) -> None:
    d = os.path.dirname(path)
    if d:
        os.makedirs(d, exist_ok=True)


def local_name(tag: str) -> str:
    """
    {ns}tag / prefix:tag / tag -> tag
    """
    s = nz(tag)
    if "}" in s:
        s = s.split("}", 1)[1]
    if ":" in s:
        s = s.split(":", 1)[1]
    return s


def normalize_tag(tag: str) -> str:
    """
    比較用の正規化タグ
    - prefix除去
    - underscore除去
    - 小文字化
    """
    s = local_name(tag)
    s = s.replace("_", "")
    s = s.lower()
    return s


def safe_text(elem: ET.Element) -> str:
    if elem.text is None:
        return ""
    return elem.text.strip()


# =========================================================
# 顧客XML読込
# =========================================================

def build_path(parent_path: str, elem: ET.Element) -> str:
    name = local_name(elem.tag)
    if parent_path == "":
        return f"/{name}"
    return f"{parent_path}/{name}"


def flatten_customer_leaf_nodes(
    elem: ET.Element,
    parent_path: str = "",
    rows: Optional[List[Dict]] = None,
) -> List[Dict]:
    if rows is None:
        rows = []

    path = build_path(parent_path, elem)
    children = list(elem)

    if children:
        for child in children:
            flatten_customer_leaf_nodes(child, path, rows)
    else:
        tag = local_name(elem.tag)
        rows.append({
            "customer_path": path,
            "customer_tag": tag,
            "normalized_tag": normalize_tag(tag),
            "sample_value": safe_text(elem),
        })

    return rows


def aggregate_customer_info(customer_rows: List[Dict]) -> Dict[str, Dict]:
    """
    normalized_tag ごとに顧客実績を集約
    """
    agg: Dict[str, Dict] = {}

    for row in customer_rows:
        key = row["normalized_tag"]
        if key not in agg:
            agg[key] = {
                "customer_tags": [],
                "sample_values": [],
                "path_examples": [],
                "occurs": 0,
            }

        bucket = agg[key]

        tag = row["customer_tag"]
        if tag and tag not in bucket["customer_tags"]:
            bucket["customer_tags"].append(tag)

        sv = row["sample_value"]
        if sv and sv not in bucket["sample_values"] and len(bucket["sample_values"]) < 5:
            bucket["sample_values"].append(sv)

        cp = row["customer_path"]
        if cp and cp not in bucket["path_examples"] and len(bucket["path_examples"]) < 5:
            bucket["path_examples"].append(cp)

        bucket["occurs"] += 1

    return agg


# =========================================================
# Schema 読込
# =========================================================

HIERARCHY_COLS = list(range(2, 12))   # B:K
ATTR_COL = 14                         # N
JP_NAME_COL = 15                      # O
SCHEMA_NO_COL = 16                    # P
SCHEMA_NAME_COL = 17                  # Q
NOTE_COL = 18                         # R
REQUIRED_COL = 19                     # S
REPEAT_COL = 20                       # T
TYPE_CLASS_COL = 21                   # U
DATA_ATTR_COL = 22                    # V
EXAMPLE_COL = 23                      # W


def normalize_required(value: str) -> str:
    s = nz(value)
    if "必須" in s:
        return "必須"
    if "任意" in s:
        return "任意"
    return s


def parse_repeat(value: str) -> Tuple[str, str]:
    """
    [1..1] -> ("1", "1")
    [0..*] -> ("0", "*")
    """
    s = nz(value)
    m = re.match(r"\[(.+)\.\.(.+)\]", s)
    if not m:
        return "", ""
    return nz(m.group(1)), nz(m.group(2))


def derive_length_from_data_attr(data_attr: str) -> str:
    """
    V列のデータ属性から長さをできる範囲で補完
    例:
      Text_Max20           -> 20
      Identifier_Num_Max13 -> 13
      Code_Num_2           -> 2
      Amount_10            -> 10
      Measure_7_3          -> 7,3
      Quantity_6           -> 6
    """
    s = nz(data_attr)

    m = re.match(r".*_(\d+)_(\d+)$", s)
    if m:
        return f"{m.group(1)},{m.group(2)}"

    m = re.match(r".*_Max(\d+)$", s)
    if m:
        return m.group(1)

    m = re.match(r".*_(\d+)$", s)
    if m:
        return m.group(1)

    return ""


def read_schema_rows(schema_xlsx: str, sheet_name: Optional[str] = None) -> List[Dict]:
    wb = openpyxl.load_workbook(schema_xlsx, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    rows: List[Dict] = []
    stack = [""] * len(HIERARCHY_COLS)

    for r in range(5, ws.max_row + 1):
        levels = [nz(ws.cell(r, c).value) for c in HIERARCHY_COLS]
        attr_name = nz(ws.cell(r, ATTR_COL).value)

        deepest_level = -1
        deepest_value = ""
        for idx, val in enumerate(levels):
            if val != "":
                deepest_level = idx
                deepest_value = val

        if deepest_level >= 0:
            stack[deepest_level] = deepest_value
            for j in range(deepest_level + 1, len(stack)):
                stack[j] = ""

        field_kind = ""
        standard_tag = ""
        standard_path = ""

        if attr_name != "":
            field_kind = "attribute"
            standard_tag = local_name(attr_name)

            parent_parts = [local_name(x) for x in stack if x]
            if parent_parts:
                standard_path = "/" + "/".join(parent_parts) + f"/@{standard_tag}"
            else:
                standard_path = f"/@{standard_tag}"

        elif deepest_level >= 0 and deepest_value != "":
            field_kind = "element"
            standard_tag = local_name(deepest_value)

            parts = [local_name(x) for x in stack if x]
            if parts:
                standard_path = "/" + "/".join(parts)
            else:
                standard_path = f"/{standard_tag}"
        else:
            continue

        japanese_name = nz(ws.cell(r, JP_NAME_COL).value)
        schema_no = nz(ws.cell(r, SCHEMA_NO_COL).value)
        schema_name = nz(ws.cell(r, SCHEMA_NAME_COL).value)
        schema_note = nz(ws.cell(r, NOTE_COL).value)
        required = normalize_required(ws.cell(r, REQUIRED_COL).value)
        repeat_raw = nz(ws.cell(r, REPEAT_COL).value)
        repeat_min, repeat_max = parse_repeat(repeat_raw)
        type_class = nz(ws.cell(r, TYPE_CLASS_COL).value)
        data_attr = nz(ws.cell(r, DATA_ATTR_COL).value)
        example_value = nz(ws.cell(r, EXAMPLE_COL).value)
        length = derive_length_from_data_attr(data_attr)

        row = {
            "message_type": "",
            "version": "",
            "field_kind": field_kind,
            "standard_path": standard_path,
            "standard_tag": standard_tag,
            "normalized_tag": normalize_tag(standard_tag),
            "japanese_name": japanese_name,
            "required": required,
            "repeat_min": repeat_min,
            "repeat_max": repeat_max,
            "repeat_raw": repeat_raw,
            "type_class": type_class,
            "data_attribute": data_attr,
            "length": length,
            "example_value": example_value,
            "schema_number": schema_no,
            "schema_name": schema_name,
            "schema_note": schema_note,
            "dictionary_source": os.path.basename(schema_xlsx),
        }
        rows.append(row)

    return rows


# =========================================================
# 既存辞書読込
# =========================================================

SYSTEM_COLUMNS = [
    "message_type",
    "version",
    "field_kind",
    "standard_path",
    "standard_tag",
    "normalized_tag",
    "japanese_name",
    "required",
    "repeat_min",
    "repeat_max",
    "repeat_raw",
    "type_class",
    "data_attribute",
    "length",
    "example_value",
    "schema_number",
    "schema_name",
    "schema_note",
    "dictionary_source",
    "customer_tag_candidates",
    "customer_sample_values",
    "customer_occurs",
    "customer_path_examples",
    "customer_match_status",
    "active",
    "note",
]

PRESERVE_IF_GENERATED_EMPTY = {
    "japanese_name",
    "required",
    "repeat_min",
    "repeat_max",
    "repeat_raw",
    "type_class",
    "data_attribute",
    "length",
    "example_value",
    "schema_number",
    "schema_name",
    "schema_note",
    "note",
    "active",
}


def dict_key(row: Dict) -> Tuple[str, str, str, str, str]:
    return (
        nz(row.get("message_type")),
        nz(row.get("version")),
        nz(row.get("field_kind")),
        nz(row.get("standard_path")),
        nz(row.get("standard_tag")),
    )


def read_existing_dictionary(path: str) -> Tuple[List[Dict], List[str]]:
    if not os.path.exists(path):
        return [], []

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [nz(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    rows = []

    for r in range(2, ws.max_row + 1):
        row = {}
        empty = True
        for c, h in enumerate(headers, start=1):
            v = ws.cell(r, c).value
            row[h] = "" if v is None else v
            if v not in (None, ""):
                empty = False
        if not empty:
            rows.append(row)

    return rows, headers


# =========================================================
# 辞書生成 / マージ
# =========================================================

def enrich_schema_rows_with_customer_info(
    schema_rows: List[Dict],
    customer_info: Dict[str, Dict],
    message_type: str,
    version: str
) -> List[Dict]:
    out = []

    for row in schema_rows:
        row = dict(row)
        row["message_type"] = message_type
        row["version"] = version

        info = customer_info.get(row["normalized_tag"], None)
        if info:
            row["customer_tag_candidates"] = " | ".join(info["customer_tags"])
            row["customer_sample_values"] = " | ".join(info["sample_values"])
            row["customer_occurs"] = str(info["occurs"])
            row["customer_path_examples"] = " | ".join(info["path_examples"])

            if row["standard_tag"] in info["customer_tags"]:
                row["customer_match_status"] = "顧客実績あり(一致)"
            else:
                row["customer_match_status"] = "顧客実績あり(表記ゆれ候補)"
        else:
            row["customer_tag_candidates"] = ""
            row["customer_sample_values"] = ""
            row["customer_occurs"] = ""
            row["customer_path_examples"] = ""
            row["customer_match_status"] = ""

        row["active"] = "○"
        row["note"] = ""

        out.append(row)

    return out


def merge_with_existing(new_rows: List[Dict], existing_rows: List[Dict], existing_headers: List[str]) -> Tuple[List[Dict], List[str]]:
    existing_map = {dict_key(r): r for r in existing_rows}

    merged_rows = []
    extra_headers = [h for h in existing_headers if h not in SYSTEM_COLUMNS]

    for new_row in new_rows:
        key = dict_key(new_row)
        if key not in existing_map:
            merged_row = dict(new_row)
            for h in extra_headers:
                merged_row[h] = ""
            merged_rows.append(merged_row)
            continue

        old_row = existing_map[key]
        merged_row = {}

        for col in SYSTEM_COLUMNS:
            new_val = nz(new_row.get(col, ""))
            old_val = nz(old_row.get(col, ""))

            if col in PRESERVE_IF_GENERATED_EMPTY and new_val == "" and old_val != "":
                merged_row[col] = old_val
            else:
                merged_row[col] = new_val if new_val != "" else old_val

        for h in extra_headers:
            merged_row[h] = old_row.get(h, "")

        merged_rows.append(merged_row)

    new_keys = {dict_key(r) for r in new_rows}
    for old_row in existing_rows:
        key = dict_key(old_row)
        if key not in new_keys:
            merged_rows.append(dict(old_row))

    final_headers = SYSTEM_COLUMNS + extra_headers
    return merged_rows, final_headers


# =========================================================
# 充足判定
# =========================================================

def build_expected_keys(rows: List[Dict], message_type: str, version: str) -> set:
    keys = set()
    for row in rows:
        keys.add((
            nz(message_type),
            nz(version),
            nz(row.get("field_kind")),
            nz(row.get("standard_path")),
            nz(row.get("standard_tag")),
        ))
    return keys


def build_existing_keys(rows: List[Dict]) -> set:
    return {dict_key(r) for r in rows}


def validate_existing_dictionary(
    existing_rows: List[Dict],
    existing_headers: List[str],
    expected_keys: set
) -> Tuple[bool, List[str]]:
    issues = []

    missing_headers = [h for h in SYSTEM_COLUMNS if h not in existing_headers]
    if missing_headers:
        issues.append(f"必須列不足: {', '.join(missing_headers)}")

    existing_keys = build_existing_keys(existing_rows)
    missing_keys = expected_keys - existing_keys
    if missing_keys:
        issues.append(f"標準項目不足: {len(missing_keys)}件")

    is_complete = len(issues) == 0
    return is_complete, issues


# =========================================================
# Excel出力
# =========================================================

OUTPUT_HEADERS = SYSTEM_COLUMNS


def write_dictionary_xlsx(output_xlsx: str, rows: List[Dict], headers: Optional[List[str]] = None) -> None:
    headers = headers or OUTPUT_HEADERS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "dictionary"

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, h in enumerate(headers, start=1):
            ws.cell(r_idx, c_idx, row.get(h, ""))

    widths = {
        "A": 14,
        "B": 10,
        "C": 12,
        "D": 42,
        "E": 24,
        "F": 22,
        "G": 28,
        "H": 10,
        "I": 10,
        "J": 10,
        "K": 12,
        "L": 24,
        "M": 20,
        "N": 10,
        "O": 20,
        "P": 10,
        "Q": 20,
        "R": 30,
        "S": 24,
        "T": 28,
        "U": 24,
        "V": 10,
        "W": 36,
        "X": 20,
        "Y": 8,
        "Z": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    ensure_dir(output_xlsx)
    wb.save(output_xlsx)


# =========================================================
# メイン
# =========================================================

def parse_args():
    parser = argparse.ArgumentParser(description="Step0: Order / v1_3 項目辞書を生成・不足補完する")
    parser.add_argument("--schema-xlsx", required=True, help="Schema_Order_20090901.xlsx")
    parser.add_argument("--customer-xml", required=True, help="顧客データxml")
    parser.add_argument("--output-xlsx", required=True, help="出力する項目辞書xlsx")
    parser.add_argument("--message-type", default="order", help="メッセージ種別。既定値: order")
    parser.add_argument("--version", default="v1_3", help="バージョン。既定値: v1_3")
    parser.add_argument("--sheet-name", default=None, help="Schemaシート名")
    return parser.parse_args()


def main():
    args = parse_args()

    if not os.path.exists(args.schema_xlsx):
        raise FileNotFoundError(f"Schema xlsx が見つかりません: {args.schema_xlsx}")
    if not os.path.exists(args.customer_xml):
        raise FileNotFoundError(f"顧客xml が見つかりません: {args.customer_xml}")

    log("[INFO] 顧客XML読み込み開始")
    customer_tree = ET.parse(args.customer_xml)
    customer_root = customer_tree.getroot()
    customer_rows = flatten_customer_leaf_nodes(customer_root)
    customer_info = aggregate_customer_info(customer_rows)
    log(f"[INFO] 顧客XML葉項目数={len(customer_rows)}")
    log(f"[INFO] 顧客XML正規化タグ数={len(customer_info)}")

    log("[INFO] Schema読み込み開始")
    schema_rows = read_schema_rows(args.schema_xlsx, sheet_name=args.sheet_name)
    log(f"[INFO] Schema辞書行数={len(schema_rows)}")

    log("[INFO] 顧客実績付加開始")
    new_rows = enrich_schema_rows_with_customer_info(
        schema_rows=schema_rows,
        customer_info=customer_info,
        message_type=args.message_type,
        version=args.version,
    )
    log(f"[INFO] 新規生成候補行数={len(new_rows)}")

    expected_keys = build_expected_keys(
        rows=schema_rows,
        message_type=args.message_type,
        version=args.version,
    )

    existing_rows, existing_headers = read_existing_dictionary(args.output_xlsx)

    if existing_rows:
        log(f"[INFO] 既存辞書行数={len(existing_rows)}")
        is_complete, issues = validate_existing_dictionary(
            existing_rows=existing_rows,
            existing_headers=existing_headers,
            expected_keys=expected_keys,
        )

        if is_complete:
            log("[INFO] 既存辞書は必要列・必要行を満たしています。Step0は何もせず終了します")
            return

        log("[INFO] 既存辞書に不足があります。不足補完を実施します")
        for issue in issues:
            log(f"[INFO] - {issue}")

        merged_rows, headers = merge_with_existing(new_rows, existing_rows, existing_headers)
        log(f"[INFO] 補完後辞書行数={len(merged_rows)}")

        write_dictionary_xlsx(args.output_xlsx, merged_rows, headers=headers)
        log(f"[INFO] 出力完了(不足補完): {args.output_xlsx}")
        return

    log("[INFO] 既存辞書なし。新規作成します")
    headers = OUTPUT_HEADERS
    write_dictionary_xlsx(args.output_xlsx, new_rows, headers=headers)
    log(f"[INFO] 出力完了(新規作成): {args.output_xlsx}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log(f"[ERROR] {e}")
        sys.exit(1)