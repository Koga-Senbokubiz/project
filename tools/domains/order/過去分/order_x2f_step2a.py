# -*- coding: utf-8 -*-
"""
order_x2f_step2a.py

Step2a:
顧客データXML と 標準BMSレイアウトXML を比較し、
order_x2f_step2-n.xlsx の初期版を新規作成する。

正式列定義:
A  No
B  変換対象
C  顧客項目
D  標準項目
E  顧客タグ
F  標準タグ
G  正規化タグ
H  XMLパス
I  XMLタグ
J  サンプル値
K  出現回数
L  BMS項目名
M  BMSタグ
N  BMSパス
O  必須
P  桁数
Q  データ型
R  状態
S  突合方法
T  備考

方針:
- 顧客XMLは葉ノード中心に抽出
- 標準XMLは「レイアウト定義XML」として読み、
  Name 属性を優先して標準タグとして扱う
- タグ完全一致 → 正規化タグ一致 → ゆるい候補一致 の順で判定
- マッチしたものは標準の属性（必須・桁数・データ型）を埋める
"""

import os
import re
import sys
import argparse
from collections import defaultdict
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET


def log(msg: str) -> None:
    print(msg, flush=True)


def nz(value, default: str = "") -> str:
    if value is None:
        return default
    return str(value).strip()


def ensure_dir(path: str) -> None:
    d = os.path.dirname(path)
    if d:
        os.makedirs(d, exist_ok=True)


def local_name(tag: str) -> str:
    s = nz(tag)
    if "}" in s:
        s = s.split("}", 1)[1]
    if ":" in s:
        s = s.split(":", 1)[1]
    return s


def normalize_tag(tag: str) -> str:
    s = local_name(tag)
    s = s.replace("_", "")
    s = s.lower()
    return s


def safe_text(elem: ET.Element) -> str:
    if elem.text is None:
        return ""
    return elem.text.strip()


def detect_data_type_from_text(value: str) -> str:
    s = nz(value)
    if s == "":
        return ""
    if s.isdigit():
        return "numeric"
    try:
        float(s)
        return "decimal"
    except Exception:
        pass
    return "string"


def build_path(parent_path: str, elem: ET.Element) -> str:
    name = local_name(elem.tag)
    if parent_path == "":
        return f"/{name}"
    return f"{parent_path}/{name}"


def flatten_xml_leaf_nodes(
    elem: ET.Element,
    parent_path: str = "",
    rows: Optional[List[Dict]] = None,
) -> List[Dict]:
    if rows is None:
        rows = []

    path = build_path(parent_path, elem)
    children = list(elem)

    if children:
        for child in children:
            flatten_xml_leaf_nodes(child, path, rows)
    else:
        raw_tag = local_name(elem.tag)
        rows.append({
            "xml_path": path,
            "xml_tag": raw_tag,
            "normalized_tag": normalize_tag(raw_tag),
            "sample_value": safe_text(elem),
        })

    return rows


def build_layout_name_path(rows: List[Dict]) -> List[Dict]:
    """
    標準XML(レイアウト定義XML)の ID / ParentID / Name から BMSパスを復元
    例:
      /order/tradeSummary/itemNetPrice
    のような Name ベースの論理パスを作る
    """
    by_id = {}
    children_map = defaultdict(list)

    for row in rows:
        row_id = row["id"]
        parent_id = row["parent_id"]
        if row_id:
            by_id[row_id] = row
        if parent_id:
            children_map[parent_id].append(row)

    def resolve_path(row: Dict) -> str:
        cached = row.get("_resolved_path")
        if cached:
            return cached

        name = nz(row["bms_tag"])
        if name == "":
            row["_resolved_path"] = ""
            return ""

        parent_id = nz(row["parent_id"])
        if parent_id == "" or parent_id not in by_id:
            path = f"/{name}"
            row["_resolved_path"] = path
            return path

        parent = by_id[parent_id]
        parent_path = resolve_path(parent)
        if parent_path == "":
            path = f"/{name}"
        else:
            path = f"{parent_path}/{name}"

        row["_resolved_path"] = path
        return path

    for row in rows:
        row["bms_path"] = resolve_path(row)

    return rows


def flatten_standard_layout_xml(
    elem: ET.Element,
    rows: Optional[List[Dict]] = None,
) -> List[Dict]:
    """
    標準XML用:
    レイアウト定義XMLとして読む
    Name 属性があるものを標準項目候補として収集する
    """
    if rows is None:
        rows = []

    elem_name = local_name(elem.tag)

    name_attr = (
        elem.attrib.get("Name", "")
        or elem.attrib.get("name", "")
        or elem.attrib.get("LABEL", "")
        or elem.attrib.get("label", "")
    )

    # レイアウト定義で項目名を持つものだけ対象
    if name_attr != "":
        row = {
            "id": nz(elem.attrib.get("ID", "") or elem.attrib.get("Id", "") or elem.attrib.get("id", "")),
            "parent_id": nz(elem.attrib.get("ParentID", "") or elem.attrib.get("ParentId", "") or elem.attrib.get("parentId", "") or elem.attrib.get("parent_id", "")),
            "bms_tag": name_attr,
            "normalized_tag": normalize_tag(name_attr),
            "bms_item_name": name_attr,
            "required": nz(elem.attrib.get("Mandatory", "") or elem.attrib.get("mandatory", "") or elem.attrib.get("Required", "") or elem.attrib.get("required", "")),
            "length": nz(elem.attrib.get("Length", "") or elem.attrib.get("length", "") or elem.attrib.get("Size", "") or elem.attrib.get("size", "") or elem.attrib.get("Bytes", "") or elem.attrib.get("bytes", "")),
            "data_type": nz(elem.attrib.get("Type", "") or elem.attrib.get("type", "") or elem.attrib.get("DataType", "") or elem.attrib.get("dataType", "") or elem.attrib.get("Format", "") or elem.attrib.get("format", "")),
            "raw_elem_name": elem_name,
            "bms_path": "",  # 後で ID / ParentID から復元
        }
        rows.append(row)

    for child in list(elem):
        flatten_standard_layout_xml(child, rows)

    return rows


def group_customer_rows(customer_rows: List[Dict]) -> List[Dict]:
    grouped: Dict[str, Dict] = {}

    for row in customer_rows:
        path = row["xml_path"]
        if path not in grouped:
            grouped[path] = {
                "xml_path": path,
                "xml_tag": row["xml_tag"],
                "normalized_tag": row["normalized_tag"],
                "sample_value": row["sample_value"],
                "occurs": 1,
            }
        else:
            grouped[path]["occurs"] += 1
            if grouped[path]["sample_value"] == "" and row["sample_value"] != "":
                grouped[path]["sample_value"] = row["sample_value"]

    return list(grouped.values())


def build_standard_indexes(standard_rows: List[Dict]):
    by_tag: Dict[str, List[Dict]] = defaultdict(list)
    by_normalized_tag: Dict[str, List[Dict]] = defaultdict(list)

    for row in standard_rows:
        by_tag[row["bms_tag"]].append(row)
        by_normalized_tag[row["normalized_tag"]].append(row)

    return by_tag, by_normalized_tag


def all_same_non_empty(candidates: List[Dict], key: str) -> str:
    values = [nz(c.get(key, "")) for c in candidates]
    values = [v for v in values if v != ""]
    if not values:
        return ""
    first = values[0]
    if all(v == first for v in values):
        return first
    return ""


def choose_best_candidate(candidates: List[Dict]) -> Optional[Dict]:
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]

    # BMSパスが短いもの優先
    sorted_candidates = sorted(candidates, key=lambda x: len(nz(x["bms_path"])))
    best = sorted_candidates[0]
    best_len = len(nz(best["bms_path"]))
    same_rank = [c for c in sorted_candidates if len(nz(c["bms_path"])) == best_len]

    if len(same_rank) == 1:
        return best

    return None


def loose_match_candidates(customer_tag: str, std_rows: List[Dict]) -> List[Dict]:
    ct = normalize_tag(customer_tag)
    out = []
    for row in std_rows:
        st = row["normalized_tag"]
        if ct == "" or st == "":
            continue
        if ct in st or st in ct:
            out.append(row)
    return out


def build_match_result(
    customer_tag: str,
    normalized_tag: str,
    candidates: List[Dict],
    status: str,
    match_method: str,
    note: str,
) -> Dict:
    if not candidates:
        return {
            "standard_item": "",
            "customer_tag": customer_tag,
            "standard_tag": "",
            "normalized_tag": normalized_tag if normalized_tag else "",
            "bms_item_name": "",
            "bms_tag": "",
            "bms_path": "",
            "required": "",
            "length": "",
            "data_type": "",
            "status": status,
            "match_method": match_method,
            "note": note,
        }

    if len(candidates) == 1:
        std = candidates[0]
        return {
            "standard_item": "○",
            "customer_tag": customer_tag,
            "standard_tag": std["bms_tag"],
            "normalized_tag": std["normalized_tag"],
            "bms_item_name": std["bms_item_name"],
            "bms_tag": std["bms_tag"],
            "bms_path": std["bms_path"],
            "required": std["required"],
            "length": std["length"],
            "data_type": std["data_type"],
            "status": status,
            "match_method": match_method,
            "note": note,
        }

    best = choose_best_candidate(candidates)
    required_consensus = all_same_non_empty(candidates, "required")
    length_consensus = all_same_non_empty(candidates, "length")
    data_type_consensus = all_same_non_empty(candidates, "data_type")
    item_name_consensus = all_same_non_empty(candidates, "bms_item_name")
    bms_tag_consensus = all_same_non_empty(candidates, "bms_tag")
    norm_tag_consensus = all_same_non_empty(candidates, "normalized_tag")

    return {
        "standard_item": "○",
        "customer_tag": customer_tag,
        "standard_tag": bms_tag_consensus or (best["bms_tag"] if best else ""),
        "normalized_tag": norm_tag_consensus or normalized_tag,
        "bms_item_name": item_name_consensus or (best["bms_item_name"] if best else ""),
        "bms_tag": bms_tag_consensus or (best["bms_tag"] if best else ""),
        "bms_path": "",  # 複数候補は手動確定
        "required": required_consensus,
        "length": length_consensus,
        "data_type": data_type_consensus,
        "status": status,
        "match_method": match_method,
        "note": note,
    }


def match_customer_to_standard(
    customer_row: Dict,
    std_by_tag: Dict[str, List[Dict]],
    std_by_normalized_tag: Dict[str, List[Dict]],
    standard_rows: List[Dict],
) -> Dict:
    customer_tag = customer_row["xml_tag"]
    normalized_tag = customer_row["normalized_tag"]

    # 1. タグ完全一致
    exact_candidates = std_by_tag.get(customer_tag, [])
    if exact_candidates:
        if len(exact_candidates) == 1:
            return build_match_result(
                customer_tag=customer_tag,
                normalized_tag=normalized_tag,
                candidates=exact_candidates,
                status="一致",
                match_method="タグ完全一致",
                note="",
            )
        return build_match_result(
            customer_tag=customer_tag,
            normalized_tag=normalized_tag,
            candidates=exact_candidates,
            status="候補あり",
            match_method="タグ完全一致(複数候補)",
            note=f"複数候補あり:{len(exact_candidates)} / BMSパスは手動確定",
        )

    # 2. 正規化タグ一致
    norm_candidates = std_by_normalized_tag.get(normalized_tag, [])
    if norm_candidates:
        if len(norm_candidates) == 1:
            return build_match_result(
                customer_tag=customer_tag,
                normalized_tag=normalized_tag,
                candidates=norm_candidates,
                status="候補あり",
                match_method="正規化タグ一致",
                note="表記ゆれ一致。妥当性確認要",
            )
        return build_match_result(
            customer_tag=customer_tag,
            normalized_tag=normalized_tag,
            candidates=norm_candidates,
            status="候補あり",
            match_method="正規化タグ一致(複数候補)",
            note=f"複数候補あり:{len(norm_candidates)} / 妥当性確認要",
        )

    # 3. ゆるい候補一致
    loose_candidates = loose_match_candidates(customer_tag, standard_rows)
    if loose_candidates:
        return build_match_result(
            customer_tag=customer_tag,
            normalized_tag=normalized_tag,
            candidates=loose_candidates,
            status="候補あり",
            match_method="ゆるい候補一致",
            note=f"部分一致候補:{len(loose_candidates)} / 自動確定禁止",
        )

    # 4. 未一致
    return {
        "standard_item": "",
        "customer_tag": customer_tag,
        "standard_tag": "",
        "normalized_tag": normalized_tag,
        "bms_item_name": "",
        "bms_tag": "",
        "bms_path": "",
        "required": "",
        "length": "",
        "data_type": "",
        "status": "未一致",
        "match_method": "",
        "note": "顧客独自項目候補",
    }


def make_step2a_rows(customer_grouped: List[Dict], standard_rows: List[Dict]) -> List[Dict]:
    std_by_tag, std_by_normalized_tag = build_standard_indexes(standard_rows)

    out_rows: List[Dict] = []
    no = 1

    for c in customer_grouped:
        matched = match_customer_to_standard(
            customer_row=c,
            std_by_tag=std_by_tag,
            std_by_normalized_tag=std_by_normalized_tag,
            standard_rows=standard_rows,
        )

        row = {
            "no": no,
            "convert_target": "",
            "customer_item": "○",
            "standard_item": matched["standard_item"],
            "customer_tag": matched["customer_tag"],
            "standard_tag": matched["standard_tag"],
            "normalized_tag": matched["normalized_tag"],
            "xml_path": c["xml_path"],
            "xml_tag": c["xml_tag"],
            "sample_value": c["sample_value"],
            "occurs": c["occurs"],
            "bms_item_name": matched["bms_item_name"],
            "bms_tag": matched["bms_tag"],
            "bms_path": matched["bms_path"],
            "required": matched["required"],
            "length": matched["length"],
            "data_type": matched["data_type"],
            "status": matched["status"],
            "match_method": matched["match_method"],
            "note": matched["note"],
        }

        if row["standard_item"] == "":
            inferred_type = detect_data_type_from_text(row["sample_value"])
            if inferred_type:
                row["note"] = (row["note"] + f" / 推定型:{inferred_type}").strip(" /")

        out_rows.append(row)
        no += 1

    return out_rows


HEADERS = [
    "No", "変換対象", "顧客項目", "標準項目",
    "顧客タグ", "標準タグ", "正規化タグ",
    "XMLパス", "XMLタグ", "サンプル値", "出現回数",
    "BMS項目名", "BMSタグ", "BMSパス", "必須", "桁数", "データ型",
    "状態", "突合方法", "備考"
]


def apply_sheet_style(ws) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="CCCCCC")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    widths = {
        "A": 6,
        "B": 10,
        "C": 10,
        "D": 10,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 40,
        "I": 18,
        "J": 22,
        "K": 10,
        "L": 24,
        "M": 18,
        "N": 40,
        "O": 10,
        "P": 10,
        "Q": 14,
        "R": 12,
        "S": 24,
        "T": 30,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    max_row = ws.max_row
    max_col = ws.max_column
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    for r in range(2, max_row + 1):
        status = nz(ws.cell(r, 18).value)  # R列
        if status == "一致":
            fill = PatternFill(fill_type="solid", fgColor="EAF4EA")
        elif status == "候補あり":
            fill = PatternFill(fill_type="solid", fgColor="FFF7E6")
        elif status == "未一致":
            fill = PatternFill(fill_type="solid", fgColor="FDECEC")
        else:
            fill = None

        if fill:
            for c in range(1, max_col + 1):
                ws.cell(r, c).fill = fill


def write_step2a_xlsx(output_xlsx: str, rows: List[Dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Step2a"

    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for idx, row in enumerate(rows, start=2):
        values = [
            row["no"],
            row["convert_target"],
            row["customer_item"],
            row["standard_item"],
            row["customer_tag"],
            row["standard_tag"],
            row["normalized_tag"],
            row["xml_path"],
            row["xml_tag"],
            row["sample_value"],
            row["occurs"],
            row["bms_item_name"],
            row["bms_tag"],
            row["bms_path"],
            row["required"],
            row["length"],
            row["data_type"],
            row["status"],
            row["match_method"],
            row["note"],
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=idx, column=col_idx, value=value)

    apply_sheet_style(ws)
    ensure_dir(output_xlsx)
    wb.save(output_xlsx)


def parse_args():
    parser = argparse.ArgumentParser(description="Step2a: 顧客XMLと標準XMLを比較してStep2初期xlsxを生成する")
    parser.add_argument("--input-xml", required=True, help="顧客データXML")
    parser.add_argument("--standard-xml", required=True, help="標準BMSレイアウトXML")
    parser.add_argument("--output-xlsx", required=True, help="出力するStep2a.xlsx")
    return parser.parse_args()


def main():
    args = parse_args()

    if not os.path.exists(args.input_xml):
        raise FileNotFoundError(f"顧客データXMLが見つかりません: {args.input_xml}")
    if not os.path.exists(args.standard_xml):
        raise FileNotFoundError(f"標準XMLが見つかりません: {args.standard_xml}")

    log("[INFO] 顧客XML読み込み開始")
    customer_tree = ET.parse(args.input_xml)
    customer_root = customer_tree.getroot()

    customer_leaf_rows = flatten_xml_leaf_nodes(customer_root)
    customer_grouped = group_customer_rows(customer_leaf_rows)
    log(f"[INFO] 顧客XML項目数={len(customer_grouped)}")

    log("[INFO] 標準XML読み込み開始")
    standard_tree = ET.parse(args.standard_xml)
    standard_root = standard_tree.getroot()

    standard_rows = flatten_standard_layout_xml(standard_root)
    standard_rows = build_layout_name_path(standard_rows)
    log(f"[INFO] 標準XML項目数={len(standard_rows)}")

    log("[INFO] Step2a行生成開始")
    step2a_rows = make_step2a_rows(customer_grouped, standard_rows)
    log(f"[INFO] Step2a行数={len(step2a_rows)}")

    exact_count = sum(1 for r in step2a_rows if r["status"] == "一致")
    candidate_count = sum(1 for r in step2a_rows if r["status"] == "候補あり")
    unmatched_count = sum(1 for r in step2a_rows if r["status"] == "未一致")

    log(f"[INFO] 一致件数={exact_count}")
    log(f"[INFO] 候補あり件数={candidate_count}")
    log(f"[INFO] 未一致件数={unmatched_count}")

    log("[INFO] Step2a.xlsx 出力開始")
    write_step2a_xlsx(args.output_xlsx, step2a_rows)
    log(f"[INFO] 出力完了: {args.output_xlsx}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log(f"[ERROR] {e}")
        sys.exit(1)