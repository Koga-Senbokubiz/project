#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import xml.etree.ElementTree as ET
from pathlib import Path
import openpyxl


def local(tag):
    return tag.split("}")[-1]


def normalize_tag(tag):
    s = "" if tag is None else str(tag).strip()
    if ":" in s:
        s = s.split(":")[-1]
    return s


# =========================
# 基本形XML
# =========================
def extract_base_leaves(xml_path):
    tree = ET.parse(xml_path)
    root = tree.getroot()

    layout_map = {}
    item_map = {}

    for elem in root.findall("項目"):
        item_id = str(elem.attrib.get("ID", "")).strip()
        if not item_id:
            continue
        item_map[item_id] = {
            "jp_name": elem.attrib.get("項目ID", "")
        }

    for elem in root.findall("レイアウト"):
        layout_id = str(elem.attrib.get("ID", "")).strip()
        if not layout_id:
            continue
        layout_map[layout_id] = {
            "name": elem.attrib.get("Name", ""),
            "type": str(elem.attrib.get("Type", "")).strip(),
            "property_id": str(elem.attrib.get("PropertyID", "")).strip(),
            "parents_id": str(elem.attrib.get("ParentsID", "")).strip(),
        }

    def build_xpath(layout_id):
        names = []
        current_id = layout_id

        while current_id and current_id in layout_map:
            node = layout_map[current_id]
            if node["type"] != "1" and node["name"]:
                names.append(normalize_tag(node["name"]))

            parent_id = node["parents_id"]
            if not parent_id or parent_id == "0":
                break
            current_id = parent_id

        names.reverse()
        return "/" + "/".join(names)

    leaves = []
    for layout_id, node in layout_map.items():
        if node["type"] != "5":
            continue

        tag_name = normalize_tag(node["name"])
        if not tag_name:
            continue

        xpath = build_xpath(layout_id)
        jp_name = item_map.get(node["property_id"], {}).get("jp_name", "")

        leaves.append({
            "xpath": xpath,
            "tag": tag_name,
            "jp_name": jp_name
        })

    return leaves


# =========================
# 顧客XML
# =========================
def extract_customer_tags(xml_path):
    tree = ET.parse(xml_path)
    root = tree.getroot()

    tags = set()
    leaves = []

    def walk(elem, path):
        name = local(elem.tag)
        tags.add(name)

        current_path = path + [normalize_tag(name)]
        children = [c for c in list(elem) if isinstance(c.tag, str)]

        if not children:
            leaves.append({
                "xpath": "/" + "/".join(current_path),
                "tag": normalize_tag(name)
            })
            return

        for c in children:
            if isinstance(c.tag, str):
                walk(c, current_path)

    walk(root, [])
    return tags, leaves


# =========================
# 辞書
# =========================
def load_dictionary(dictionary_path):
    wb = openpyxl.load_workbook(dictionary_path, data_only=True)
    ws = wb["customer_dictionary"]

    header = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str):
            header[v.strip()] = c

    col_status = header.get("status")
    col_candidate = header.get("candidate_bms_tag")

    adopted = set()

    for r in range(3, ws.max_row + 1):
        status = str(ws.cell(r, col_status).value or "").strip()
        candidate = str(ws.cell(r, col_candidate).value or "").strip()

        if status == "標準項目" and candidate:
            adopted.add(normalize_tag(candidate))

    return adopted


# =========================
# main
# =========================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--base-xml", required=True)
    parser.add_argument("--dictionary", required=True)
    parser.add_argument("--customer-xml", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    base_leaves = extract_base_leaves(args.base_xml)
    dict_tags = load_dictionary(args.dictionary)
    customer_tags_raw, customer_leaf_rows = extract_customer_tags(args.customer_xml)
    customer_tags = {normalize_tag(x) for x in customer_tags_raw}

    wb = openpyxl.Workbook()
    ws = wb.active

    # ヘッダ
    ws.append(["No", "基本形1_3：発注XPath", "タグ", "対象外", "顧客XML", "顧客辞書", "備考"])

    # 列幅
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 56
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 9
    ws.column_dimensions["F"].width = 9
    ws.column_dimensions["G"].width = 22

    no = 1
    base_tag_set = set()

    for leaf in base_leaves:
        tag = normalize_tag(leaf["tag"])
        base_tag_set.add(tag)

        in_dict = tag in dict_tags
        in_xml = tag in customer_tags

        delete_flag = "○" if not (in_dict or in_xml) else ""
        xml_flag = "○" if in_xml else "×"
        dict_flag = "○" if in_dict else "×"

        memo = f"{leaf['jp_name']}" if leaf.get("jp_name") else ""

        ws.append([
            no,
            leaf["xpath"],
            leaf["tag"],
            delete_flag,
            xml_flag,
            dict_flag,
            memo
        ])
        no += 1

    # 末尾に顧客XMLのみに存在するタグを追記
    customer_only_rows = []
    seen_customer_only = set()

    for row in customer_leaf_rows:
        tag = normalize_tag(row["tag"])
        xpath = row["xpath"]

        if tag in base_tag_set:
            continue

        key = (xpath, tag)
        if key in seen_customer_only:
            continue
        seen_customer_only.add(key)

        customer_only_rows.append({
            "xpath": xpath,
            "tag": tag
        })

    if customer_only_rows:
        ws.append(["", "", "", "", "", "", ""])
        ws.append(["", "顧客データxmlのみに存在するタグ", "", "", "", "", ""])
        ws.append(["No", "顧客XML XPath", "タグ", "", "", "", "備考"])

        for row in customer_only_rows:
            ws.append([
                no,
                row["xpath"],
                row["tag"],
                "",
                "○",
                "×",
                "基本形1_3：発注Ver1_3.xml になし"
            ])
            no += 1

    wb.save(args.out)

    print("[OK] Step2 completed")


if __name__ == "__main__":
    main()