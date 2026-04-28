# -*- coding: utf-8 -*-
"""
order_x2f_step3.py
Step3: Step2のxlsxをもとに、変換元レイアウトXMLを生成する。

前提:
- .bat から TEMPLATE_FILE を直接受け取る
- テンプレートXMLの構造は維持し、H/D/T レコード配下の項目だけ差し替える
- MappingEditor で読めることを優先する
"""

import os
import sys
import copy
import argparse
from typing import Dict, List, Optional, Tuple

import openpyxl
import xml.etree.ElementTree as ET


# =========================================================
# 共通
# =========================================================

def log(msg: str) -> None:
    print(msg, flush=True)


def nz(value, default="") -> str:
    if value is None:
        return default
    return str(value).strip()


def safe_int(value, default=0) -> int:
    try:
        s = nz(value)
        if s == "":
            return default
        return int(float(s))
    except Exception:
        return default


def ensure_dir(path: str) -> None:
    d = os.path.dirname(path)
    if d:
        os.makedirs(d, exist_ok=True)


def local_name(tag: str) -> str:
    if "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def ns_prefix(tag: str) -> str:
    if tag.startswith("{") and "}" in tag:
        return tag.split("}", 1)[0] + "}"
    return ""


# =========================================================
# Step2 xlsx 読み込み
# =========================================================

COLUMN_ALIASES = {
    "record_type": ["H/D/T", "区分", "record_type", "record", "rec_type", "レコード区分"],
    "item_name": ["項目名", "item_name", "name", "論理名", "項目名称"],
    "item_id": ["項目ID", "item_id", "id", "物理名", "タグ名", "element_name"],
    "data_type": ["型", "type", "data_type", "項目型", "データ型"],
    "length": ["桁数", "length", "size", "bytes", "項目長"],
    "required": ["必須", "required", "mandatory"],
    "repeat": ["繰返", "repeat", "occurs", "repeat_count"],
    "adopt": ["採用", "マッピング採用", "use", "adopt"],
    "note": ["備考", "note", "memo"],
}


def normalize_header(s: str) -> str:
    return nz(s).lower().replace(" ", "").replace("　", "")


def find_header_map(ws) -> Dict[str, int]:
    """
    先頭10行くらいを見て、ヘッダ行を自動検出
    """
    best_row = None
    best_score = -1
    best_map = {}

    for row_idx in range(1, min(ws.max_row, 10) + 1):
        values = [normalize_header(ws.cell(row=row_idx, column=c).value) for c in range(1, ws.max_column + 1)]

        tmp_map = {}
        score = 0
        for key, aliases in COLUMN_ALIASES.items():
            alias_norms = [normalize_header(x) for x in aliases]
            for col_idx, v in enumerate(values, start=1):
                if v in alias_norms:
                    tmp_map[key] = col_idx
                    score += 1
                    break

        if score > best_score:
            best_score = score
            best_row = row_idx
            best_map = tmp_map

    if best_row is None or best_score <= 0:
        raise RuntimeError("Step2 xlsx のヘッダ行を特定できませんでした。")

    best_map["_header_row"] = best_row
    return best_map


def cell(ws, row_idx: int, header_map: Dict[str, int], key: str):
    col = header_map.get(key)
    if not col:
        return None
    return ws.cell(row=row_idx, column=col).value


def is_adopted(adopt_value) -> bool:
    s = nz(adopt_value)
    if s in ("", "-", "×", "x", "X", "FALSE", "False", "false", "0", "候補あり"):
        return False
    if s in ("○", "〇", "o", "O", "1", "TRUE", "True", "true", "採用", "yes", "Yes", "YES"):
        return True
    # 採用列がある場合、空以外を採用とみなしたいケースもあるので緩めに
    return s != ""


def load_step2_items(step2_xlsx: str, sheet_name: Optional[str] = None) -> Dict[str, List[Dict]]:
    wb = openpyxl.load_workbook(step2_xlsx, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    header_map = find_header_map(ws)
    header_row = header_map["_header_row"]

    records = {"H": [], "D": [], "T": []}
    seq_map = {"H": 1, "D": 1, "T": 1}

    for r in range(header_row + 1, ws.max_row + 1):
        rec_type = nz(cell(ws, r, header_map, "record_type")).upper()
        if rec_type not in ("H", "D", "T"):
            continue

        adopt_col_exists = "adopt" in header_map
        adopt_value = cell(ws, r, header_map, "adopt") if adopt_col_exists else "○"
        if adopt_col_exists and not is_adopted(adopt_value):
            continue

        item_name = nz(cell(ws, r, header_map, "item_name"))
        item_id = nz(cell(ws, r, header_map, "item_id"))
        data_type = nz(cell(ws, r, header_map, "data_type"), "X")
        length = safe_int(cell(ws, r, header_map, "length"), 1)
        required = nz(cell(ws, r, header_map, "required"), "false")
        repeat = safe_int(cell(ws, r, header_map, "repeat"), 1)
        note = nz(cell(ws, r, header_map, "note"))

        if item_name == "" and item_id == "":
            continue

        if item_id == "":
            item_id = item_name
        if item_name == "":
            item_name = item_id

        item = {
            "seq": seq_map[rec_type],
            "record_type": rec_type,
            "item_name": item_name,
            "item_id": item_id,
            "data_type": data_type if data_type else "X",
            "length": max(length, 1),
            "required": required if required else "false",
            "repeat": max(repeat, 1),
            "note": note,
        }
        records[rec_type].append(item)
        seq_map[rec_type] += 1

    return records


# =========================================================
# テンプレートXML解析
# =========================================================

def find_candidate_record_nodes(root: ET.Element) -> List[ET.Element]:
    """
    Record / RecordLayout 相当を広めに拾う
    """
    result = []
    for elem in root.iter():
        lname = local_name(elem.tag).lower()
        if lname in ("record", "recordlayout", "layoutrecord", "row", "segment"):
            result.append(elem)
    return result


def node_match_record_type(elem: ET.Element, record_type: str) -> bool:
    """
    H / D / T を、name/id/code/label 等の属性から推定
    """
    candidates = []
    for k, v in elem.attrib.items():
        candidates.append((k.lower(), nz(v).upper()))

    text_val = nz(elem.text).upper()
    if text_val:
        candidates.append(("text", text_val))

    patterns = {
        "H": ["H", "HEADER", "HEAD", "HDR"],
        "D": ["D", "DETAIL", "DTL", "BODY"],
        "T": ["T", "TRAILER", "TAIL", "TRL", "FOOTER"],
    }

    for _, v in candidates:
        for p in patterns[record_type]:
            if v == p or v.endswith("_" + p) or v.startswith(p + "_") or p in v:
                return True
    return False


def find_record_node(root: ET.Element, record_type: str) -> Optional[ET.Element]:
    for elem in find_candidate_record_nodes(root):
        if node_match_record_type(elem, record_type):
            return elem
    return None


def find_first_field_prototype(record_node: ET.Element) -> Optional[ET.Element]:
    """
    テンプレート中の既存項目ノードを雛形にする
    """
    for child in list(record_node):
        lname = local_name(child.tag).lower()
        if lname in ("field", "item", "column", "layoutitem", "element"):
            return child
    return None


def make_fallback_field(record_node: ET.Element) -> ET.Element:
    """
    テンプレートに項目雛形が無い場合の最終フォールバック
    """
    ns = ns_prefix(record_node.tag)
    return ET.Element(f"{ns}Field")


def clear_field_children(record_node: ET.Element, field_prototype: ET.Element) -> None:
    field_local = local_name(field_prototype.tag)
    remove_targets = []

    for child in list(record_node):
        if local_name(child.tag) == field_local:
            remove_targets.append(child)

    for child in remove_targets:
        record_node.remove(child)


def set_attr_if_exists_or_common(elem: ET.Element, candidates: List[str], value: str) -> bool:
    """
    既存属性名に合わせてセット。既存候補がなければ先頭名で新規作成。
    """
    lower_map = {k.lower(): k for k in elem.attrib.keys()}

    for cand in candidates:
        if cand.lower() in lower_map:
            elem.set(lower_map[cand.lower()], value)
            return True

    elem.set(candidates[0], value)
    return True


def apply_item_to_field(field_elem: ET.Element, item: Dict) -> None:
    """
    雛形項目ノードに対して、一般的な属性をできるだけ自然に流し込む
    """
    set_attr_if_exists_or_common(field_elem, ["id", "itemId", "fieldId", "code", "seq"], str(item["seq"]))
    set_attr_if_exists_or_common(field_elem, ["name", "itemName", "fieldName", "label"], item["item_name"])
    set_attr_if_exists_or_common(field_elem, ["physicalName", "item_id", "tag", "xmlName"], item["item_id"])
    set_attr_if_exists_or_common(field_elem, ["type", "dataType", "format"], item["data_type"])
    set_attr_if_exists_or_common(field_elem, ["length", "size", "bytes", "maxLength"], str(item["length"]))
    set_attr_if_exists_or_common(field_elem, ["required", "mandatory"], str(item["required"]).lower())
    set_attr_if_exists_or_common(field_elem, ["repeat", "occurs", "repeatCount", "maxOccurs"], str(item["repeat"]))

    # 雛形に text があって空なら項目名を入れておく
    if nz(field_elem.text) == "":
        field_elem.text = item["item_name"]


def indent_xml(elem: ET.Element, level: int = 0) -> None:
    i = "\n" + level * "  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "  "
        for child in elem:
            indent_xml(child, level + 1)
        if not child.tail or not child.tail.strip():
            child.tail = i
    if level and (not elem.tail or not elem.tail.strip()):
        elem.tail = i


# =========================================================
# メイン処理
# =========================================================

def build_layout_xml(step2_xlsx: str, template_file: str, output_xml: str, sheet_name: Optional[str] = None) -> None:
    if not os.path.exists(step2_xlsx):
        raise FileNotFoundError(f"Step2 xlsx が見つかりません: {step2_xlsx}")
    if not os.path.exists(template_file):
        raise FileNotFoundError(f"テンプレートXMLが見つかりません: {template_file}")

    log("[INFO] Step2 xlsx 読み込み開始")
    records = load_step2_items(step2_xlsx, sheet_name=sheet_name)
    log(f"[INFO] H={len(records['H'])}件 D={len(records['D'])}件 T={len(records['T'])}件")

    log("[INFO] テンプレートXML読み込み開始")
    tree = ET.parse(template_file)
    root = tree.getroot()

    for rec_type in ("H", "D", "T"):
        record_node = find_record_node(root, rec_type)
        if record_node is None:
            log(f"[WARN] {rec_type} レコードがテンプレートXML内で見つかりません。スキップします。")
            continue

        field_prototype = find_first_field_prototype(record_node)
        if field_prototype is None:
            field_prototype = make_fallback_field(record_node)

        clear_field_children(record_node, field_prototype)

        for item in records[rec_type]:
            new_field = copy.deepcopy(field_prototype)
            # 雛形にぶら下がっている不要情報をできるだけクリア
            for sub in list(new_field):
                new_field.remove(sub)
            new_field.text = None

            apply_item_to_field(new_field, item)
            record_node.append(new_field)

        log(f"[INFO] {rec_type} レコードに {len(records[rec_type])} 項目を設定しました。")

    indent_xml(root)
    ensure_dir(output_xml)
    tree.write(output_xml, encoding="utf-8", xml_declaration=True)
    log(f"[INFO] 出力完了: {output_xml}")


def parse_args():
    parser = argparse.ArgumentParser(description="Step3: Step2 xlsx から変換元レイアウトXMLを生成する")
    parser.add_argument("--step2-xlsx", required=True, help="Step2のxlsxファイル")
    parser.add_argument("--template-file", required=True, help="テンプレートXMLファイル")
    parser.add_argument("--output-xml", required=True, help="出力する変換元レイアウトXML")
    parser.add_argument("--sheet-name", required=False, default=None, help="対象シート名")
    return parser.parse_args()


def main():
    args = parse_args()
    build_layout_xml(
        step2_xlsx=args.step2_xlsx,
        template_file=args.template_file,
        output_xml=args.output_xml,
        sheet_name=args.sheet_name,
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log(f"[ERROR] {e}")
        sys.exit(1)