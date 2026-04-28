from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import BinaryIO
import logging
import xml.etree.ElementTree as ET
import openpyxl

CRLF = b"\r\n"
X01 = bytes([0x01])
XFF = bytes([0xFF])
FULLWIDTH_SPACE = "　"


def local_name(tag: str) -> str:
    return tag.rsplit('}', 1)[-1] if '}' in tag else tag


def child_text(elem: ET.Element | None, *names: str, default: str = "") -> str:
    cur = elem
    for name in names:
        if cur is None:
            return default
        nxt = None
        for ch in list(cur):
            if local_name(ch.tag) == name:
                nxt = ch
                break
        cur = nxt
    if cur is None or cur.text is None:
        return default
    return cur.text.strip()


def children(elem: ET.Element | None, name: str) -> list[ET.Element]:
    if elem is None:
        return []
    return [ch for ch in list(elem) if local_name(ch.tag) == name]


def all_fullwidth(text: str) -> bool:
    if not text:
        return False
    b = text.encode("cp932", errors="ignore")
    return len(text) * 2 == len(b)


def cut_to_cp932_bytes(text: str, length: int) -> bytes:
    out = bytearray()
    for ch in text:
        enc = ch.encode("cp932", errors="ignore")
        if len(out) + len(enc) > length:
            break
        out.extend(enc)
    return bytes(out)


def cf_str(val: str, length: int) -> bytes:
    s = (val or "").strip()
    b = s.encode("cp932", errors="ignore")
    if len(b) > length:
        b = b[:length]
    return b + b" " * (length - len(b))


def cf_int(val: str | int | Decimal, length: int) -> bytes:
    s = str(val).strip()
    if s and not s.isdigit():
        return b" " * length
    s = s.zfill(length)
    if len(s) > length:
        s = s[-length:]
    return s.encode("ascii")


def cf_jpn(val: str, length: int) -> bytes:
    s = (val or "").strip()
    if all_fullwidth(s):
        pad_chars = length // 2
        s2 = (s + FULLWIDTH_SPACE * pad_chars)[:pad_chars]
        return s2.encode("cp932", errors="ignore")[:length]
    b = cut_to_cp932_bytes(s, length)
    return b + b" " * (length - len(b))


def cf_jpn2(val: str, length: int) -> bytes:
    s = (val or "").strip()
    if all_fullwidth(s):
        pad_chars = length // 2
        s2 = (s + FULLWIDTH_SPACE * pad_chars)[:pad_chars]
        return s2.encode("cp932", errors="ignore")[:length]
    b = cut_to_cp932_bytes(s, length)
    return b + b" " * (length - len(b))


def cf_jpn_name(val: str, length: int) -> bytes:
    return cf_jpn(val, length)


def cf_jpn_003(val: str, length: int) -> bytes:
    #return cf_jpn2(val, length)
    s = (val or "").strip()
    raw = mb_strcut_cp932(s, length)

    # 全角判定（PHPと同じ）
    if len(s.encode("cp932")) == len(s) * 2:
        # 全角 → 全角空白で埋め
        pad = length - len(raw)
        return raw + ("　".encode("cp932") * (pad // 2)) + (b" " if pad % 2 else b"")
    else:
        # 混在 → 半角空白
        return raw + b" " * (length - len(raw))


def cf_jpn_note(val: str, length: int) -> bytes:
    s = (val or "").strip()
    raw = cut_to_cp932_bytes(s, length)
    pad = length - len(raw)
    if pad <= 0:
        return raw
    return raw + ("　".encode("cp932") * (pad // 2)) + (b" " if pad % 2 else b"")


def cf_jpn_t0(val: str, length: int) -> bytes:
    s = (val or "").strip()
    b = cut_to_cp932_bytes(s, length)
    return b + b" " * (length - len(b))

def mb_strcut_cp932(text: str, length: int) -> bytes:
    """
    PHPの mb_strcut($text, 0, length, 'SJIS') 相当（簡易版）
    - CP932(SJIS-win)前提
    - 文字途中で切れない
    """
    if not text:
        return b""

    encoded = text.encode("cp932", errors="ignore")

    if len(encoded) <= length:
        return encoded

    # いったん切る
    cut = encoded[:length]

    # 末尾が文字途中なら削る
    while True:
        try:
            cut.decode("cp932")
            break
        except UnicodeDecodeError:
            cut = cut[:-1]

    return cut


def cf_ymd(val: str, length: int) -> bytes:
    s = (val or "").strip()
    parts = s.split("-")
    if len(parts) == 3 and all(parts):
        try:
            s = f"{int(parts[0]):04d}{int(parts[1]):02d}{int(parts[2]):02d}"
        except ValueError:
            pass
    b = s.encode("ascii", errors="ignore")
    if len(b) > length:
        b = b[:length]
    return b + b" " * (length - len(b))


@dataclass
class LineItem:
    line_no: str
    gtin: str
    order_item_code: str
    code_type: str
    item_name: str
    spec: str
    unit_price: str
    net_amount: str
    quantity: str
    unit_multiple: str
    num_order_units: str
    unit_of_measure: str


@dataclass
class OrderRecord:
    trade_number: str
    receiver_code: str
    receiver_gln: str
    receiver_name: str
    transfer_code: str
    transfer_gln: str
    payee_code: str
    payee_gln: str
    seller_code: str
    seller_gln: str
    seller_name: str
    buyer_name: str
    major_category: str
    sub_major_category: str
    order_date_yy_mm_dd: str
    delivery_date_yy_mm_dd: str
    trade_type_code: str
    goods_classification_code: str
    note_text: str
    fax_no: str
    items: list[LineItem] = field(default_factory=list)


class BBORDXMLtoDFAX:
    def __init__(
        self,
        out_fp: BinaryIO,
        bb_fax_no: str,
        debug_mode: str,
        debug_fax_no: str,
        now: datetime | None = None,
        def_xlsx: str | Path | None = None,
        logger: logging.Logger | None = None,
    ) -> None:
        self.out_fp = out_fp
        self.bb_fax_no = bb_fax_no
        self.debug_mode = debug_mode
        self.debug_fax_no = debug_fax_no
        self.now = now or datetime.now()
        self.current_fax_no = ""
        self.current_orocd = ""
        self.count_page = 0
        self.count_line = 0
        self.def_xlsx = Path(def_xlsx) if def_xlsx else None
        self.logger = logger or logging.getLogger(__name__)
        self._defs_page_header: list[dict[str, object]] = []
        self._defs_11: list[dict[str, object]] = []
        self._defs_20: list[dict[str, object]] = []
        self._defs_21: list[dict[str, object]] = []
        self._defs_22: list[dict[str, object]] = []
        self._defs_23: list[dict[str, object]] = []
        if self.def_xlsx and self.def_xlsx.exists():
            self._load_excel_defs()
        elif self.def_xlsx:
            self.logger.warning("Excel定義ファイルが見つかりません: %s", self.def_xlsx)

    def _load_excel_defs(self) -> None:
        try:
            self.logger.debug("Excel定義読込開始: %s", self.def_xlsx)
            wb = openpyxl.load_workbook(self.def_xlsx, data_only=True)

            if "10_out04_ページヘッダ" in wb.sheetnames:
                ws = wb["10_out04_ページヘッダ"]
                defs: list[dict[str, object]] = []
                for row in ws.iter_rows(min_row=5, values_only=True):
                    enabled, line_offset, col, source_kind, source_name, formatter, length, fixed_value, notes = row[:9]
                    if str(enabled).strip().upper() != "Y":
                        continue
                    if col in (None, ""):
                        continue
                    defs.append({
                        "line_offset": int(line_offset or 0),
                        "col": str(col).zfill(3),
                        "source_kind": (source_kind or "").strip(),
                        "source_name": (source_name or "").strip(),
                        "formatter": (formatter or "").strip(),
                        "length": None if length in (None, "") else int(length),
                        "fixed_value": "" if fixed_value is None else str(fixed_value),
                        "notes": "" if notes is None else str(notes),
                    })
                self._defs_page_header = defs

            if "11_out04_注文ヘッダ" in wb.sheetnames:
                ws = wb["11_out04_注文ヘッダ"]
                defs11: list[dict[str, object]] = []
                for row in ws.iter_rows(min_row=5, values_only=True):
                    enabled, line_offset, col, source_kind, source_name, formatter, length, fixed_value, notes = row[:9]
                    if str(enabled).strip().upper() != "Y":
                        continue
                    if col in (None, ""):
                        continue
                    defs11.append({
                        "line_offset": int(line_offset or 0),
                        "col": str(col).zfill(3),
                        "source_kind": (source_kind or "").strip(),
                        "source_name": (source_name or "").strip(),
                        "formatter": (formatter or "").strip(),
                        "length": None if length in (None, "") else int(length),
                        "fixed_value": "" if fixed_value is None else str(fixed_value),
                        "notes": "" if notes is None else str(notes),
                    })
                self._defs_11 = defs11

            if "20_out05_明細1行目" in wb.sheetnames:
                ws = wb["20_out05_明細1行目"]
                defs20: list[dict[str, object]] = []
                for row in ws.iter_rows(min_row=5, values_only=True):
                    enabled, line_offset, col, source_kind, source_name, formatter, length, fixed_value, notes = row[:9]
                    if str(enabled).strip().upper() != "Y":
                        continue
                    if col in (None, ""):
                        continue
                    defs20.append({
                        "line_offset": int(line_offset or 0),
                        "col": str(col).zfill(3),
                        "source_kind": (source_kind or "").strip(),
                        "source_name": (source_name or "").strip(),
                        "formatter": (formatter or "").strip(),
                        "length": None if length in (None, "") else int(length),
                        "fixed_value": "" if fixed_value is None else str(fixed_value),
                        "notes": "" if notes is None else str(notes),
                    })
                self._defs_20 = defs20

            if "21_out05_明細2行目" in wb.sheetnames:
                ws = wb["21_out05_明細2行目"]
                defs21: list[dict[str, object]] = []
                for row in ws.iter_rows(min_row=5, values_only=True):
                    enabled, line_offset, col, source_kind, source_name, formatter, length, fixed_value, notes = row[:9]
                    if str(enabled).strip().upper() != "Y":
                        continue
                    if col in (None, ""):
                        continue
                    defs21.append({
                        "line_offset": int(line_offset or 0),
                        "col": str(col).zfill(3),
                        "source_kind": (source_kind or "").strip(),
                        "source_name": (source_name or "").strip(),
                        "formatter": (formatter or "").strip(),
                        "length": None if length in (None, "") else int(length),
                        "fixed_value": "" if fixed_value is None else str(fixed_value),
                        "notes": "" if notes is None else str(notes),
                    })
                self._defs_21 = defs21

            if "22_out05_合計" in wb.sheetnames:
                ws = wb["22_out05_合計"]
                defs22: list[dict[str, object]] = []
                for row in ws.iter_rows(min_row=5, values_only=True):
                    enabled, line_offset, col, source_kind, source_name, formatter, length, fixed_value, notes = row[:9]
                    if str(enabled).strip().upper() != "Y":
                        continue
                    if col in (None, ""):
                        continue
                    defs22.append({
                        "line_offset": int(line_offset or 0),
                        "col": str(col).zfill(3),
                        "source_kind": (source_kind or "").strip(),
                        "source_name": (source_name or "").strip(),
                        "formatter": (formatter or "").strip(),
                        "length": None if length in (None, "") else int(length),
                        "fixed_value": "" if fixed_value is None else str(fixed_value),
                        "notes": "" if notes is None else str(notes),
                    })
                self._defs_22 = defs22

            if "23_out05_備考" in wb.sheetnames:
                ws = wb["23_out05_備考"]
                defs23: list[dict[str, object]] = []
                for row in ws.iter_rows(min_row=5, values_only=True):
                    enabled, line_offset, col, source_kind, source_name, formatter, length, fixed_value, notes = row[:9]
                    if str(enabled).strip().upper() != "Y":
                        continue
                    if col in (None, ""):
                        continue
                    defs23.append({
                        "line_offset": int(line_offset or 0),
                        "col": str(col).zfill(3),
                        "source_kind": (source_kind or "").strip(),
                        "source_name": (source_name or "").strip(),
                        "formatter": (formatter or "").strip(),
                        "length": None if length in (None, "") else int(length),
                        "fixed_value": "" if fixed_value is None else str(fixed_value),
                        "notes": "" if notes is None else str(notes),
                    })
                self._defs_23 = defs23

        except Exception:
            self.logger.exception("Excel定義読込失敗: %s", self.def_xlsx)
            self._defs_page_header = []
            self._defs_11 = []
            self._defs_20 = []
            self._defs_21 = []
            self._defs_22 = []
            self._defs_23 = []
        else:
            self.logger.debug(
                "Excel定義読込完了 page_header=%s header=%s detail1=%s detail2=%s total=%s note=%s",
                len(self._defs_page_header), len(self._defs_11), len(self._defs_20), len(self._defs_21), len(self._defs_22), len(self._defs_23)
            )
            pass

    def _format_excel_value(self, value: str, formatter: str, length: int | None) -> bytes:
        if formatter == "literal":
            return value.encode("cp932", errors="ignore")
        if formatter == "ascii":
            return value.encode("ascii", errors="ignore")
        if formatter == "cf_str":
            return cf_str(value, int(length or 0))
        if formatter == "cf_jpn":
            return cf_jpn(value, int(length or 0))
        if formatter == "cf_jpn2":
            return cf_jpn2(value, int(length or 0))
        if formatter == "cf_jpn_name":
            return cf_jpn_name(value, int(length or 0))
        if formatter == "cf_jpn_003":
            return cf_jpn_003(value, int(length or 0))
        if formatter == "cf_int":
            return cf_int(value, int(length or 0))
        if formatter == "cf_ymd":
            return cf_ymd(value, int(length or 0))
        return value.encode("cp932", errors="ignore")

    def _resolve_excel_value(self, spec: dict[str, object], order: OrderRecord, fax_no: str) -> bytes:
        source_kind = str(spec["source_kind"])
        source_name = str(spec["source_name"])
        formatter = str(spec["formatter"])
        length = spec["length"]

        if source_kind == "fixed":
            fixed_value = str(spec["fixed_value"])
            if fixed_value == "":
                return b""
            return self._format_excel_value(fixed_value, formatter, length)

        if source_kind == "system":
            if source_name == "now_text":
                value = self.now.strftime("%y.%m.%d %H:%M")
            elif source_name == "count_page":
                value = str(self.count_page)
            else:
                value = ""
            return self._format_excel_value(value, formatter, length)

        if source_kind == "header":
            if source_name == "fax_no":
                value = fax_no
            else:
                value = str(getattr(order, source_name, ""))
            return self._format_excel_value(value, formatter, length)

        return b""

    def _build_page_header_from_excel(self, order: OrderRecord, fax_no: str) -> bytes | None:
        if not self._defs_page_header:
            return None
        parts: list[tuple[bytes, bytes, bytes]] = []
        current_line = 1
        for spec in self._defs_page_header:
            if str(spec["source_kind"]) == "fixed" and str(spec["fixed_value"]) == "":
                current_line += 1
                continue
            value = self._resolve_excel_value(spec, order, fax_no)
            if value == b"":
                continue
            line_no = cf_int(current_line + int(spec["line_offset"]), 3)
            col = str(spec["col"]).encode("ascii")
            if parts and parts[-1][0] == line_no and parts[-1][1] == col:
                parts[-1] = (parts[-1][0], parts[-1][1], parts[-1][2] + value)
            else:
                parts.append((line_no, col, value))
        rec = bytearray(b"D1")
        for line_no, col, value in parts:
            rec += line_no + X01 + col + X01 + value + XFF
        return bytes(rec)

    def _find_def_11(self, line_offset: int, col: str) -> dict[str, object] | None:
        target_col = str(col).zfill(3)
        for spec in self._defs_11:
            if int(spec["line_offset"]) == int(line_offset) and str(spec["col"]) == target_col:
                return spec
        return None

    def _field_or_default_11(
        self,
        line_offset: int,
        col: str,
        default_value: bytes,
        order: OrderRecord,
        fax_no: str,
    ) -> bytes:
        spec = self._find_def_11(line_offset, col)
        if spec is None:
            return default_value
        value = self._resolve_excel_value(spec, order, fax_no)
        if value == b"" and str(spec["source_kind"]) != "fixed":
            return default_value
        return value

    def _find_def_20(self, line_offset: int, col: str) -> dict[str, object] | None:
        target_col = str(col).zfill(3)
        for spec in self._defs_20:
            if int(spec["line_offset"]) == int(line_offset) and str(spec["col"]) == target_col:
                return spec
        return None

    def _field_or_default_20(
        self,
        line_offset: int,
        col: str,
        default_value: bytes,
        order: OrderRecord,
        fax_no: str,
        item: LineItem,
    ) -> bytes:
        spec = self._find_def_20(line_offset, col)
        if spec is None:
            return default_value
        value = self._resolve_excel_value_21(spec, order, fax_no, item)
        if value == b"" and str(spec["source_kind"]) != "fixed":
            return default_value
        return value

    def _find_def_22(self, line_offset: int, col: str) -> dict[str, object] | None:
        target_col = str(col).zfill(3)
        for spec in self._defs_22:
            if int(spec["line_offset"]) == int(line_offset) and str(spec["col"]) == target_col:
                return spec
        return None

    def _resolve_excel_value_22(self, spec: dict[str, object], order: OrderRecord, fax_no: str, extra: dict[str, int]) -> bytes:
        source_kind = str(spec["source_kind"])
        source_name = str(spec["source_name"])
        formatter = str(spec["formatter"])
        length = spec["length"]

        if source_kind == "fixed":
            fixed_value = str(spec["fixed_value"])
            if fixed_value == "":
                return b""
            return self._format_excel_value(fixed_value, formatter, length)

        if source_kind == "system":
            if source_name == "total_qty":
                value = str(extra.get("total_qty", ""))
            elif source_name == "total_amount":
                value = str(extra.get("total_amount", ""))
            elif source_name == "now_text":
                value = self.now.strftime("%y.%m.%d %H:%M")
            elif source_name == "count_page":
                value = str(self.count_page)
            else:
                value = ""
            return self._format_excel_value(value, formatter, length)

        if source_kind == "header":
            if source_name == "fax_no":
                value = fax_no
            else:
                value = str(getattr(order, source_name, ""))
            return self._format_excel_value(value, formatter, length)

        return b""

    def _field_or_default_22(
        self,
        line_offset: int,
        col: str,
        default_value: bytes,
        order: OrderRecord,
        fax_no: str,
        extra: dict[str, int],
    ) -> bytes:
        spec = self._find_def_22(line_offset, col)
        if spec is None:
            return default_value
        value = self._resolve_excel_value_22(spec, order, fax_no, extra)
        if value == b"" and str(spec["source_kind"]) != "fixed":
            return default_value
        return value

    def _find_def_23(self, line_offset: int, col: str) -> dict[str, object] | None:
        target_col = str(col).zfill(3)
        for spec in self._defs_23:
            if int(spec["line_offset"]) == int(line_offset) and str(spec["col"]) == target_col:
                return spec
        return None

    def _resolve_excel_value_23(self, spec: dict[str, object], order: OrderRecord, fax_no: str) -> bytes:
        source_kind = str(spec["source_kind"])
        source_name = str(spec["source_name"])
        formatter = str(spec["formatter"])
        length = spec["length"]

        if source_kind == "fixed":
            fixed_value = str(spec["fixed_value"])
            if fixed_value == "":
                return b""
            return self._format_excel_value(fixed_value, formatter, length)

        if source_kind == "header":
            if source_name == "fax_no":
                value = fax_no
            else:
                value = str(getattr(order, source_name, ""))
            return self._format_excel_value(value, formatter, length)

        if source_kind == "system":
            if source_name == "now_text":
                value = self.now.strftime("%y.%m.%d %H:%M")
            elif source_name == "count_page":
                value = str(self.count_page)
            else:
                value = ""
            return self._format_excel_value(value, formatter, length)

        return b""

    def _field_or_default_23(
        self,
        line_offset: int,
        col: str,
        default_value: bytes,
        order: OrderRecord,
        fax_no: str,
    ) -> bytes:
        spec = self._find_def_23(line_offset, col)
        if spec is None:
            return default_value
        value = self._resolve_excel_value_23(spec, order, fax_no)
        if value == b"" and str(spec["source_kind"]) != "fixed":
            return default_value
        return value

    def _find_def_21(self, line_offset: int, col: str) -> dict[str, object] | None:
        target_col = str(col).zfill(3)
        for spec in self._defs_21:
            if int(spec["line_offset"]) == int(line_offset) and str(spec["col"]) == target_col:
                return spec
        return None

    def _resolve_excel_value_21(self, spec: dict[str, object], order: OrderRecord, fax_no: str, item: LineItem) -> bytes:
        source_kind = str(spec["source_kind"])
        source_name = str(spec["source_name"])
        formatter = str(spec["formatter"])
        length = spec["length"]

        if source_kind == "fixed":
            fixed_value = str(spec["fixed_value"])
            if fixed_value == "":
                return b""
            return self._format_excel_value(fixed_value, formatter, length)

        if source_kind == "detail":
            if source_name == "unit_of_measure_display":
                value = "" if item.unit_of_measure == "00" else item.unit_of_measure
            else:
                value = str(getattr(item, source_name, ""))
            return self._format_excel_value(value, formatter, length)

        if source_kind == "header":
            if source_name == "fax_no":
                value = fax_no
            else:
                value = str(getattr(order, source_name, ""))
            return self._format_excel_value(value, formatter, length)

        if source_kind == "system":
            if source_name == "now_text":
                value = self.now.strftime("%y.%m.%d %H:%M")
            elif source_name == "count_page":
                value = str(self.count_page)
            else:
                value = ""
            return self._format_excel_value(value, formatter, length)

        return b""

    def _field_or_default_21(
        self,
        line_offset: int,
        col: str,
        default_value: bytes,
        order: OrderRecord,
        fax_no: str,
        item: LineItem,
    ) -> bytes:
        spec = self._find_def_21(line_offset, col)
        if spec is None:
            return default_value
        value = self._resolve_excel_value_21(spec, order, fax_no, item)
        if value == b"" and str(spec["source_kind"]) != "fixed":
            return default_value
        return value

    def do_tran(self, infile: str | Path) -> tuple[bool, str]:
        try:
            root = ET.parse(str(infile)).getroot()
            orders = self._parse_orders(root)

            for idx, order in enumerate(orders, start=1):
                self._out04(order)
                self._out05(order)

            return True, str(infile)
        except Exception as exc:
            self.logger.exception("XML処理失敗: %s", infile)
            return False, f"{infile}: {exc}"

    def _parse_orders(self, root: ET.Element) -> list[OrderRecord]:
        message = next((ch for ch in root if local_name(ch.tag) == "message"), None)
        list_of_orders = None
        if message is not None:
            for ch in list(message):
                if local_name(ch.tag) == "listOfOrders":
                    list_of_orders = ch
                    break
        if list_of_orders is None:
            return []

        buyer = next((ch for ch in list(list_of_orders) if local_name(ch.tag) == "buyer"), None)
        buyer_name = child_text(buyer, "name")
        orders: list[OrderRecord] = []
        for order in children(list_of_orders, "order"):
            parties = next((ch for ch in list(order) if local_name(ch.tag) == "parties"), None)
            receiver = next((ch for ch in list(parties) if local_name(ch.tag) == "receiver"), None) if parties is not None else None
            transfer = next((ch for ch in list(parties) if local_name(ch.tag) == "transferOfOwnershipLocation"), None) if parties is not None else None
            payee = next((ch for ch in list(parties) if local_name(ch.tag) == "payee"), None) if parties is not None else None
            seller = next((ch for ch in list(parties) if local_name(ch.tag) == "seller"), None) if parties is not None else None
            summary = next((ch for ch in list(order) if local_name(ch.tag) == "tradeSummary"), None)
            goods_major = next((ch for ch in list(summary) if local_name(ch.tag) == "goodsMajorCategory"), None) if summary is not None else None
            dates = next((ch for ch in list(summary) if local_name(ch.tag) == "dates"), None) if summary is not None else None
            instructions = next((ch for ch in list(summary) if local_name(ch.tag) == "instructions"), None) if summary is not None else None
            note = next((ch for ch in list(summary) if local_name(ch.tag) == "note"), None) if summary is not None else None
            trade_id = next((ch for ch in list(order) if local_name(ch.tag) == "tradeID"), None)

            note_text = child_text(note, "text")
            note_text_sbcs = child_text(note, "text_sbcs")
            fax_no = ""
            if note_text_sbcs.strip().upper().startswith("FAX"):
                fax_no = note_text_sbcs[4:].strip()
            record = OrderRecord(
                trade_number=child_text(trade_id, "tradeNumber"),
                receiver_code=child_text(receiver, "code"),
                receiver_gln=child_text(receiver, "gln"),
                receiver_name=child_text(receiver, "name"),
                transfer_code=child_text(transfer, "code"),
                transfer_gln=child_text(transfer, "gln"),
                payee_code=child_text(payee, "code"),
                payee_gln=child_text(payee, "gln"),
                seller_code=child_text(seller, "code"),
                seller_gln=child_text(seller, "gln"),
                seller_name=child_text(seller, "name"),
                buyer_name=buyer_name,
                major_category=child_text(goods_major, "majorCategory"),
                sub_major_category=child_text(goods_major, "subMajorCategory"),
                order_date_yy_mm_dd=self._yy_mm_dd(child_text(dates, "orderDate")),
                delivery_date_yy_mm_dd=self._yy_mm_dd(child_text(dates, "deliveryDateToReceiver")),
                trade_type_code=child_text(instructions, "tradeTypeCode"),
                goods_classification_code=child_text(instructions, "goodsClassificationCode"),
                note_text=note_text,
                fax_no=fax_no,
            )
            for line in children(order, "lineItem"):
                item_id = next((ch for ch in list(line) if local_name(ch.tag) == "itemID"), None)
                item_info = next((ch for ch in list(line) if local_name(ch.tag) == "itemInfo"), None)
                item_spec = next((ch for ch in list(item_info) if local_name(ch.tag) == "itemSpec"), None) if item_info is not None else None
                amounts = next((ch for ch in list(line) if local_name(ch.tag) == "amounts"), None)
                item_net_price = next((ch for ch in list(amounts) if local_name(ch.tag) == "itemNetPrice"), None) if amounts is not None else None
                quantities = next((ch for ch in list(line) if local_name(ch.tag) == "quantities"), None)
                order_qty = next((ch for ch in list(quantities) if local_name(ch.tag) == "orderQuantity"), None) if quantities is not None else None
                order_item_code_elem = next((ch for ch in list(item_id) if local_name(ch.tag) == "orderItemCode"), None) if item_id is not None else None
                line_id = next((ch for ch in list(line) if local_name(ch.tag) == "lineID"), None)
                record.items.append(LineItem(
                    line_no=child_text(line_id, "lineNumber"),
                    gtin=child_text(item_id, "gtin"),
                    order_item_code=(order_item_code_elem.text or "").strip() if order_item_code_elem is not None and order_item_code_elem.text else "",
                    code_type=(order_item_code_elem.attrib.get("codeType", "") if order_item_code_elem is not None else ""),
                    item_name=child_text(item_id, "name"),
                    spec=child_text(item_spec, "spec"),
                    unit_price=(item_net_price.attrib.get("unitPrice", "") if item_net_price is not None else ""),
                    net_amount=(item_net_price.text or "").strip() if item_net_price is not None and item_net_price.text else "",
                    quantity=child_text(order_qty, "quantity"),
                    unit_multiple=child_text(quantities, "unitMultiple"),
                    num_order_units=child_text(order_qty, "numOfOrderUnits"),
                    unit_of_measure=child_text(quantities, "unitOfMeasure"),
                ))
            orders.append(record)
        return orders

    @staticmethod
    def _yy_mm_dd(iso_date: str) -> str:
        if len(iso_date) == 10 and iso_date[4] == '-' and iso_date[7] == '-':
            return f"{iso_date[2:4]}.{iso_date[5:7]}.{iso_date[8:10]}"
        return iso_date

    def _write(self, payload: bytes) -> None:
        self.out_fp.write(payload + CRLF)

    def _out04(self, order: OrderRecord) -> None:
        fax_no = self.debug_fax_no if self.debug_mode == "Y" else (order.fax_no or self.bb_fax_no)
        if self.current_fax_no != fax_no or self.current_orocd != order.seller_code:
            self.current_fax_no = fax_no
            self.current_orocd = order.seller_code
            #self.logger.info("T0切替 fax_no=%s seller_code=%s trade_number=%s", fax_no, order.seller_code, order.trade_number)
            rec = (
                b"T0"
                + b"VANBB1"
                + cf_int(order.seller_code, 6)
                + b" " * 4
                + cf_str(fax_no, 30)
                + cf_str(self.bb_fax_no, 30)
                + cf_jpn2(order.seller_name, 30)
            )
            self.count_line = 0
            self._write(rec)

        if self.count_line > 55:
            self._write(b"D1" + b"999" + X01 + b"999" + XFF)
            self.count_line = 0

        if self.count_line == 0:
            self.count_page += 1
            #self.logger.info("ページ開始 page=%s trade_number=%s", self.count_page, order.trade_number)
            rec = self._build_page_header_from_excel(order, fax_no)
            if rec is None:
                now_text = self.now.strftime("%y.%m.%d %H:%M")
                rec = bytearray(b"D1")
                rec += b"001" + X01 + b"002" + X01 + cf_str(fax_no, 20) + XFF
                rec += b"002" + X01 + b"002" + X01 + cf_str(order.seller_code, 4) + b":" + cf_jpn(order.seller_name, 40) + XFF
                rec += b"002" + X01 + b"011" + X01 + now_text.encode("ascii") + XFF
                rec += b"002" + X01 + b"015" + X01 + str(self.count_page).encode("ascii") + XFF
                self._write(bytes(rec))
            else:
                self._write(rec)

        self.count_line += 4
        rec = bytearray(b"D1")
        rec += cf_int(self.count_line, 3) + X01 + b"006" + X01 + self._field_or_default_11(
            0, "006", cf_str(order.trade_number, 9), order, fax_no
        ) + XFF
        rec += cf_int(self.count_line, 3) + X01 + b"012" + X01 + self._field_or_default_11(
            0, "012", cf_ymd(order.order_date_yy_mm_dd, 8), order, fax_no
        ) + XFF
        rec += cf_int(self.count_line, 3) + X01 + b"014" + X01 + self._field_or_default_11(
            0, "014", cf_ymd(order.delivery_date_yy_mm_dd, 8), order, fax_no
        ) + XFF

        self.count_line += 1
        rec += cf_int(self.count_line, 3) + X01 + b"003" + X01 + self._field_or_default_11(
            1, "003", cf_jpn(order.buyer_name, 40), order, fax_no
        ) + XFF

        self.count_line += 1
        rec += cf_int(self.count_line, 3) + X01 + b"003" + X01 + self._field_or_default_11(
            2, "003", cf_jpn(order.receiver_name, 40), order, fax_no
        ) + XFF
        rec += cf_int(self.count_line, 3) + X01 + b"005" + X01 + self._field_or_default_11(
            2, "005", cf_str(order.receiver_code, 6), order, fax_no
        ) + XFF
        rec += cf_int(self.count_line, 3) + X01 + b"006" + X01 + self._field_or_default_11(
            2, "006", cf_str(order.sub_major_category, 10), order, fax_no
        ) + XFF
        rec += cf_int(self.count_line, 3) + X01 + b"007" + X01 + self._field_or_default_11(
            2, "007", cf_str(order.trade_type_code, 4), order, fax_no
        ) + XFF
        self._write(bytes(rec))

        rec2 = bytearray(b"D1")
        rec2 += cf_int(self.count_line, 3) + X01 + b"008" + X01 + self._field_or_default_11(
            2, "008", cf_str(order.seller_code, 4), order, fax_no
        ) + XFF
        rec2 += cf_int(self.count_line, 3) + X01 + b"010" + X01 + self._field_or_default_11(
            2, "010", cf_jpn(order.seller_name, 40), order, fax_no
        ) + XFF
        self._write(bytes(rec2))
        self.logger.debug("注文ヘッダ出力完了 trade_number=%s current_line=%s", order.trade_number, self.count_line)

    def _out05(self, order: OrderRecord) -> None:
        self.count_line += 2
        total_qty = Decimal("0")
        total_amount = Decimal("0")

        for item in order.items:
            self.logger.debug("明細出力開始 trade_number=%s line_no=%s item_name=%s", order.trade_number, item.line_no, item.item_name)
            unit_of_measure = "" if item.unit_of_measure == "00" else item.unit_of_measure

            self.count_line += 1
            detail_line_no = self.count_line + 1

            rec = bytearray(b"D1")
            rec += cf_int(self.count_line, 3) + X01 + b"002" + X01 + self._field_or_default_20(
                0, "002", cf_jpn(item.item_name, 50), order, self.current_fax_no or self.bb_fax_no, item
            ) + XFF

            rec += cf_int(detail_line_no, 3) + X01 + b"002" + X01 + self._field_or_default_21(
                0, "002", b"", order, self.current_fax_no or self.bb_fax_no, item
            ) + XFF
            rec += cf_int(detail_line_no, 3) + X01 + b"004" + X01 + self._field_or_default_21(
                0, "004", cf_jpn(item.spec, 20), order, self.current_fax_no or self.bb_fax_no, item
            ) + XFF
            rec += cf_int(detail_line_no, 3) + X01 + b"005" + X01 + self._field_or_default_21(
                0, "005", cf_str(item.order_item_code, 13), order, self.current_fax_no or self.bb_fax_no, item
            ) + XFF
            rec += cf_int(detail_line_no, 3) + X01 + b"007" + X01 + self._field_or_default_21(
                0, "007", cf_int(item.unit_multiple, 4), order, self.current_fax_no or self.bb_fax_no, item
            ) + XFF
            rec += cf_int(detail_line_no, 3) + X01 + b"008" + X01 + self._field_or_default_21(
                0, "008", item.num_order_units.encode("ascii", errors="ignore"), order, self.current_fax_no or self.bb_fax_no, item
            ) + XFF
            rec += cf_int(detail_line_no, 3) + X01 + b"009" + X01 + self._field_or_default_21(
                0, "009", unit_of_measure.encode("ascii", errors="ignore"), order, self.current_fax_no or self.bb_fax_no, item
            ) + XFF
            rec += cf_int(detail_line_no, 3) + X01 + b"010" + X01 + self._field_or_default_21(
                0, "010", item.quantity.encode("ascii", errors="ignore"), order, self.current_fax_no or self.bb_fax_no, item
            ) + XFF
            rec += cf_int(detail_line_no, 3) + X01 + b"012" + X01 + self._field_or_default_21(
                0, "012", item.unit_price.encode("ascii", errors="ignore"), order, self.current_fax_no or self.bb_fax_no, item
            ) + XFF
            rec += cf_int(detail_line_no, 3) + X01 + b"014" + X01 + self._field_or_default_21(
                0, "014", item.net_amount.encode("ascii", errors="ignore"), order, self.current_fax_no or self.bb_fax_no, item
            ) + XFF
            self._write(bytes(rec))

            self.count_line = detail_line_no

            try:
                total_qty += Decimal(item.quantity)
            except Exception:
                pass
            try:
                total_amount += Decimal(item.net_amount)
            except Exception:
                pass

        self.count_line += ((9 - len(order.items)) * 2) + 1
        rec = bytearray(b"D1")
        extra_total = {"total_qty": int(total_qty), "total_amount": int(total_amount)}
        rec += cf_int(self.count_line, 3) + X01 + b"010" + X01 + self._field_or_default_22(
            0, "010", cf_int(int(total_qty), 6), order, self.current_fax_no or self.bb_fax_no, extra_total
        ) + XFF
        rec += cf_int(self.count_line, 3) + X01 + b"014" + X01 + self._field_or_default_22(
            0, "014", cf_int(int(total_amount), 10), order, self.current_fax_no or self.bb_fax_no, extra_total
        ) + XFF
        self._write(bytes(rec))
        #self.logger.info("合計出力 trade_number=%s total_qty=%s total_amount=%s", order.trade_number, int(total_qty), int(total_amount))

        self.count_line += 1
        rec2 = bytearray(b"D1")
        rec2 += cf_int(self.count_line, 3) + X01 + b"003" + X01 + self._field_or_default_23(
            0, "003", cf_jpn_note(order.note_text, 20), order, self.current_fax_no or self.bb_fax_no
        ) + XFF
        self._write(bytes(rec2))
        self.logger.debug("備考出力完了 trade_number=%s current_line=%s", order.trade_number, self.count_line)