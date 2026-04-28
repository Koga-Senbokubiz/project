#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

DEFAULT_BOOK = "サンダイコー_order_データ変換定義書.xlsx"
DEFAULT_FROM_SHEET = "02.FromLayout"
DEFAULT_TO_SHEET = "03.ToLayout"
DEFAULT_MAPPING_SHEET = "04.MappingLogic"
DEFAULT_ENCODING = "cp932"

THIN = Side(style="thin", color="A6A6A6")
ALL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL_HEADER = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")
FONT_HEADER = Font(bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

@dataclass
class LayoutItem:
    group_id: str
    field_id: str
    field_name: str
    col: int
    start_pos: Optional[int] = None
    length: Optional[int] = None
    rec_type: str = ""

@dataclass
class BrxordSourceVar:
    var_name: str
    rec_type: str
    start1: Optional[int]
    length: Optional[int]
    comment: str
    source_kind: str
    literal_value: str

@dataclass
class BrxordMapping:
    to_group: str
    to_tag: str
    source_var: str
    source_kind: str
    rec_type: str
    start1: Optional[int]
    length: Optional[int]
    comment: str
    literal_value: str

ASSIGN_FIELD_RE = re.compile(
    r"""^\s*
        \$(?P<var>BMS[0-9A-Za-z_]+)\s*=\s*
        (?:(?:trim)\(\s*)?
        (?P<kind>substr|getJpn|getANK|getYMD2)
        \(
            \s*\$vXML\['(?P<rec>[A-Z])'\]\s*,
            \s*(?P<start>\d+)\s*,
            \s*(?P<length>\d+)
    """,
    re.X,
)

ASSIGN_LITERAL_RE = re.compile(
    r"""^\s*
        \$(?P<var>BMS[0-9A-Za-z_]+)\s*=\s*
        (?P<quote>['\"])(?P<val>.*?)(?P=quote)\s*;
    """,
    re.X,
)

COMMENT_RE = re.compile(r"//\s*(.+?)\s*$")
OPEN_TAG_RE = re.compile(r"<(?P<tag>[A-Za-z0-9_:-]+)>")
CLOSE_TAG_RE = re.compile(r"</(?P<tag>[A-Za-z0-9_:-]+)>")
BMS_VAR_RE = re.compile(r"\$(BMS[0-9A-Za-z_]+)")

def normalize_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip()

def normalize_name_for_match(text: str) -> str:
    s = normalize_value(text)
    s = s.replace(" ", "").replace("　", "")
    s = s.replace("（", "(").replace("）", ")")
    s = s.replace("／", "/")
    s = s.replace("－", "-").replace("―", "-")
    return s

def local_tag_name(tag: str) -> str:
    return tag.split(":", 1)[-1]

def find_label_rows(ws, max_scan_row: int = 120) -> Dict[str, List[int]]:
    labels: Dict[str, List[int]] = {}
    for r in range(1, min(ws.max_row, max_scan_row) + 1):
        label = normalize_value(ws.cell(r, 1).value)
        if label:
            labels.setdefault(label, []).append(r)
    return labels

def resolve_layout_rows(ws) -> Dict[str, int | List[int]]:
    labels = find_label_rows(ws)

    def require_last(label: str) -> int:
        rows = labels.get(label, [])
        if not rows:
            raise ValueError(f"{ws.title}: A列にラベル [{label}] が見つかりません。")
        return rows[-1]

    def require_one(label: str) -> int:
        rows = labels.get(label, [])
        if not rows:
            raise ValueError(f"{ws.title}: A列にラベル [{label}] が見つかりません。")
        return rows[0]

    group_rows: List[int] = []
    for i in range(20):
        key = f"グループID_lv{i}"
        rows = labels.get(key, [])
        if rows:
            group_rows.append(rows[0])
        else:
            if i > 0:
                break

    if not group_rows:
        raise ValueError(f"{ws.title}: グループID_lv0 以降が見つかりません。")

    result: Dict[str, int | List[int]] = {
        "group_rows": group_rows,
        "field_id_row": require_last("項目ID"),
        "field_name_row": require_last("項目名"),
    }
    if "開始位置" in labels:
        result["start_pos_row"] = require_one("開始位置")
    if "桁数" in labels:
        result["length_row"] = require_one("桁数")
    return result

def pick_effective_group_id(group_values: List[str]) -> str:
    non_empty = [g for g in group_values if g]
    return non_empty[-1] if non_empty else ""

def infer_rec_type_from_group(group_id: str) -> str:
    gid = group_id.upper()
    if gid.startswith("B_"):
        return "B"
    if gid.startswith("D_"):
        return "D"
    if gid.startswith("A_"):
        return "A"
    return ""

def read_layout_items(ws, is_from_fixed: bool) -> List[LayoutItem]:
    rows = resolve_layout_rows(ws)
    group_rows: List[int] = rows["group_rows"]
    field_id_row: int = rows["field_id_row"]
    field_name_row: int = rows["field_name_row"]
    start_pos_row = rows.get("start_pos_row")
    length_row = rows.get("length_row")

    items: List[LayoutItem] = []
    for col in range(2, ws.max_column + 1):
        field_id = normalize_value(ws.cell(field_id_row, col).value)
        if not field_id:
            continue

        field_name = normalize_value(ws.cell(field_name_row, col).value)
        group_values = [normalize_value(ws.cell(r, col).value) for r in group_rows]
        group_id = pick_effective_group_id(group_values)

        start_pos = None
        length = None
        rec_type = ""

        if is_from_fixed:
            if start_pos_row:
                raw = ws.cell(start_pos_row, col).value
                try:
                    start_pos = int(raw) if raw not in (None, "") else None
                except Exception:
                    start_pos = None
            if length_row:
                raw = ws.cell(length_row, col).value
                try:
                    length = int(raw) if raw not in (None, "") else None
                except Exception:
                    length = None
            rec_type = infer_rec_type_from_group(group_id)

        items.append(LayoutItem(group_id, field_id, field_name, col, start_pos, length, rec_type))
    return items

def read_php_from_zip(zip_path: Path, target_name: str, encoding: str) -> str:
    with zipfile.ZipFile(zip_path) as zf:
        candidates = [n for n in zf.namelist() if n.endswith("/" + target_name) or n.endswith(target_name)]
        if not candidates:
            raise FileNotFoundError(f"{target_name} が zip 内に見つかりません: {zip_path}")
        return zf.read(candidates[0]).decode(encoding, errors="replace")

def infer_group_for_line(lines: List[str], idx: int) -> str:
    stack: List[str] = []
    for j in range(max(0, idx - 40), idx + 1):
        line = lines[j]

        for m in OPEN_TAG_RE.finditer(line):
            tag = local_tag_name(m.group("tag"))
            if f"</{m.group('tag')}>" not in line:
                stack.append(tag)

        for m in CLOSE_TAG_RE.finditer(line):
            tag = local_tag_name(m.group("tag"))
            if stack and stack[-1] == tag:
                stack.pop()

    return stack[-1] if stack else ""

def parse_brxord_xml(php_text: str) -> List[BrxordMapping]:
    lines = php_text.splitlines()
    source_vars: Dict[str, BrxordSourceVar] = {}
    mappings: List[BrxordMapping] = []
    seen: set[Tuple[str, str, str]] = set()

    for line in lines:
        m = ASSIGN_FIELD_RE.match(line)
        if m:
            var = m.group("var")
            rec = m.group("rec")
            start1 = int(m.group("start")) + 1
            length = int(m.group("length"))
            cm = COMMENT_RE.search(line)
            comment = cm.group(1).strip() if cm else ""
            source_vars[var] = BrxordSourceVar(var, rec, start1, length, comment, "field", "")
            continue

        lm = ASSIGN_LITERAL_RE.match(line)
        if lm:
            var = lm.group("var")
            val = lm.group("val")
            cm = COMMENT_RE.search(line)
            comment = cm.group(1).strip() if cm else ""
            source_vars[var] = BrxordSourceVar(var, "", None, None, comment, "literal", val)

    for i, line in enumerate(lines):
        if "$BMS" not in line:
            continue
        tag_m = OPEN_TAG_RE.search(line)
        var_m = BMS_VAR_RE.search(line)
        if not (tag_m and var_m):
            continue

        to_tag = local_tag_name(tag_m.group("tag"))
        source_var = var_m.group(1)
        src = source_vars.get(source_var)
        if not src:
            continue

        to_group = infer_group_for_line(lines, i)

        key = (to_group, to_tag, source_var)
        if key in seen:
            continue
        seen.add(key)

        mappings.append(
            BrxordMapping(
                to_group=to_group,
                to_tag=to_tag,
                source_var=source_var,
                source_kind=src.source_kind,
                rec_type=src.rec_type,
                start1=src.start1,
                length=src.length,
                comment=src.comment,
                literal_value=src.literal_value,
            )
        )
    return mappings

def build_from_index(from_items: List[LayoutItem]) -> Dict[Tuple[str, int, int], LayoutItem]:
    idx: Dict[Tuple[str, int, int], LayoutItem] = {}
    for item in from_items:
        if item.rec_type and item.start_pos and item.length:
            idx[(item.rec_type, item.start_pos, item.length)] = item
    return idx

def match_from_by_name(comment: str, from_items: List[LayoutItem], rec_type: str) -> Optional[LayoutItem]:
    comment_norm = normalize_name_for_match(comment)
    if not comment_norm:
        return None

    candidates = [f for f in from_items if f.rec_type == rec_type]
    if not candidates:
        candidates = from_items

    for f in candidates:
        field_norm = normalize_name_for_match(f.field_name)
        if not field_norm:
            continue
        if comment_norm == field_norm or comment_norm in field_norm or field_norm in comment_norm:
            return f
    return None

def create_row(target_mark: str, from_group: str, from_id: str, from_name: str,
               to_group: str, to_id: str, to_name: str, fixed_or_expr: str, note: str) -> List[str]:
    return ["", target_mark, from_group, from_id, from_name, to_group, to_id, to_name, fixed_or_expr, note]

def build_mapping_rows(from_items: List[LayoutItem], to_items: List[LayoutItem], brxord_mappings: List[BrxordMapping]) -> List[List[str]]:
    from_idx = build_from_index(from_items)
    rows: List[List[str]] = []

    brxord_by_to: Dict[Tuple[str, str], BrxordMapping] = {}
    for m in brxord_mappings:
        key = (m.to_group, m.to_tag)
        if key not in brxord_by_to:
            brxord_by_to[key] = m

    for t in to_items:
        key = (t.group_id, t.field_id)
        bm = brxord_by_to.get(key)

        if bm is None:
            rows.append(create_row("", "", "", "", t.group_id, t.field_id, t.field_name, "", "BRXORD未設定"))
            continue

        if bm.source_kind == "literal":
            rows.append(create_row("〇", "", "", "", t.group_id, t.field_id, t.field_name, bm.literal_value, "BRXORD固定値"))
            continue

        f = None
        if bm.start1 and bm.length:
            f = from_idx.get((bm.rec_type, bm.start1, bm.length))
        if not f:
            f = match_from_by_name(bm.comment, from_items, bm.rec_type)

        if f:
            rows.append(create_row("〇", f.group_id, f.field_id, f.field_name, t.group_id, t.field_id, t.field_name, "", "BRXORD項目"))
        else:
            rows.append(create_row("△", f"{bm.rec_type}_候補" if bm.rec_type else "PHP候補", bm.source_var, bm.comment,
                                   t.group_id, t.field_id, t.field_name, "", "BRXORD候補 / From未特定"))

    return rows

def sort_rows_by_target(rows: List[List[str]]) -> List[List[str]]:
    def row_key(row: List[str]):
        to_group = normalize_value(row[5])
        to_field = normalize_value(row[6])
        from_group = normalize_value(row[2])
        from_field = normalize_value(row[3])
        return (to_group, to_field, from_group, from_field)
    return sorted(rows, key=row_key)

def renumber_rows(rows: List[List[str]]) -> List[List[str]]:
    for idx, row in enumerate(rows, start=1):
        row[0] = str(idx)
    return rows

def reset_mapping_sheet(ws) -> None:
    merged_ranges = list(ws.merged_cells.ranges)
    for mr in merged_ranges:
        ws.unmerge_cells(str(mr))
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.font = Font(bold=False)
            cell.alignment = Alignment()

def write_mapping_sheet(ws, rows: List[List[str]]) -> None:
    reset_mapping_sheet(ws)

    headers = [
        "項番", "変換対象",
        "変換元グループID", "変換元項目ID", "変換元項目名",
        "変換先グループID", "変換先項目ID", "変換先項目名",
        "固定値/式", "備考",
    ]
    for col, val in enumerate(headers, start=1):
        cell = ws.cell(1, col)
        cell.value = val
        cell.fill = FILL_HEADER
        cell.border = ALL_BORDER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx)
            cell.value = val
            cell.border = ALL_BORDER
            cell.alignment = ALIGN_LEFT if c_idx in (5, 8, 9, 10) else ALIGN_CENTER

    widths = {"A": 8, "B": 10, "C": 20, "D": 22, "E": 28, "F": 24, "G": 22, "H": 28, "I": 18, "J": 26}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

def main() -> None:
    parser = argparse.ArgumentParser(description="03.ToLayout を正として BRXORD_XML.php ベースで 04.MappingLogic を初期作成する")
    parser.add_argument("--book", default=DEFAULT_BOOK, help="データ変換定義書.xlsx")
    parser.add_argument("--zip", required=True, help="BRXORD_XML.php を含む zip")
    parser.add_argument("--from-sheet", default=DEFAULT_FROM_SHEET, help="変換元シート名")
    parser.add_argument("--to-sheet", default=DEFAULT_TO_SHEET, help="変換先シート名")
    parser.add_argument("--mapping-sheet", default=DEFAULT_MAPPING_SHEET, help="出力シート名")
    parser.add_argument("--encoding", default=DEFAULT_ENCODING, help="zip内PHPの文字コード")
    parser.add_argument("--out", default="", help="出力先xlsx。省略時は元ブックに _mapping_tobase を付ける")
    args = parser.parse_args()

    book_path = Path(args.book)
    zip_path = Path(args.zip)
    if not book_path.exists():
        raise FileNotFoundError(f"ブックが見つかりません: {book_path}")
    if not zip_path.exists():
        raise FileNotFoundError(f"zipが見つかりません: {zip_path}")

    wb = load_workbook(book_path)
    if args.from_sheet not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {args.from_sheet}")
    if args.to_sheet not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {args.to_sheet}")

    from_items = read_layout_items(wb[args.from_sheet], is_from_fixed=True)
    to_items = read_layout_items(wb[args.to_sheet], is_from_fixed=False)

    brxord_text = read_php_from_zip(zip_path, "BRXORD_XML.php", args.encoding)
    brxord_mappings = parse_brxord_xml(brxord_text)

    rows = build_mapping_rows(from_items, to_items, brxord_mappings)
    rows = sort_rows_by_target(rows)
    rows = renumber_rows(rows)

    if args.mapping_sheet in wb.sheetnames:
        mapping_ws = wb[args.mapping_sheet]
    else:
        mapping_ws = wb.create_sheet(args.mapping_sheet)

    write_mapping_sheet(mapping_ws, rows)

    out_path = Path(args.out) if args.out else book_path.with_name(f"{book_path.stem}_mapping_tobase{book_path.suffix}")
    wb.save(out_path)

    print("完了")
    print(f"入力ブック      : {book_path}")
    print(f"入力ZIP        : {zip_path}")
    print(f"変換元項目数    : {len(from_items)}")
    print(f"変換先項目数    : {len(to_items)}")
    print(f"BRXORD候補数   : {len(brxord_mappings)}")
    print(f"出力行数        : {len(rows)}")
    print(f"出力ファイル    : {out_path}")

if __name__ == "__main__":
    main()
