#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# =========================
# 設定
# =========================
DEFAULT_BOOK = "サンダイコー_order_データ変換定義書.xlsx"
DEFAULT_FROM_SHEET = "02.FromLayout"
DEFAULT_TO_SHEET = "03.ToLayout"
DEFAULT_MAPPING_SHEET = "04.MappingLogic"
DEFAULT_ENCODING = "cp932"

THIN = Side(style="thin", color="A6A6A6")
ALL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL_HEADER = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")
FONT_HEADER = Font(bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# =========================
# データクラス
# =========================
@dataclass
class LayoutItem:
    group_id: str
    field_id: str
    field_name: str
    col: int
    start_pos: Optional[int] = None
    length: Optional[int] = None
    rec_type: str = ""


@dataclass
class AstSourceVar:
    var_name: str
    rec_type: str
    start0: int
    length: int
    comment: str


@dataclass
class MappingCandidate:
    rec_type: str
    start1: int
    length: int
    source_comment: str
    to_tag: str
    note: str


# =========================
# 正規表現
# =========================
ASSIGN_RE = re.compile(
    r"""^\s*
        \$(?P<var>BMS[0-9A-Za-z_]+)\s*=\s*
        trim\(\s*substr\(\$rec,\s*(?P<start>\d+)\s*,\s*(?P<length>\d+)\)\s*\)\s*;
        \s*//\s*(?P<comment>.+?)\s*$
    """,
    re.X,
)

FUNC_START_RE = re.compile(r"^\s*function\s+(?P<name>makeXML[A-Z])\s*\(")
TAG_NAME_RE = re.compile(r"<(?P<tag>[A-Za-z0-9_:-]+)(?:\s+[^>]*)?>")
BMS_VAR_RE = re.compile(r"\$(BMS[0-9A-Za-z_]+)")


# =========================
# 共通
# =========================
def normalize_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def norm_text(s: str) -> str:
    return normalize_value(s).replace(" ", "").replace("　", "")


# =========================
# レイアウト読取
# =========================
def find_label_rows(ws, max_scan_row: int = 120) -> Dict[str, List[int]]:
    labels: Dict[str, List[int]] = {}
    for r in range(1, min(ws.max_row, max_scan_row) + 1):
        label = normalize_value(ws.cell(r, 1).value)
        if label:
            labels.setdefault(label, []).append(r)
    return labels


def resolve_layout_rows(ws) -> Dict[str, int | List[int]]:
    labels = find_label_rows(ws)

    def require_last(label: str) -> int:
        rows = labels.get(label, [])
        if not rows:
            raise ValueError(f"{ws.title}: A列にラベル [{label}] が見つかりません。")
        return rows[-1]

    def require_one(label: str) -> int:
        rows = labels.get(label, [])
        if not rows:
            raise ValueError(f"{ws.title}: A列にラベル [{label}] が見つかりません。")
        return rows[0]

    group_rows: List[int] = []
    for i in range(20):
        key = f"グループID_lv{i}"
        rows = labels.get(key, [])
        if rows:
            group_rows.append(rows[0])
        else:
            if i > 0:
                break

    if not group_rows:
        raise ValueError(f"{ws.title}: グループID_lv0 以降が見つかりません。")

    result: Dict[str, int | List[int]] = {
        "group_rows": group_rows,
        "field_id_row": require_last("項目ID"),
        "field_name_row": require_last("項目名"),
    }

    if "開始位置" in labels:
        result["start_pos_row"] = require_one("開始位置")
    if "桁数" in labels:
        result["length_row"] = require_one("桁数")

    return result


def pick_effective_group_id(group_values: List[str]) -> str:
    non_empty = [g for g in group_values if g]
    return non_empty[-1] if non_empty else ""


def infer_rec_type_from_group(group_id: str) -> str:
    gid = group_id.upper()
    if gid == "COMMON":
        return "COMMON"
    if gid.startswith("B_"):
        return "B"
    if gid.startswith("D_"):
        return "D"
    if gid.startswith("A_"):
        return "A"
    return ""


def read_layout_items(ws, is_from_fixed: bool) -> List[LayoutItem]:
    rows = resolve_layout_rows(ws)
    group_rows: List[int] = rows["group_rows"]  # type: ignore[assignment]
    field_id_row: int = rows["field_id_row"]  # type: ignore[assignment]
    field_name_row: int = rows["field_name_row"]  # type: ignore[assignment]

    start_pos_row = rows.get("start_pos_row")
    length_row = rows.get("length_row")

    items: List[LayoutItem] = []
    for col in range(2, ws.max_column + 1):
        field_id = normalize_value(ws.cell(field_id_row, col).value)
        if not field_id:
            continue

        field_name = normalize_value(ws.cell(field_name_row, col).value)
        group_values = [normalize_value(ws.cell(r, col).value) for r in group_rows]
        group_id = pick_effective_group_id(group_values)

        start_pos = None
        length = None
        rec_type = ""

        if is_from_fixed:
            if start_pos_row:
                raw = ws.cell(start_pos_row, col).value
                try:
                    start_pos = int(raw) if raw not in (None, "") else None
                except Exception:
                    start_pos = None
            if length_row:
                raw = ws.cell(length_row, col).value
                try:
                    length = int(raw) if raw not in (None, "") else None
                except Exception:
                    length = None
            rec_type = infer_rec_type_from_group(group_id)

        items.append(
            LayoutItem(
                group_id=group_id,
                field_id=field_id,
                field_name=field_name,
                col=col,
                start_pos=start_pos,
                length=length,
                rec_type=rec_type,
            )
        )

    return items


# =========================
# AST_Order.php 読取
# =========================
def read_ast_order_from_zip(zip_path: Path, encoding: str) -> str:
    with zipfile.ZipFile(zip_path) as zf:
        candidates = [n for n in zf.namelist() if n.endswith("/AST_Order.php") or n.endswith("AST_Order.php")]
        if not candidates:
            raise FileNotFoundError(f"AST_Order.php が zip 内に見つかりません: {zip_path}")
        target = candidates[0]
        return zf.read(target).decode(encoding, errors="replace")


def parse_ast_order_candidates(ast_text: str) -> List[MappingCandidate]:
    lines = ast_text.splitlines()

    current_func = ""
    source_vars: Dict[str, AstSourceVar] = {}
    candidates: List[MappingCandidate] = []

    # 1) BMS変数定義を収集
    for line in lines:
        fm = FUNC_START_RE.search(line)
        if fm:
            current_func = fm.group("name")
            continue

        am = ASSIGN_RE.match(line)
        if am:
            var = am.group("var")
            start0 = int(am.group("start"))
            length = int(am.group("length"))
            comment = am.group("comment")

            if current_func == "makeXMLB":
                rec_type = "B"
            elif current_func == "makeXMLD":
                rec_type = "D"
            elif current_func == "makeXMLC":
                rec_type = "B"
            else:
                rec_type = ""

            source_vars[var] = AstSourceVar(
                var_name=var,
                rec_type=rec_type,
                start0=start0,
                length=length,
                comment=comment,
            )

    # 2) tag と BMS 変数の対応を収集
    current_func = ""
    seen: Set[Tuple[str, int, int, str]] = set()

    for line in lines:
        fm = FUNC_START_RE.search(line)
        if fm:
            current_func = fm.group("name")
            continue

        tag_m = TAG_NAME_RE.search(line)
        var_m = BMS_VAR_RE.search(line)
        if not (tag_m and var_m):
            continue

        tag = tag_m.group("tag")
        var = var_m.group(1)

        if var not in source_vars:
            continue

        src = source_vars[var]

        if current_func not in ("makeXMLB", "makeXMLC", "makeXMLD"):
            continue

        key = (src.rec_type, src.start0, src.length, tag)
        if key in seen:
            continue
        seen.add(key)

        candidates.append(
            MappingCandidate(
                rec_type=src.rec_type,
                start1=src.start0 + 1,
                length=src.length,
                source_comment=src.comment,
                to_tag=tag,
                note=f"AST_Order.php:{var}",
            )
        )

    return candidates


# =========================
# マッチング
# =========================
def build_from_index(from_items: List[LayoutItem]) -> Dict[Tuple[str, int, int], LayoutItem]:
    idx: Dict[Tuple[str, int, int], LayoutItem] = {}
    for item in from_items:
        if item.rec_type and item.start_pos and item.length:
            key = (item.rec_type, item.start_pos, item.length)
            if key not in idx:
                idx[key] = item
    return idx


def build_to_index(to_items: List[LayoutItem]) -> Dict[str, LayoutItem]:
    idx: Dict[str, LayoutItem] = {}
    for item in to_items:
        if item.field_id and item.field_id not in idx:
            idx[item.field_id] = item
    return idx


def match_from_by_name(comment: str, from_items: List[LayoutItem], rec_type: str) -> Optional[LayoutItem]:
    c = norm_text(comment)
    if not c:
        return None

    # rec_type一致優先
    candidates = [f for f in from_items if f.rec_type == rec_type]
    if not candidates:
        candidates = from_items

    # 完全一致/部分一致
    for f in candidates:
        fn = norm_text(f.field_name)
        if not fn:
            continue
        if c == fn or c in fn or fn in c:
            return f

    return None


def build_mapping_rows(
    from_items: List[LayoutItem],
    to_items: List[LayoutItem],
    ast_candidates: List[MappingCandidate],
) -> List[List[str]]:
    from_idx = build_from_index(from_items)
    to_idx = build_to_index(to_items)

    rows: List[List[str]] = []
    used_from: Set[str] = set()
    used_to: Set[str] = set()

    seq = 1

    # 1. AST候補を優先
    for cand in ast_candidates:
        # ① rec_type + 開始位置 + 桁数 完全一致
        f = from_idx.get((cand.rec_type, cand.start1, cand.length))

        # ② 項目名で既存FromLayoutに寄せる
        if not f:
            f = match_from_by_name(cand.source_comment, from_items, cand.rec_type)

        t = to_idx.get(cand.to_tag)

        if not f and not t:
            continue

        rows.append([
            str(seq),                               # A 項番
            "項目" if (f and t) else "",           # B 変換対象
            f.group_id if f else "",               # C 変換元グループID
            f.field_id if f else "",               # D 変換元項目ID
            f.field_name if f else cand.source_comment,  # E 変換元項目名
            t.group_id if t else "",               # F 変換先グループID
            t.field_id if t else cand.to_tag,      # G 変換先項目ID
            t.field_name if t else "",             # H 変換先項目名
            "",                                    # I 固定値/式
            "候補あり / 要確認",                   # J 備考
        ])
        seq += 1

        if f:
            used_from.add(f.field_id)
        if t:
            used_to.add(t.field_id)

    # 2. 残った From
    for f in from_items:
        if f.field_id in used_from:
            continue

        rows.append([
            str(seq),
            "",
            f.group_id,
            f.field_id,
            f.field_name,
            "",
            "",
            "",
            "",
            "変換先未設定",
        ])
        seq += 1

    # 3. 残った To
    for t in to_items:
        if t.field_id in used_to:
            continue

        rows.append([
            str(seq),
            "",
            "",
            "",
            "",
            t.group_id,
            t.field_id,
            t.field_name,
            "",
            "変換元未設定",
        ])
        seq += 1

    return rows


# =========================
# Excel出力
# =========================
def reset_mapping_sheet(ws) -> None:
    merged_ranges = list(ws.merged_cells.ranges)
    for mr in merged_ranges:
        ws.unmerge_cells(str(mr))

    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.font = Font(bold=False)
            cell.alignment = Alignment()


def write_mapping_sheet(ws, rows: List[List[str]]) -> None:
    reset_mapping_sheet(ws)

    headers = [
        "項番", "変換対象",
        "変換元グループID", "変換元項目ID", "変換元項目名",
        "変換先グループID", "変換先項目ID", "変換先項目名",
        "固定値/式", "備考",
    ]

    for col, val in enumerate(headers, start=1):
        cell = ws.cell(1, col)
        cell.value = val
        cell.fill = FILL_HEADER
        cell.border = ALL_BORDER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx)
            cell.value = val
            cell.border = ALL_BORDER
            cell.alignment = ALIGN_LEFT if c_idx in (5, 8, 9, 10) else ALIGN_CENTER

    widths = {
        "A": 8,
        "B": 12,
        "C": 20,
        "D": 18,
        "E": 28,
        "F": 20,
        "G": 18,
        "H": 28,
        "I": 18,
        "J": 24,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


# =========================
# main
# =========================
def main() -> None:
    parser = argparse.ArgumentParser(description="04.MappingLogic の初期候補を AST_Order.php ベースで作成する")
    parser.add_argument("--book", default=DEFAULT_BOOK, help="データ変換定義書.xlsx")
    parser.add_argument("--zip", required=True, help="ogv-bat1.zip")
    parser.add_argument("--from-sheet", default=DEFAULT_FROM_SHEET, help="変換元シート名")
    parser.add_argument("--to-sheet", default=DEFAULT_TO_SHEET, help="変換先シート名")
    parser.add_argument("--mapping-sheet", default=DEFAULT_MAPPING_SHEET, help="出力シート名")
    parser.add_argument("--encoding", default=DEFAULT_ENCODING, help="zip内PHPの文字コード")
    parser.add_argument("--out", default="", help="出力先xlsx。省略時は元ブックに _mapping_init を付ける")
    args = parser.parse_args()

    book_path = Path(args.book)
    zip_path = Path(args.zip)

    if not book_path.exists():
        raise FileNotFoundError(f"ブックが見つかりません: {book_path}")
    if not zip_path.exists():
        raise FileNotFoundError(f"zipが見つかりません: {zip_path}")

    wb = load_workbook(book_path)

    if args.from_sheet not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {args.from_sheet}")
    if args.to_sheet not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {args.to_sheet}")

    from_ws = wb[args.from_sheet]
    to_ws = wb[args.to_sheet]

    from_items = read_layout_items(from_ws, is_from_fixed=True)
    to_items = read_layout_items(to_ws, is_from_fixed=False)

    ast_text = read_ast_order_from_zip(zip_path, encoding=args.encoding)
    ast_candidates = parse_ast_order_candidates(ast_text)

    mapping_rows = build_mapping_rows(from_items, to_items, ast_candidates)

    if args.mapping_sheet in wb.sheetnames:
        mapping_ws = wb[args.mapping_sheet]
    else:
        mapping_ws = wb.create_sheet(args.mapping_sheet)

    write_mapping_sheet(mapping_ws, mapping_rows)

    if args.out:
        out_path = Path(args.out)
    else:
        out_path = book_path.with_name(f"{book_path.stem}_mapping_init{book_path.suffix}")

    wb.save(out_path)

    print("完了")
    print(f"入力ブック   : {book_path}")
    print(f"入力ZIP     : {zip_path}")
    print(f"変換元シート : {args.from_sheet}")
    print(f"変換先シート : {args.to_sheet}")
    print(f"出力シート   : {args.mapping_sheet}")
    print(f"AST候補数    : {len(ast_candidates)}")
    print(f"変換元項目数 : {len(from_items)}")
    print(f"変換先項目数 : {len(to_items)}")
    print(f"出力行数     : {len(mapping_rows)}")
    print(f"出力ファイル : {out_path}")


if __name__ == "__main__":
    main()