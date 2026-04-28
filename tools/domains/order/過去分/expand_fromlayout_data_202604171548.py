#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
データ変換定義書の 02.FromLayout に、実データ(DAT)を横展開するツール。

使い方例:
python expand_fromlayout_data.py ^
  --book "サンダイコー_order_データ変換定義書.xlsx" ^
  --dat  "020120260128123042BRXORD.DAT"

任意:
  --sheet 02.FromLayout
  --start-row 16
  --max-group-level 5

ポイント:
- 横持ち定義 (1列=1項目) を前提
- 実データは start-row 行目以降へ展開
- A列に連番、B列以降に項目値を出力
- COMMON は全レコードに展開
- B_HEADER は B レコードだけ
- D_DETAIL は D レコードだけ
- 既存の実データ部分はクリアしてから再展開
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


# ---- 定義書内の固定行（テンプレート前提） -----------------------------
ROW_FIELD_ID_TOP = 1
ROW_FIELD_NAME_TOP = 2
ROW_GROUP_LV0 = 3
ROW_REPEAT_GROUP_ID = 9
ROW_REPEAT_ID = 10
ROW_FIELD_ID_DEF = 11
ROW_FIELD_NAME_DEF = 12
ROW_DATA_TYPE = 13
ROW_LENGTH = 14
ROW_START_POS = 15
DEFAULT_DATA_START_ROW = 16
DEFAULT_MAX_GROUP_LEVEL = 5


def normalize_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def detect_line_separator(raw: bytes) -> bytes:
    if b"\r\n" in raw:
        return b"\r\n"
    if b"\n" in raw:
        return b"\n"
    if b"\r" in raw:
        return b"\r"
    return b""


def read_records(dat_path: Path, encoding: str = "cp932") -> List[str]:
    raw = dat_path.read_bytes()
    sep = detect_line_separator(raw)
    if sep:
        parts = raw.split(sep)
    else:
        parts = [raw]

    records: List[str] = []
    for part in parts:
        if not part:
            continue
        records.append(part.decode(encoding, errors="replace"))
    return records


def get_group_rows(max_group_level: int) -> List[int]:
    # lv0 から lvN まで
    return [ROW_GROUP_LV0 + i for i in range(max_group_level + 1)]


def get_max_used_column(ws) -> int:
    return ws.max_column


def clear_existing_sample_area(ws, start_row: int) -> None:
    max_col = ws.max_column
    max_row = ws.max_row
    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).value = None


def parse_field_definitions(ws, max_group_level: int) -> List[Dict]:
    group_rows = get_group_rows(max_group_level)
    max_col = get_max_used_column(ws)

    fields: List[Dict] = []

    for col in range(2, max_col + 1):  # B列以降が項目
        field_id = normalize_value(ws.cell(ROW_FIELD_ID_DEF, col).value)
        if not field_id:
            continue

        field_name = normalize_value(ws.cell(ROW_FIELD_NAME_DEF, col).value)
        data_type = normalize_value(ws.cell(ROW_DATA_TYPE, col).value)
        length_val = ws.cell(ROW_LENGTH, col).value
        start_pos_val = ws.cell(ROW_START_POS, col).value
        repeat_group_id = normalize_value(ws.cell(ROW_REPEAT_GROUP_ID, col).value)
        repeat_id = normalize_value(ws.cell(ROW_REPEAT_ID, col).value)

        group_levels = []
        for row in group_rows:
            group_levels.append(normalize_value(ws.cell(row, col).value))

        if start_pos_val in (None, "") or length_val in (None, ""):
            continue

        try:
            start_pos = int(start_pos_val)
            length = int(length_val)
        except ValueError:
            continue

        fields.append(
            {
                "col": col,
                "field_id": field_id,
                "field_name": field_name,
                "data_type": data_type,
                "length": length,
                "start_pos": start_pos,
                "group_levels": group_levels,
                "repeat_group_id": repeat_group_id,
                "repeat_id": repeat_id,
            }
        )

    return fields


def infer_target_record_type(group_levels: List[str]) -> Optional[str]:
    """
    group_levels から対象レコード種別を推定する。

    例:
      COMMON        -> None (全レコード)
      B_HEADER      -> B
      D_DETAIL      -> D
      X_xxx         -> X
    """
    for group in group_levels:
        if not group:
            continue
        if group == "COMMON":
            return None
        m = re.match(r"^([A-Z])(?:_|$)", group)
        if m:
            return m.group(1)
    return None


def field_applies_to_record(field_def: Dict, rec_type: str) -> bool:
    target = infer_target_record_type(field_def["group_levels"])
    if target is None:
        return True
    return target == rec_type


def slice_value(record: str, start_pos: int, length: int) -> str:
    """
    固定長の開始位置は 1 始まり。
    """
    start_idx = start_pos - 1
    end_idx = start_idx + length
    return record[start_idx:end_idx]


def expand_samples(ws, records: List[str], fields: List[Dict], start_row: int) -> int:
    row = start_row
    seq = 1

    for record in records:
        if not record:
            continue

        rec_type = record[:1]
        ws.cell(row, 1).value = seq

        for field_def in fields:
            col = field_def["col"]
            if field_applies_to_record(field_def, rec_type):
                value = slice_value(record, field_def["start_pos"], field_def["length"])
                ws.cell(row, col).value = value
            else:
                ws.cell(row, col).value = None

        row += 1
        seq += 1

    return seq - 1


def main() -> None:
    parser = argparse.ArgumentParser(description="02.FromLayout に実データを横展開する")
    parser.add_argument("--book", required=True, help="データ変換定義書.xlsx")
    parser.add_argument("--dat", required=True, help="固定長DATファイル")
    parser.add_argument("--sheet", default="02.FromLayout", help="対象シート名")
    parser.add_argument("--start-row", type=int, default=DEFAULT_DATA_START_ROW, help="実データ展開開始行")
    parser.add_argument("--encoding", default="cp932", help="DATファイルの文字コード")
    parser.add_argument(
        "--max-group-level",
        type=int,
        default=DEFAULT_MAX_GROUP_LEVEL,
        help="グループ階層の最大深さ（lv0 を除いた数。例: 5 -> lv0～lv5 を読む）",
    )
    parser.add_argument("--out", default="", help="出力先xlsx。省略時は元ブックに _expanded を付けて保存")
    args = parser.parse_args()

    book_path = Path(args.book)
    dat_path = Path(args.dat)

    if not book_path.exists():
        raise FileNotFoundError(f"定義書が見つかりません: {book_path}")
    if not dat_path.exists():
        raise FileNotFoundError(f"DATファイルが見つかりません: {dat_path}")

    wb = load_workbook(book_path)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {args.sheet}")

    ws = wb[args.sheet]

    fields = parse_field_definitions(ws, args.max_group_level)
    if not fields:
        raise ValueError("項目定義を読めませんでした。02.FromLayout の行構成を確認してください。")

    records = read_records(dat_path, encoding=args.encoding)
    if not records:
        raise ValueError("DATファイルにレコードが見つかりませんでした。")

    clear_existing_sample_area(ws, args.start_row)
    rec_count = expand_samples(ws, records, fields, args.start_row)

    if args.out:
        out_path = Path(args.out)
    else:
        out_path = book_path.with_name(f"{book_path.stem}_expanded{book_path.suffix}")

    wb.save(out_path)

    print("完了")
    print(f"入力ブック: {book_path}")
    print(f"DATファイル : {dat_path}")
    print(f"対象シート : {args.sheet}")
    print(f"読込項目数 : {len(fields)}")
    print(f"展開件数   : {rec_count}")
    print(f"出力ファイル: {out_path}")


if __name__ == "__main__":
    main()
