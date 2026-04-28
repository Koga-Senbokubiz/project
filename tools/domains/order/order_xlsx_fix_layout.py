#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
order_fix_layout_v2.py

PHPソース(zip) と 固定長実データ(DAT) から、
データ変換定義書.xlsx の 02.FromLayout / 03.FromLayout を作成する。

【特徴】
- 固定長専用。max-group-level には依存しない
- 既存のデータ変換定義書.xlsx をテンプレートとして使い、見た目を壊さない
- zip を複数指定可能
- zip の指定順を優先順として扱う
- zip 内の PHP を総なめして、$vXML['B'] / $vXML['D'] への
  substr / getJpn / getANK / getYMD2 を抽出
- 同一 rec_type + start + length の候補が複数あれば、先に指定した zip / php を優先
- 実データDATから B / D のサンプルを流し込む
- COMMON と B_HEADER / D_DETAIL を自動設定
- 既存シートのA列ラベル・色・枠線・太字・中央揃え等はそのまま利用

【入力】
- --book      データ変換定義書.xlsx
- --zip       PHPソースを含むzip（複数指定可、先に書いたものが優先）
- --dat       固定長DAT
- --sheet     出力シート（既定: 02.FromLayout）
- --encoding  文字コード（既定: cp932）
- --out       出力先xlsx

【使い方例】
python order_fix_layout.py ^
  --book "データ変換定義書.xlsx" ^
  --zip  "src.zip" ^
  --zip  "ogv-bat22.zip" ^
  --dat  "020120260128123042BRXORD.DAT" ^
  --sheet "02.FromLayout"
"""

from __future__ import annotations

import argparse
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


DEFAULT_SHEET = "02.FromLayout"
DEFAULT_ENCODING = "cp932"


@dataclass
class SourceField:
    rec_type: str           # COMMON / B / D
    start_pos0: int         # 0-based
    length: int
    item_id: str
    item_name: str
    source_kind: str        # substr / getJpn / getANK / getYMD2
    comments: List[str]
    php_source: str
    priority: int           # 小さいほど優先


def normalize_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def detect_line_separator(raw: bytes) -> bytes:
    if b"\r\n" in raw:
        return b"\r\n"
    if b"\n" in raw:
        return b"\n"
    if b"\r" in raw:
        return b"\r"
    return b""


def read_records(dat_path: Path, encoding: str = DEFAULT_ENCODING) -> List[str]:
    raw = dat_path.read_bytes()
    sep = detect_line_separator(raw)
    parts = raw.split(sep) if sep else [raw]

    records: List[str] = []
    for part in parts:
        if not part:
            continue
        records.append(part.decode(encoding, errors="replace"))
    return records


def first_record_by_type(records: List[str]) -> Dict[str, str]:
    result: Dict[str, str] = {}
    for rec in records:
        if not rec:
            continue
        rt = rec[:1]
        if rt not in result:
            result[rt] = rec
    return result


EXTRACT_RE = re.compile(
    r"""(?P<var>\$[A-Za-z0-9_]+)\s*=\s*
        (?P<func>substr|getJpn|getANK|getYMD2)\s*
        \(\s*\$vXML\['(?P<rec>[A-Z])'\]\s*,\s*
        (?P<start>\d+)\s*,\s*
        (?P<length>\d+)
    """,
    re.X,
)
COMMENT_RE = re.compile(r"//\s*(.+?)\s*$")


def read_all_php_from_zips(zip_paths: List[Path], encoding: str) -> List[Tuple[int, str, str]]:
    """
    zip 内のすべての PHP を読む。
    戻り値: [(priority, source_name, php_text), ...]
    priority は zip の指定順（0始まり）
    """
    php_texts: List[Tuple[int, str, str]] = []

    for priority, zip_path in enumerate(zip_paths):
        with zipfile.ZipFile(zip_path) as zf:
            for name in zf.namelist():
                if not name.lower().endswith(".php"):
                    continue
                try:
                    text = zf.read(name).decode(encoding, errors="replace")
                except Exception:
                    continue
                php_texts.append((priority, f"{zip_path.name}:{name}", text))

    return php_texts


def parse_php_fields(php_texts: List[Tuple[int, str, str]]) -> List[SourceField]:
    """
    zip 内の複数 PHP から B/D レコードの抽出位置を拾う。
    同じ rec_type + start + length は1項目にまとめる。
    候補が複数ある場合は priority が最小（先指定zip）が優先。
    """
    merged: Dict[Tuple[str, int, int], Dict] = {}

    for priority, source_name, php_text in php_texts:
        for line in php_text.splitlines():
            m = EXTRACT_RE.search(line)
            if not m:
                continue

            var_name = m.group("var").lstrip("$")
            func = m.group("func")
            rec_type = m.group("rec")
            start0 = int(m.group("start"))
            length = int(m.group("length"))

            cm = COMMENT_RE.search(line)
            comment = cm.group(1).strip() if cm else ""

            key = (rec_type, start0, length)
            if key not in merged:
                merged[key] = {
                    "rec_type": rec_type,
                    "start0": start0,
                    "length": length,
                    "func": func,
                    "comments": [],
                    "vars": [],
                    "sources": [],
                    "priority": priority,
                    "best_comment": comment,
                    "best_var": var_name,
                    "best_source": source_name,
                }

            current = merged[key]

            if comment and comment not in current["comments"]:
                current["comments"].append(comment)
            if var_name and var_name not in current["vars"]:
                current["vars"].append(var_name)
            if source_name not in current["sources"]:
                current["sources"].append(source_name)

            # 優先度が高いものを代表候補にする
            if priority < current["priority"]:
                current["priority"] = priority
                current["func"] = func
                current["best_comment"] = comment
                current["best_var"] = var_name
                current["best_source"] = source_name

    fields: List[SourceField] = []

    # COMMON: 先頭1桁をレコードタイプとして固定追加
    fields.append(
        SourceField(
            rec_type="COMMON",
            start_pos0=0,
            length=1,
            item_id="rec_type",
            item_name="レコードタイプ",
            source_kind="substr",
            comments=["レコードタイプ"],
            php_source="COMMON",
            priority=-1,
        )
    )

    items = sorted(
        [v for v in merged.values() if v["rec_type"] in ("B", "D")],
        key=lambda x: (x["rec_type"], x["start0"], x["length"]),
    )

    for item in items:
        rec_type = item["rec_type"]
        start0 = item["start0"]
        length = item["length"]
        comments = item["comments"]
        vars_ = item["vars"]

        # 優先代表
        best_comment = item["best_comment"]
        best_var = item["best_var"]
        best_source = item["best_source"]

        if best_comment:
            item_name = best_comment
        elif comments:
            item_name = comments[0]
        elif best_var:
            item_name = best_var
        elif vars_:
            item_name = vars_[0]
        else:
            item_name = f"{rec_type}_{start0+1}_{length}"

        item_id = f"{rec_type.lower()}_{start0+1:03d}_{length:03d}"

        fields.append(
            SourceField(
                rec_type=rec_type,
                start_pos0=start0,
                length=length,
                item_id=item_id,
                item_name=item_name,
                source_kind=item["func"],
                comments=comments,
                php_source=best_source,
                priority=item["priority"],
            )
        )

    return fields


def find_label_rows(ws, max_scan_row: int = 120) -> Dict[str, List[int]]:
    labels: Dict[str, List[int]] = {}
    for r in range(1, min(ws.max_row, max_scan_row) + 1):
        label = normalize_value(ws.cell(r, 1).value)
        if label:
            labels.setdefault(label, []).append(r)
    return labels


def resolve_layout_rows(ws) -> Dict[str, int | List[int]]:
    labels = find_label_rows(ws)

    def require(label: str) -> int:
        rows = labels.get(label, [])
        if not rows:
            raise ValueError(f"A列にラベル [{label}] が見つかりません。")
        return rows[0]

    group_rows = []
    for i in range(6):   # 固定長は lv0～lv5 想定
        key = f"グループID_lv{i}"
        rows = labels.get(key, [])
        if rows:
            group_rows.append(rows[0])

    if not group_rows:
        raise ValueError("グループID_lv0～lv5 が見つかりません。")

    row_repeat_group = require("繰返しグループID")
    row_repeat_id = require("繰返しID")
    row_field_id_rows = labels.get("項目ID", [])
    row_field_name_rows = labels.get("項目名", [])
    row_data_type = require("属性")
    row_length = require("桁数")
    row_start_pos = require("開始位置")
    row_required_rows = labels.get("必須", [])

    if not row_field_id_rows or not row_field_name_rows:
        raise ValueError("項目ID / 項目名 行が見つかりません。")

    row_field_id = row_field_id_rows[-1]
    row_field_name = row_field_name_rows[-1]
    row_required = row_required_rows[0] if row_required_rows else -1
    sample_start = (row_required if row_required > 0 else row_start_pos) + 1

    return {
        "group_rows": group_rows,
        "row_repeat_group": row_repeat_group,
        "row_repeat_id": row_repeat_id,
        "row_field_id": row_field_id,
        "row_field_name": row_field_name,
        "row_data_type": row_data_type,
        "row_length": row_length,
        "row_start_pos": row_start_pos,
        "row_required": row_required,
        "row_sample_start": sample_start,
    }


def clear_data_area(ws, sample_start_row: int) -> None:
    """
    値だけ消して書式は残す。
    """
    for r in range(1, ws.max_row + 1):
        for c in range(2, ws.max_column + 1):
            ws.cell(r, c).value = None

    for r in range(sample_start_row, ws.max_row + 1):
        ws.cell(r, 1).value = None


def slice_safe(record: str, start1: int, length: int) -> str:
    start0 = start1 - 1
    end0 = start0 + length
    if start0 >= len(record):
        return ""
    return record[start0:end0]


def write_layout(ws, rows: Dict[str, int | List[int]], fields: List[SourceField], samples: Dict[str, str]) -> None:
    group_rows: List[int] = rows["group_rows"]  # type: ignore[assignment]
    row_repeat_group: int = rows["row_repeat_group"]  # type: ignore[assignment]
    row_repeat_id: int = rows["row_repeat_id"]  # type: ignore[assignment]
    row_field_id: int = rows["row_field_id"]  # type: ignore[assignment]
    row_field_name: int = rows["row_field_name"]  # type: ignore[assignment]
    row_data_type: int = rows["row_data_type"]  # type: ignore[assignment]
    row_length: int = rows["row_length"]  # type: ignore[assignment]
    row_start_pos: int = rows["row_start_pos"]  # type: ignore[assignment]
    row_required: int = rows["row_required"]  # type: ignore[assignment]
    row_sample_start: int = rows["row_sample_start"]  # type: ignore[assignment]

    clear_data_area(ws, row_sample_start)

    col = 2
    for f in fields:
        level_values = [""] * len(group_rows)
        if len(group_rows) >= 1:
            level_values[0] = "order"
        if len(group_rows) >= 2:
            if f.rec_type == "COMMON":
                level_values[1] = "COMMON"
            elif f.rec_type == "B":
                level_values[1] = "B_HEADER"
            elif f.rec_type == "D":
                level_values[1] = "D_DETAIL"

        for idx, rownum in enumerate(group_rows):
            ws.cell(rownum, col).value = level_values[idx]

        ws.cell(row_repeat_group, col).value = "D_DETAIL" if f.rec_type == "D" else ""
        ws.cell(row_repeat_id, col).value = "1" if f.rec_type == "D" else ""
        ws.cell(row_field_id, col).value = f.item_id
        ws.cell(row_field_name, col).value = f.item_name
        ws.cell(row_data_type, col).value = "String"
        ws.cell(row_length, col).value = f.length
        ws.cell(row_start_pos, col).value = f.start_pos0 + 1
        if row_required > 0:
            ws.cell(row_required, col).value = "〇"

        col += 1

    # サンプル行: 1行目=B、2行目=D
    sample_rows: List[Tuple[str, str]] = []
    if "B" in samples:
        sample_rows.append(("B", samples["B"]))
    if "D" in samples:
        sample_rows.append(("D", samples["D"]))
    if not sample_rows and samples:
        for k, v in samples.items():
            sample_rows.append((k, v))

    for idx, (rt, rec) in enumerate(sample_rows):
        excel_row = row_sample_start + idx
        ws.cell(excel_row, 1).value = idx + 1

        for cidx, f in enumerate(fields, start=2):
            if f.rec_type == "COMMON":
                value = slice_safe(rec, f.start_pos0 + 1, f.length)
            elif f.rec_type == rt:
                value = slice_safe(rec, f.start_pos0 + 1, f.length)
            else:
                value = ""
            ws.cell(excel_row, cidx).value = value


def main() -> None:
    parser = argparse.ArgumentParser(description="PHPソース(zip)と固定長実データから order固定長レイアウトを作成する")
    parser.add_argument("--book", required=True, help="データ変換定義書.xlsx")
    parser.add_argument("--zip", action="append", required=True, help="PHPソースを含むzip（複数指定可、先指定が優先）")
    parser.add_argument("--dat", required=True, help="固定長DAT")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="出力シート名")
    parser.add_argument("--encoding", default=DEFAULT_ENCODING, help="文字コード")
    parser.add_argument("--out", default="", help="出力先xlsx。省略時は元ブックに _fix_layout を付けて保存")
    args = parser.parse_args()

    book_path = Path(args.book)
    dat_path = Path(args.dat)
    zip_paths = [Path(z) for z in args.zip]

    if not book_path.exists():
        raise FileNotFoundError(f"定義書が見つかりません: {book_path}")
    if not dat_path.exists():
        raise FileNotFoundError(f"DATが見つかりません: {dat_path}")
    for zp in zip_paths:
        if not zp.exists():
            raise FileNotFoundError(f"zipが見つかりません: {zp}")

    php_texts = read_all_php_from_zips(zip_paths, encoding=args.encoding)
    if not php_texts:
        raise ValueError("zip内にPHPが見つかりませんでした。")

    fields = parse_php_fields(php_texts)
    if not fields:
        raise ValueError("PHPから固定長項目を抽出できませんでした。対象zipを確認してください。")

    records = read_records(dat_path, encoding=args.encoding)
    sample_map = first_record_by_type(records)

    wb = load_workbook(book_path)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {args.sheet}")
    ws = wb[args.sheet]

    rows = resolve_layout_rows(ws)
    write_layout(ws, rows, fields, sample_map)

    if args.out:
        out_path = Path(args.out)
    else:
        out_path = book_path.with_name(f"{book_path.stem}_fix_layout{book_path.suffix}")

    wb.save(out_path)

    print("完了")
    print(f"入力ブック : {book_path}")
    print(f"入力ZIP   : {', '.join(str(z) for z in zip_paths)}")
    print(f"入力DAT   : {dat_path}")
    print(f"対象シート: {args.sheet}")
    print(f"抽出項目数: {len(fields)}")
    print(f"出力      : {out_path}")


if __name__ == "__main__":
    main()