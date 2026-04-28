#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
create_order_sd_to_layout.py

サンダイコー案件用：
03.ToLayout から EE XML形式レイアウトXMLを作成する。

方針：
- 基本形1_3：発注Ver1_3.xml をテンプレとして使う
- テンプレの schema / データストアXML属性 / 各種デフォルト属性を利用する
- XML形式レイアウトに必要な以下を一体で再生成する
    レイアウト
    データストア
    データストアXML
    レコード
    レコードXML
    レコード補助情報
    項目グループ
    項目
    項目XML
    項目補助情報
"""

import argparse
from collections import OrderedDict, defaultdict
from pathlib import Path
import xml.etree.ElementTree as ET

from openpyxl import load_workbook


XS_NS = "http://www.w3.org/2001/XMLSchema"
MSDATA_NS = "urn:schemas-microsoft-com:xml-msdata"

ET.register_namespace("xs", XS_NS)
ET.register_namespace("msdata", MSDATA_NS)


# =========================
# 共通
# =========================
def local_name(tag: str) -> str:
    return tag.split("}", 1)[1] if "}" in tag else tag


def cell_str(ws, row: int, col: int) -> str:
    v = ws.cell(row, col).value
    if v is None:
        return ""
    return str(v).strip()


def find_row(ws, label: str) -> int:
    for r in range(1, min(ws.max_row, 300) + 1):
        if cell_str(ws, r, 1) == label:
            return r
    raise ValueError(f"A列に [{label}] が見つかりません。")


def to_int(value: str, default: int = 0) -> int:
    try:
        if value == "":
            return default
        return int(float(value))
    except Exception:
        return default


def add_elem(root, tag: str, attrs: dict) -> ET.Element:
    e = ET.Element(tag)
    for k, v in attrs.items():
        e.set(k, "" if v is None else str(v))
    root.append(e)
    return e


def attr_to_ee(attr: str) -> str:
    a = (attr or "").strip().lower()

    if a in ("date", "datetime", "日付"):
        return "8"

    if a in ("time", "時刻"):
        return "10"

    if a in ("number", "numeric", "decimal", "integer", "int", "数値"):
        return "5"

    return "1"


# =========================
# Excel読取
# =========================
def read_to_layout(ws):
    group_rows = []

    i = 0
    while True:
        label = f"グループID_lv{i}"
        try:
            group_rows.append(find_row(ws, label))
            i += 1
        except ValueError:
            break

    if not group_rows:
        raise ValueError("グループID_lv0～ が見つかりません。")

    try:
        row_repeat_group = find_row(ws, "繰返しグループID")
    except ValueError:
        row_repeat_group = None

    try:
        row_repeat_id = find_row(ws, "繰返しID")
    except ValueError:
        row_repeat_id = None

    row_item_id = find_row(ws, "項目ID")
    row_item_name = find_row(ws, "項目名")
    row_attr = find_row(ws, "属性")

    try:
        row_length = find_row(ws, "桁数")
    except ValueError:
        row_length = None

    try:
        row_required = find_row(ws, "必須")
    except ValueError:
        row_required = None

    fields = []

    for col in range(2, ws.max_column + 1):
        item_id = cell_str(ws, row_item_id, col)
        if not item_id:
            continue

        groups = []
        for r in group_rows:
            g = cell_str(ws, r, col)
            if g:
                groups.append(g)

        if not groups:
            continue

        fields.append({
            "col": col,
            "groups": groups,
            "repeat_group": cell_str(ws, row_repeat_group, col) if row_repeat_group else "",
            "repeat_id": cell_str(ws, row_repeat_id, col) if row_repeat_id else "",
            "item_id": item_id,
            "item_name": cell_str(ws, row_item_name, col) or item_id,
            "attr": cell_str(ws, row_attr, col) or "String",
            "length": to_int(cell_str(ws, row_length, col), 0) if row_length else 0,
            "required": cell_str(ws, row_required, col) if row_required else "",
        })

    if not fields:
        raise ValueError("03.ToLayout から項目を取得できませんでした。")

    return fields


# =========================
# テンプレート解析
# =========================
def first_attrs(root, tag_name: str, fallback: dict) -> dict:
    for e in root:
        if local_name(e.tag) == tag_name:
            return dict(e.attrib)
    return dict(fallback)


def collect_template_name_map(root):
    """
    テンプレ側の sh: / common: / order: 付き名称をできる限り再利用する。
    同一ローカル名が一意に決まる場合だけ補正する。
    """
    names = defaultdict(set)

    for e in root:
        if local_name(e.tag) == "レイアウト":
            name = e.attrib.get("Name", "")
            if not name:
                continue
            local = name.split(":")[-1]
            names[local].add(name)

    resolved = {}
    for local, values in names.items():
        if len(values) == 1:
            resolved[local] = next(iter(values))

    return resolved


def normalize_template_name(name: str, name_map: dict) -> str:
    if not name:
        return name
    if ":" in name:
        return name
    return name_map.get(name, name)


def get_template_defaults(template_root):
    return {
        "datastore": first_attrs(template_root, "データストア", {
            "ID": "1",
            "フォーマット": "6",
        }),
        "datastore_xml": first_attrs(template_root, "データストアXML", {
            "ID": "1",
            "ルートタグ名": "sh:StandardBusinessDocument",
            "XMLスキーマ指定": "1",
            "XMLスキーマファイルパス": "",
            "桁あふれ": "1",
            "名前空間指定": "0",
            "名前空間定義": "",
            "文字コード": "12",
            "改行処理": "0",
            "インデント処理": "1",
            "インデント文字コード": "",
            "SchemaLocation": "",
            "SchemaLocationを出力する": "True",
        }),
        "record_xml": first_attrs(template_root, "レコードXML", {
            "ID": "1",
            "階層番号": "1",
            "必須": "1",
            "親レコード名": "",
            "挿入する直前の項目名": "",
        }),
        "record_aux": first_attrs(template_root, "レコード補助情報", {
            "ID": "1",
            "レコード出力タイプ": "2",
            "項目ID配列": "",
            "評価方法": "0",
            "評価順序": "0",
            "with空削除": "False",
        }),
        "item_group": first_attrs(template_root, "項目グループ", {
            "ID": "1",
            "項目グループ名": "",
            "項目グループの繰り返し回数": "1",
            "項目グループ種別": "0",
            "必須": "1",
            "タグの出現": "毎回",
            "項目グループの種別_EDIFACT": "通常の項目グループ",
            "必須_EDIFACT": "任意",
        }),
        "item_xml": first_attrs(template_root, "項目XML", {
            "ID": "1",
            "必須": "0",
            "データ型": "0",
            "全体の桁数": "0",
            "小数部桁数": "0",
            "パディングを行う": "False",
            "パディング": "2",
            "位置": "1",
            "空白を削除する": "True",
            "金融EDI情報_外部ファイルを使用する": "False",
            "金融EDI情報_外部ファイルフォーマット": "1",
            "金融EDI情報_外部ファイル": "",
        }),
        "item_aux": first_attrs(template_root, "項目補助情報", {
            "ID": "1",
            "全角": "False",
            "大文字": "False",
            "暦書式": "",
            "暦エラータイプ": "0",
            "レコード出力タイプ": "0",
            "最終出力タイプ": "0",
            "最終出力値": "",
            "項目出力": "True",
            "全角半角変換": "True",
            "大文字小文字変換": "True",
        }),
    }


def keep_schema_only(root):
    schema_nodes = [c for c in list(root) if local_name(c.tag) == "schema"]

    for c in list(root):
        root.remove(c)

    for s in schema_nodes:
        root.append(s)


# =========================
# 構造生成
# =========================
def split_record_and_path(field):
    groups = field["groups"]

    if len(groups) >= 2:
        root_name = groups[0]
        record_name = groups[1]
        inner_groups = groups[2:]
    else:
        root_name = groups[0]
        record_name = groups[0]
        inner_groups = []

    return root_name, record_name, inner_groups


def build_record_map(fields, name_map):
    records = OrderedDict()

    for f in fields:
        root_name, record_name, inner_groups = split_record_and_path(f)
        record_name = normalize_template_name(record_name, name_map)

        if record_name not in records:
            records[record_name] = {
                "root_name": normalize_template_name(root_name, name_map),
                "fields": [],
            }

        f2 = dict(f)
        f2["record_name"] = record_name
        f2["inner_groups"] = [normalize_template_name(g, name_map) for g in inner_groups]
        f2["item_name_norm"] = normalize_template_name(f2["item_name"], name_map)
        f2["item_id_norm"] = f2["item_id"]

        records[record_name]["fields"].append(f2)

    return records


# =========================
# EE XML形式レイアウト生成
# =========================
def build_ee_xml_layout(template_path: Path, fields):
    template_tree = ET.parse(template_path)
    template_root = template_tree.getroot()

    defaults = get_template_defaults(template_root)
    name_map = collect_template_name_map(template_root)
    records = build_record_map(fields, name_map)

    tree = ET.parse(template_path)
    root = tree.getroot()

    # schema以外は再生成する
    keep_schema_only(root)

    layout_id = 1
    record_id = 1
    item_group_id = 1
    item_id = 1

    # -------------------------
    # レイアウト：データストア
    # -------------------------
    add_elem(root, "レイアウト", {
        "ID": str(layout_id),
        "Name": "データストア",
        "Type": "1",
        "PropertyID": "1",
        "ParentsID": "0",
    })
    datastore_layout_id = layout_id
    layout_id += 1

    # -------------------------
    # データストア実体
    # -------------------------
    ds = dict(defaults["datastore"])
    ds["ID"] = "1"
    add_elem(root, "データストア", ds)

    dsxml = dict(defaults["datastore_xml"])
    dsxml["ID"] = "1"
    add_elem(root, "データストアXML", dsxml)

    # -------------------------
    # レコード
    # -------------------------
    record_layout_id_map = {}
    record_id_map = {}

    for record_name, record_info in records.items():
        rid = record_id
        record_id_map[record_name] = rid

        add_elem(root, "レイアウト", {
            "ID": str(layout_id),
            "Name": record_name,
            "Type": "3",
            "PropertyID": str(rid),
            "ParentsID": str(datastore_layout_id),
        })
        record_layout_id_map[record_name] = layout_id
        layout_id += 1

        add_elem(root, "レコード", {
            "ID": str(rid),
            "レコード名": record_name,
        })

        rx = dict(defaults["record_xml"])
        rx["ID"] = str(rid)
        rx["階層番号"] = rx.get("階層番号", "1")
        rx["必須"] = rx.get("必須", "1")
        rx["親レコード名"] = record_info["root_name"]
        add_elem(root, "レコードXML", rx)

        ra = dict(defaults["record_aux"])
        ra["ID"] = str(rid)
        add_elem(root, "レコード補助情報", ra)

        record_id += 1

    # -------------------------
    # 項目グループ・項目
    # -------------------------
    group_map = {}

    def get_group_layout(record_name, group_path):
        nonlocal layout_id, item_group_id

        parent_layout_id = record_layout_id_map[record_name]
        current_path = []

        for group_name in group_path:
            current_path.append(group_name)
            key = (record_name, tuple(current_path))

            if key in group_map:
                parent_layout_id = group_map[key]["layout_id"]
                continue

            gid = item_group_id

            add_elem(root, "項目グループ", {
                **dict(defaults["item_group"]),
                "ID": str(gid),
                "項目グループ名": group_name,
            })

            add_elem(root, "レイアウト", {
                "ID": str(layout_id),
                "Name": group_name,
                "Type": "4",
                "PropertyID": str(gid),
                "ParentsID": str(parent_layout_id),
            })

            group_map[key] = {
                "group_id": gid,
                "layout_id": layout_id,
            }

            parent_layout_id = layout_id
            layout_id += 1
            item_group_id += 1

        return parent_layout_id

    for record_name, record_info in records.items():
        for f in record_info["fields"]:
            parent_layout_id = get_group_layout(record_name, f["inner_groups"])

            iid = item_id

            add_elem(root, "レイアウト", {
                "ID": str(layout_id),
                "Name": f["item_name_norm"],
                "Type": "5",
                "PropertyID": str(iid),
                "ParentsID": str(parent_layout_id),
            })
            layout_id += 1

            add_elem(root, "項目", {
                "ID": str(iid),
                "項目名": f["item_name_norm"],
                "項目ID": f["item_id_norm"],
                "属性": attr_to_ee(f["attr"]),
                "属性チェック": "0",
            })

            ix = dict(defaults["item_xml"])
            ix["ID"] = str(iid)

            if f["required"] in ("〇", "○", "1", "true", "True", "TRUE", "必須"):
                ix["必須"] = "1"
            else:
                ix["必須"] = ix.get("必須", "0")

            if f["length"] > 0:
                ix["全体の桁数"] = str(f["length"])

            add_elem(root, "項目XML", ix)

            ia = dict(defaults["item_aux"])
            ia["ID"] = str(iid)
            add_elem(root, "項目補助情報", ia)

            item_id += 1

    return tree


# =========================
# 出力
# =========================
def indent(elem, level=0):
    i = "\n" + "  " * level

    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "  "

        for child in elem:
            indent(child, level + 1)

        if not child.tail or not child.tail.strip():
            child.tail = i

    if level and (not elem.tail or not elem.tail.strip()):
        elem.tail = i


def write_ee_xml(tree, out_path: Path):
    indent(tree.getroot())

    tmp_path = out_path.with_suffix(out_path.suffix + ".tmp")
    tree.write(tmp_path, encoding="utf-8", xml_declaration=False, short_empty_elements=True)

    body = tmp_path.read_bytes()

    out_path.parent.mkdir(parents=True, exist_ok=True)

    with open(out_path, "wb") as f:
        f.write(b'<?xml version="1.0" standalone="yes"?>\r\n')
        f.write(body)

    tmp_path.unlink(missing_ok=True)


# =========================
# main
# =========================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xml", required=True, help="EE XML形式レイアウトXMLテンプレート")
    parser.add_argument("--book", required=True, help="サンダイコー_order_データ変換定義書.xlsx")
    parser.add_argument("--sheet", default="03.ToLayout", help="ToLayoutシート名")
    parser.add_argument("--out", required=True, help="出力するEE XML形式レイアウトXML")
    args = parser.parse_args()

    template_path = Path(args.xml)
    book_path = Path(args.book)
    out_path = Path(args.out)

    wb = load_workbook(book_path, data_only=True)

    if args.sheet not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {args.sheet}")

    ws = wb[args.sheet]
    fields = read_to_layout(ws)

    tree = build_ee_xml_layout(template_path, fields)
    write_ee_xml(tree, out_path)

    print("OK")
    print(f"template : {template_path}")
    print(f"book     : {book_path}")
    print(f"sheet    : {args.sheet}")
    print(f"fields   : {len(fields)}")
    print(f"out      : {out_path}")


if __name__ == "__main__":
    main()