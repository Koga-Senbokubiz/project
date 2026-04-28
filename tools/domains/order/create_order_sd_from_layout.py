#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
create_order_sd_from_layout.py

サンダイコー案件用：
02.FromLayout から EE固定長レイアウトXMLを作成する。

重要方針：
- テンプレXMLの schema は残す
- EE固定長に必要な以下を一体で再生成する
  - レイアウト
  - データストア
  - データストアFixed
  - レコード
  - レコードFixed
  - レコード補助情報
  - レコードFixed区分値判定情報
  - 項目
  - 項目Fixed
  - 項目補助情報
- COMMON項目は各レコードへ展開する
"""

import argparse
from collections import OrderedDict
from pathlib import Path
import xml.etree.ElementTree as ET

from openpyxl import load_workbook


XS_NS = "http://www.w3.org/2001/XMLSchema"
MSDATA_NS = "urn:schemas-microsoft-com:xml-msdata"

ET.register_namespace("xs", XS_NS)
ET.register_namespace("msdata", MSDATA_NS)


# =========================
# 共通
# =========================
def local_name(tag: str) -> str:
    return tag.split("}", 1)[1] if "}" in tag else tag


def cell_str(ws, row: int, col: int) -> str:
    v = ws.cell(row, col).value
    if v is None:
        return ""
    return str(v).strip()


def find_row(ws, label: str) -> int:
    for r in range(1, min(ws.max_row, 300) + 1):
        if cell_str(ws, r, 1) == label:
            return r
    raise ValueError(f"A列に [{label}] が見つかりません。")


def to_int(value: str, default: int = 0) -> int:
    try:
        if value == "":
            return default
        return int(float(value))
    except Exception:
        return default


def add_elem(root, tag: str, attrs: dict) -> ET.Element:
    e = ET.Element(tag)
    for k, v in attrs.items():
        e.set(k, "" if v is None else str(v))
    root.append(e)
    return e


# =========================
# Excel読取
# =========================
def read_from_layout(ws):
    group_rows = []
    for i in range(6):
        label = f"グループID_lv{i}"
        try:
            group_rows.append(find_row(ws, label))
        except ValueError:
            pass

    row_repeat_group = find_row(ws, "繰返しグループID")
    row_repeat_id = find_row(ws, "繰返しID")
    row_item_id = find_row(ws, "項目ID")
    row_item_name = find_row(ws, "項目名")
    row_attr = find_row(ws, "属性")
    row_length = find_row(ws, "桁数")
    row_start = find_row(ws, "開始位置")

    fields = []

    for col in range(2, ws.max_column + 1):
        item_id = cell_str(ws, row_item_id, col)
        if not item_id:
            continue

        groups = []
        for r in group_rows:
            g = cell_str(ws, r, col)
            if g:
                groups.append(g)

        if not groups:
            continue

        lv1 = groups[1] if len(groups) >= 2 else groups[0]

        fields.append({
            "col": col,
            "groups": groups,
            "record_group": lv1,
            "repeat_group": cell_str(ws, row_repeat_group, col),
            "repeat_id": cell_str(ws, row_repeat_id, col),
            "item_id": item_id,
            "item_name": cell_str(ws, row_item_name, col) or item_id,
            "attr": cell_str(ws, row_attr, col) or "String",
            "length": to_int(cell_str(ws, row_length, col), 0),
            "start": to_int(cell_str(ws, row_start, col), 0),
        })

    if not fields:
        raise ValueError("02.FromLayout から項目を取得できませんでした。")

    return fields


def attr_to_ee(attr: str) -> str:
    a = (attr or "").strip().lower()

    if a in ("date", "datetime", "日付"):
        return "8"

    if a in ("time", "時刻"):
        return "10"

    if a in ("number", "numeric", "decimal", "integer", "int", "数値"):
        return "5"

    return "1"


def guess_record_code(record_name: str) -> str:
    name = (record_name or "").upper()

    if name.startswith("B"):
        return "B"
    if name.startswith("D"):
        return "D"
    if name.startswith("A"):
        return "A"

    if "HEADER" in name:
        return "B"
    if "DETAIL" in name:
        return "D"

    return name[:1] if name else " "


def build_records(fields):
    common_fields = [f for f in fields if f["record_group"].upper() == "COMMON"]

    records = OrderedDict()
    for f in fields:
        rg = f["record_group"]
        if rg.upper() == "COMMON":
            continue
        if rg not in records:
            records[rg] = []

    for record_name in records:
        own = [f for f in fields if f["record_group"] == record_name]
        records[record_name] = common_fields + own

    if not records:
        raise ValueError("B_HEADER / D_DETAIL などのレコードグループが見つかりません。")

    return records


# =========================
# XMLテンプレ処理
# =========================
def keep_schema_only(root):
    schema_nodes = [c for c in list(root) if local_name(c.tag) == "schema"]

    for c in list(root):
        root.remove(c)

    for s in schema_nodes:
        root.append(s)


def get_template_defaults(template_root):
    def first_attrs(tag_name: str, fallback: dict) -> dict:
        for e in template_root:
            if local_name(e.tag) == tag_name:
                return dict(e.attrib)
        return dict(fallback)

    return {
        "datastore": first_attrs("データストア", {
            "ID": "1",
            "フォーマット": "1",
        }),
        "datastore_fixed": first_attrs("データストアFixed", {
            "ID": "1",
            "レコード判別": "0",
            "EOF挿入": "0",
            "レコードセパレータフラグ": "0",
            "レコードセパレータコード": "",
            "ゼロ出力": "1",
            "小数部フォーマット": "0",
            "桁あふれ": "1",
            "不明レコードスキップ": "False",
            "スキップサイズ": "0",
            "変換先ゼロ出力": "0",
        }),
        "record_fixed": first_attrs("レコードFixed", {
            "ID": "1",
            "区分値": "",
            "開始桁": "1",
            "桁数": "1",
            "文字コード": "3",
            "属性": "1",
            "位置": "0",
            "パディング": "1",
            "SISO": "0",
            "EBCDICコードタイプ": "1",
            "JEF_KEIS漢字コードタイプ": "1",
            "レコード数": "0",
            "外字使用": "False",
        }),
        "record_aux": first_attrs("レコード補助情報", {
            "ID": "1",
            "レコード出力タイプ": "0",
            "項目ID配列": "",
            "評価方法": "0",
            "評価順序": "0",
            "with空削除": "False",
        }),
        "record_judge": first_attrs("レコードFixed区分値判定情報", {
            "ID": "1",
            "区分値": "",
            "開始桁": "1",
            "桁数": "1",
            "文字コード": "3",
            "属性": "1",
            "位置": "0",
            "パディング": "1",
            "SISO": "0",
            "EBCDICコードタイプ": "1",
            "JEF_KEIS漢字コードタイプ": "1",
            "外字使用": "False",
            "ParentID": "1",
        }),
        "item_fixed": first_attrs("項目Fixed", {
            "ID": "1",
            "文字コード": "3",
            "桁数": "1",
            "小数部桁数": "0",
            "位置": "0",
            "パディング": "1",
            "変換エラー": "0",
            "置換コード": "",
            "SISO": "0",
            "EBCDICコードタイプ": "1",
            "JEF_KEIS漢字コードタイプ": "1",
            "外字使用": "False",
            "SISOを出力する": "True",
            "全角空白のパディングコード": "0",
        }),
        "item_aux": first_attrs("項目補助情報", {
            "ID": "1",
            "全角": "False",
            "大文字": "False",
            "暦書式": "",
            "暦エラータイプ": "0",
            "レコード出力タイプ": "0",
            "最終出力タイプ": "0",
            "最終出力値": "",
            "項目出力": "True",
            "全角半角変換": "True",
            "大文字小文字変換": "True",
        }),
    }


# =========================
# EE固定長XML生成
# =========================
def build_ee_xml(template_path: Path, fields):
    template_tree = ET.parse(template_path)
    template_root = template_tree.getroot()
    defaults = get_template_defaults(template_root)

    records = build_records(fields)

    tree = ET.parse(template_path)
    root = tree.getroot()
    keep_schema_only(root)

    layout_id = 1
    record_id = 1
    item_id = 1

    # レイアウト：データストア
    add_elem(root, "レイアウト", {
        "ID": layout_id,
        "Name": "データストア",
        "Type": "1",
        "PropertyID": "1",
        "ParentsID": "0",
    })
    datastore_layout_id = layout_id
    layout_id += 1

    # レイアウト：レコード
    record_layout_id_map = {}
    record_id_map = {}

    for record_name in records.keys():
        rid = record_id
        record_id_map[record_name] = rid

        add_elem(root, "レイアウト", {
            "ID": layout_id,
            "Name": record_name,
            "Type": "3",
            "PropertyID": rid,
            "ParentsID": datastore_layout_id,
        })
        record_layout_id_map[record_name] = layout_id

        layout_id += 1
        record_id += 1

    # レイアウト：項目
    layout_item_rows = []

    for record_name, record_fields in records.items():
        parent_layout_id = record_layout_id_map[record_name]

        for f in record_fields:
            current_item_id = item_id

            add_elem(root, "レイアウト", {
                "ID": layout_id,
                "Name": f["item_name"],
                "Type": "5",
                "PropertyID": current_item_id,
                "ParentsID": parent_layout_id,
            })

            layout_item_rows.append((current_item_id, f, record_name))

            layout_id += 1
            item_id += 1

    # データストア
    ds = dict(defaults["datastore"])
    ds["ID"] = "1"
    ds["フォーマット"] = "1"
    add_elem(root, "データストア", ds)

    dsf = dict(defaults["datastore_fixed"])
    dsf["ID"] = "1"
    add_elem(root, "データストアFixed", dsf)

    # レコード系
    for record_name, rid in record_id_map.items():
        record_code = guess_record_code(record_name)

        add_elem(root, "レコード", {
            "ID": rid,
            "レコード名": record_name,
        })

        rf = dict(defaults["record_fixed"])
        rf["ID"] = str(rid)
        rf["区分値"] = record_code
        rf["開始桁"] = "1"
        rf["桁数"] = "1"
        add_elem(root, "レコードFixed", rf)

        ra = dict(defaults["record_aux"])
        ra["ID"] = str(rid)
        add_elem(root, "レコード補助情報", ra)

        rj = dict(defaults["record_judge"])
        rj["ID"] = str(rid)
        rj["区分値"] = record_code
        rj["開始桁"] = "1"
        rj["桁数"] = "1"
        rj["ParentID"] = str(rid)
        add_elem(root, "レコードFixed区分値判定情報", rj)

    # 項目系
    for iid, f, record_name in layout_item_rows:
        add_elem(root, "項目", {
            "ID": iid,
            "項目名": f["item_name"],
            "項目ID": f["item_id"],
            "属性": attr_to_ee(f["attr"]),
            "属性チェック": "0",
        })

        item_fixed = dict(defaults["item_fixed"])
        item_fixed["ID"] = str(iid)
        item_fixed["桁数"] = str(f["length"])
        add_elem(root, "項目Fixed", item_fixed)

        item_aux = dict(defaults["item_aux"])
        item_aux["ID"] = str(iid)

        # COMMONのレコードタイプは、各レコードで固定値にする
        if f["record_group"].upper() == "COMMON" and f["start"] == 1 and f["length"] == 1:
            item_aux["最終出力タイプ"] = "1"
            item_aux["最終出力値"] = guess_record_code(record_name)
        else:
            item_aux["最終出力タイプ"] = item_aux.get("最終出力タイプ", "0")
            item_aux["最終出力値"] = item_aux.get("最終出力値", "")

        add_elem(root, "項目補助情報", item_aux)

    return tree


# =========================
# 出力
# =========================
def indent(elem, level=0):
    i = "\n" + "  " * level
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "  "
        for child in elem:
            indent(child, level + 1)
        if not child.tail or not child.tail.strip():
            child.tail = i
    if level and (not elem.tail or not elem.tail.strip()):
        elem.tail = i


def write_ee_xml(tree, out_path: Path):
    indent(tree.getroot())

    tmp_path = out_path.with_suffix(out_path.suffix + ".tmp")
    tree.write(tmp_path, encoding="utf-8", xml_declaration=False, short_empty_elements=True)

    body = tmp_path.read_bytes()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with open(out_path, "wb") as f:
        f.write(b'<?xml version="1.0" standalone="yes"?>\r\n')
        f.write(body)

    tmp_path.unlink(missing_ok=True)


# =========================
# main
# =========================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xml", required=True, help="EE固定長レイアウトXMLテンプレート")
    parser.add_argument("--book", required=True, help="サンダイコー_order_データ変換定義書.xlsx")
    parser.add_argument("--sheet", default="02.FromLayout", help="FromLayoutシート名")
    parser.add_argument("--out", required=True, help="出力するEE固定長レイアウトXML")
    args = parser.parse_args()

    template_path = Path(args.xml)
    book_path = Path(args.book)
    out_path = Path(args.out)

    wb = load_workbook(book_path, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {args.sheet}")

    ws = wb[args.sheet]
    fields = read_from_layout(ws)

    tree = build_ee_xml(template_path, fields)
    write_ee_xml(tree, out_path)

    print("OK")
    print(f"template : {template_path}")
    print(f"book     : {book_path}")
    print(f"sheet    : {args.sheet}")
    print(f"fields   : {len(fields)}")
    print(f"out      : {out_path}")


if __name__ == "__main__":
    main()