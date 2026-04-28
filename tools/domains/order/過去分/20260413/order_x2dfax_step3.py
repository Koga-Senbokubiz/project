#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter


def normalize_text(value):
    return "" if value is None else str(value).strip()


def load_dfax_definition(definition_xlsx):
    """
    BigBoss_DFAX_変換定義.xlsx の 01_項目辞書 シートから
    seq / record_id / field_no / source_name をそのまま読む
    """
    wb = openpyxl.load_workbook(definition_xlsx, data_only=True)

    if "01_項目辞書" not in wb.sheetnames:
        raise RuntimeError("BigBoss_DFAX_変換定義.xlsx に '01_項目辞書' シートがありません。")

    ws = wb["01_項目辞書"]

    # 1行目が英語列名
    header = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str):
            header[v.strip()] = c

    required = ["seq", "record_id", "field_no", "source_name"]
    missing = [name for name in required if name not in header]
    if missing:
        raise RuntimeError(
            "BigBoss_DFAX_変換定義.xlsx の必要列が見つかりません: " + ", ".join(missing)
        )

    col_seq = header["seq"]
    col_record_id = header["record_id"]
    col_field_no = header["field_no"]
    col_source_name = header["source_name"]

    rows = []
    for r in range(3, ws.max_row + 1):  # 3行目からデータ
        seq = normalize_text(ws.cell(r, col_seq).value)
        record_id = normalize_text(ws.cell(r, col_record_id).value)
        field_no = normalize_text(ws.cell(r, col_field_no).value)
        source_name = normalize_text(ws.cell(r, col_source_name).value)

        if not (seq or record_id or field_no or source_name):
            continue

        rows.append({
            "no": seq,
            "record_id": record_id,
            "field_no": field_no,
            "field_name": source_name,
        })

    return rows


def set_column_widths(ws):
    widths = {
        "H": 12,   # Mapping先
        "I": 8,    # No
        "J": 14,   # レコードID
        "K": 10,   # 項目番号
        "L": 28,   # 項目名
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def write_headers(ws):
    ws["H1"] = "Mapping先"
    ws["I1"] = "No"
    ws["J1"] = "レコードID"
    ws["K1"] = "項目番号"
    ws["L1"] = "項目名"


def write_dfax_rows(ws, dfax_rows):
    start_row = 2
    for idx, row in enumerate(dfax_rows, start=start_row):
        ws.cell(idx, 9).value = row["no"]          # I
        ws.cell(idx, 10).value = row["record_id"]  # J
        ws.cell(idx, 11).value = row["field_no"]   # K
        ws.cell(idx, 12).value = row["field_name"] # L


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--step2-xlsx", required=True)
    parser.add_argument("--dfax-definition-xlsx", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    step2_xlsx = Path(args.step2_xlsx)
    dfax_xlsx = Path(args.dfax_definition_xlsx)
    out_xlsx = Path(args.out)

    print(f"[step2-xlsx] {step2_xlsx}")
    print(f"[dfax-definition-xlsx] {dfax_xlsx}")
    print(f"[out] {out_xlsx}")
    print()

    if not step2_xlsx.exists():
        print("ERROR: step2 xlsx not found")
        return 1

    if not dfax_xlsx.exists():
        print("ERROR: dfax definition xlsx not found")
        return 1

    step2_wb = openpyxl.load_workbook(step2_xlsx)
    ws = step2_wb.active

    dfax_rows = load_dfax_definition(dfax_xlsx)

    write_headers(ws)
    set_column_widths(ws)
    write_dfax_rows(ws, dfax_rows)

    step2_wb.save(out_xlsx)

    print("[OK] Step3 completed")
    print(f"step2 rows : {max(ws.max_row - 1, 0)}")
    print(f"dfax rows  : {len(dfax_rows)}")
    print(f"output     : {out_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())