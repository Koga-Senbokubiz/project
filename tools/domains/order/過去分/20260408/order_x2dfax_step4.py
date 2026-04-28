#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
order_x2dfax_step2.py

用途:
  BigBoss_DFAX_変換定義.xlsx をもとに、
  EasyExchange 用ロジック XML (bbord_dfax_fax.xml) を生成する。

前提:
  - Step2 の出力は DFAXデータではなく、EEロジックXML
  - 入力 -i は EEテンプレートXML
  - 入力 -d は 変換定義Excel
  - 出力 -o は 生成するロジックXML
  - 使用シートは 01_項目辞書
  - Excelの3行目以降を実データとして扱う
  - 項目順は seq 順
  - レコード順は 01_項目辞書上の record_id 初出順
  - 各レコードの出力項目は以下の順で生成する
      1. {record_id}_record_id
      2. 辞書行に対応する項目
      3. 項目間の SEP (01)
      4. REC_END (ff)

使い方:
  python order_x2dfax_step2.py ^
    -i template.xml ^
    -d BigBoss_DFAX_変換定義.xlsx ^
    -o bbord_dfax_fax.xml
"""

from __future__ import annotations

import argparse
import re
import sys
import xml.etree.ElementTree as ET
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


ITEM_ID_BASE = 9000000


def to_str(v) -> str:
    return "" if v is None else str(v).strip()


def normalize_key(s: str) -> str:
    return re.sub(r"[^0-9A-Za-z一-龥ぁ-んァ-ヶー]+", "", to_str(s)).lower()


def local_name(tag: str) -> str:
    if "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def indent_xml(elem: ET.Element, level: int = 0) -> None:
    i = "\n" + level * "  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "  "
        last_child = None
        for child in elem:
            indent_xml(child, level + 1)
            last_child = child
        if last_child is not None and (not last_child.tail or not last_child.tail.strip()):
            last_child.tail = i
    if level and (not elem.tail or not elem.tail.strip()):
        elem.tail = i


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create EE layout XML for Order-X2DFAX Step2.")
    parser.add_argument("-i", required=True, help="テンプレートXML")
    parser.add_argument("-d", required=True, help="変換定義Excel")
    parser.add_argument("-o", required=True, help="出力ロジックXML")
    return parser.parse_args()


def clear_non_schema(root: ET.Element) -> None:
    for child in list(root):
        if local_name(child.tag) != "schema":
            root.remove(child)


def add(root: ET.Element, name: str, attrs: dict[str, str]) -> ET.Element:
    elem = ET.Element(name)
    for k, v in attrs.items():
        elem.set(k, to_str(v))
    root.append(elem)
    return elem


def read_definition_rows(xlsx_path: Path) -> list[dict[str, str]]:
    wb = load_workbook(xlsx_path, data_only=True)

    if "01_項目辞書" not in wb.sheetnames:
        raise ValueError("シート 01_項目辞書 が見つかりません。")

    ws = wb["01_項目辞書"]

    headers = [normalize_key(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    header_index = {h: i + 1 for i, h in enumerate(headers)}

    required = {
        "seq",
        "blockname",
        "sourcesheet",
        "physicallineno",
        "lineoffset",
        "recordid",
        "fieldno",
        "key",
        "enabled",
        "sourcekind",
        "sourcename",
        "sourceexample",
        "logicalnotes",
        "formatter",
        "formattermeaning",
        "length",
        "fixedvalue",
        "notes",
        "mappingbasis",
    }

    missing = [k for k in required if k not in header_index]
    if missing:
        raise ValueError(f"01_項目辞書 の見出しが不足しています: {', '.join(missing)}")

    rows: list[dict[str, str]] = []

    for r in range(3, ws.max_row + 1):
        row = {
            key: to_str(ws.cell(row=r, column=col).value)
            for key, col in header_index.items()
            if key in required
        }

        if not row["recordid"] and not row["fieldno"]:
            continue

        enabled = row["enabled"].upper()
        if enabled in {"N", "FALSE", "0"}:
            continue

        rows.append(row)

    if not rows:
        raise ValueError("01_項目辞書 に有効なデータ行がありません。")

    rows.sort(key=lambda x: int(float(x["seq"])) if x["seq"] else 0)
    return rows


def derive_attr_calendar_position_padding(formatter: str) -> tuple[str, str, str, str]:
    f = to_str(formatter).lower()

    attr = "1"
    calendar = ""
    position = "0"
    padding = "1"

    if f == "cf_ymd":
        attr = "8"
        calendar = "yyMMdd"
    elif f == "cf_int":
        attr = "5"
        position = "1"
        padding = "2"

    return attr, calendar, position, padding


def derive_length(row: dict[str, str]) -> str:
    if row["length"]:
        return str(int(float(row["length"])))

    if row["sourcekind"].lower() == "fixed":
        return "1"

    return "20"


def build_item_name(row: dict[str, str]) -> str:
    record_id = row["recordid"]
    field_no = row["fieldno"]
    source_kind = row["sourcekind"].lower()

    if source_kind == "fixed":
        label = row["notes"] or row["fixedvalue"] or row["sourcename"] or f"seq_{row['seq']}"
        return f"{record_id}_{field_no}_fixed_{label}"

    source_name = row["sourcename"] or row["notes"] or f"seq_{row['seq']}"
    return f"{record_id}_{field_no}_{source_name}"


def build_xml(template_xml: Path, definition_xlsx: Path, output_xml: Path) -> int:
    rows = read_definition_rows(definition_xlsx)

    tree = ET.parse(template_xml)
    root = tree.getroot()

    if local_name(root.tag) != "マッピングレイアウト":
        raise ValueError("テンプレートXMLのルートが マッピングレイアウト ではありません。")

    clear_non_schema(root)

    add(root, "レイアウト", {
        "ID": "1",
        "Name": "データストア",
        "Type": "1",
        "PropertyID": "1",
        "ParentsID": "0",
    })
    add(root, "データストア", {
        "ID": "1",
        "フォーマット": "1",
    })
    add(root, "データストアFixed", {
        "ID": "1",
        "レコード判別": "0",
        "EOF挿入": "0",
        "レコードセパレータフラグ": "0",
        "レコードセパレータコード": "",
        "ゼロ出力": "1",
        "小数部フォーマット": "0",
        "桁あふれ": "1",
        "不明レコードスキップ": "False",
        "スキップサイズ": "0",
    })

    next_layout_id = 2
    next_property_id = 1

    record_groups: OrderedDict[str, list[dict[str, str]]] = OrderedDict()
    for row in rows:
        record_groups.setdefault(row["recordid"], []).append(row)

    for record_seq, (record_id, rec_rows) in enumerate(record_groups.items(), start=1):
        record_layout_id = next_layout_id
        next_layout_id += 1

        add(root, "レイアウト", {
            "ID": str(record_layout_id),
            "Name": record_id,
            "Type": "3",
            "PropertyID": str(record_seq),
            "ParentsID": "1",
        })
        add(root, "レコード", {
            "ID": str(record_seq),
            "レコード名": record_id,
        })
        add(root, "レコードFixed", {
            "ID": str(record_seq),
            "区分値": record_id,
            "開始桁": "1",
            "桁数": "5",
            "文字コード": "3",
            "属性": "1",
            "位置": "0",
            "パディング": "1",
            "SISO": "0",
            "EBCDICコードタイプ": "1",
            "JEF_KEIS漢字コードタイプ": "1",
            "レコード数": "0",
            "外字使用": "False",
        })
        add(root, "レコード補助情報", {
            "ID": str(record_seq),
            "レコード出力タイプ": "0",
            "項目ID配列": "",
            "評価方法": "0",
            "評価順序": "0",
            "with空削除": "False",
        })
        add(root, "レコードFixed区分値判定情報", {
            "ID": str(record_seq),
            "区分値": record_id,
            "開始桁": "1",
            "桁数": "5",
            "文字コード": "3",
            "属性": "1",
            "位置": "0",
            "パディング": "1",
            "SISO": "0",
            "EBCDICコードタイプ": "1",
            "JEF_KEIS漢字コードタイプ": "1",
            "外字使用": "False",
            "ParentID": str(record_seq),
        })

        prop = next_property_id
        next_property_id += 1
        item_layout_id = next_layout_id
        next_layout_id += 1

        add(root, "レイアウト", {
            "ID": str(item_layout_id),
            "Name": f"{record_id}_record_id",
            "Type": "5",
            "PropertyID": str(prop),
            "ParentsID": str(record_layout_id),
        })
        add(root, "項目", {
            "ID": str(prop),
            "項目名": f"{record_id}_record_id",
            "項目ID": str(ITEM_ID_BASE + prop),
            "属性": "1",
            "属性チェック": "0",
        })
        add(root, "項目Fixed", {
            "ID": str(prop),
            "文字コード": "3",
            "桁数": "5",
            "小数部桁数": "0",
            "位置": "0",
            "パディング": "1",
            "変換エラー": "0",
            "置換コード": "",
            "SISO": "0",
            "EBCDICコードタイプ": "1",
            "JEF_KEIS漢字コードタイプ": "1",
            "外字使用": "False",
            "SISOを出力する": "True",
        })
        add(root, "項目補助情報", {
            "ID": str(prop),
            "全角": "False",
            "大文字": "False",
            "暦書式": "",
            "暦エラータイプ": "0",
            "レコード出力タイプ": "0",
            "最終出力タイプ": "1",
            "最終出力値": record_id,
        })

        for idx, row in enumerate(rec_rows, start=1):
            attr, calendar, position, padding = derive_attr_calendar_position_padding(row["formatter"])

            if row["sourcekind"].lower() == "fixed":
                attr = "1"
                calendar = ""
                position = "0"
                padding = "1"

            prop = next_property_id
            next_property_id += 1
            item_layout_id = next_layout_id
            next_layout_id += 1

            item_name = build_item_name(row)

            add(root, "レイアウト", {
                "ID": str(item_layout_id),
                "Name": item_name,
                "Type": "5",
                "PropertyID": str(prop),
                "ParentsID": str(record_layout_id),
            })
            add(root, "項目", {
                "ID": str(prop),
                "項目名": item_name,
                "項目ID": str(ITEM_ID_BASE + prop),
                "属性": attr,
                "属性チェック": "0",
            })
            add(root, "項目Fixed", {
                "ID": str(prop),
                "文字コード": "3",
                "桁数": derive_length(row),
                "小数部桁数": "0",
                "位置": position,
                "パディング": padding,
                "変換エラー": "0",
                "置換コード": "",
                "SISO": "0",
                "EBCDICコードタイプ": "1",
                "JEF_KEIS漢字コードタイプ": "1",
                "外字使用": "False",
                "SISOを出力する": "True",
            })

            out_type = "1" if row["sourcekind"].lower() == "fixed" else "0"
            out_value = row["fixedvalue"] if row["sourcekind"].lower() == "fixed" else ""

            add(root, "項目補助情報", {
                "ID": str(prop),
                "全角": "False",
                "大文字": "False",
                "暦書式": calendar,
                "暦エラータイプ": "0",
                "レコード出力タイプ": "0",
                "最終出力タイプ": out_type,
                "最終出力値": out_value,
            })

            if idx < len(rec_rows):
                prop = next_property_id
                next_property_id += 1
                item_layout_id = next_layout_id
                next_layout_id += 1

                sep_name = f"{record_id}_SEP_{idx}"

                add(root, "レイアウト", {
                    "ID": str(item_layout_id),
                    "Name": sep_name,
                    "Type": "5",
                    "PropertyID": str(prop),
                    "ParentsID": str(record_layout_id),
                })
                add(root, "項目", {
                    "ID": str(prop),
                    "項目名": sep_name,
                    "項目ID": str(ITEM_ID_BASE + prop),
                    "属性": "1",
                    "属性チェック": "0",
                })
                add(root, "項目Fixed", {
                    "ID": str(prop),
                    "文字コード": "3",
                    "桁数": "1",
                    "小数部桁数": "0",
                    "位置": "0",
                    "パディング": "1",
                    "変換エラー": "0",
                    "置換コード": "",
                    "SISO": "0",
                    "EBCDICコードタイプ": "1",
                    "JEF_KEIS漢字コードタイプ": "1",
                    "外字使用": "False",
                    "SISOを出力する": "True",
                })
                add(root, "項目補助情報", {
                    "ID": str(prop),
                    "全角": "False",
                    "大文字": "False",
                    "暦書式": "",
                    "暦エラータイプ": "0",
                    "レコード出力タイプ": "0",
                    "最終出力タイプ": "1",
                    "最終出力値": "01",
                })

        prop = next_property_id
        next_property_id += 1
        item_layout_id = next_layout_id
        next_layout_id += 1

        rec_end_name = f"{record_id}_REC_END"

        add(root, "レイアウト", {
            "ID": str(item_layout_id),
            "Name": rec_end_name,
            "Type": "5",
            "PropertyID": str(prop),
            "ParentsID": str(record_layout_id),
        })
        add(root, "項目", {
            "ID": str(prop),
            "項目名": rec_end_name,
            "項目ID": str(ITEM_ID_BASE + prop),
            "属性": "1",
            "属性チェック": "0",
        })
        add(root, "項目Fixed", {
            "ID": str(prop),
            "文字コード": "3",
            "桁数": "1",
            "小数部桁数": "0",
            "位置": "0",
            "パディング": "1",
            "変換エラー": "0",
            "置換コード": "",
            "SISO": "0",
            "EBCDICコードタイプ": "1",
            "JEF_KEIS漢字コードタイプ": "1",
            "外字使用": "False",
            "SISOを出力する": "True",
        })
        add(root, "項目補助情報", {
            "ID": str(prop),
            "全角": "False",
            "大文字": "False",
            "暦書式": "",
            "暦エラータイプ": "0",
            "レコード出力タイプ": "0",
            "最終出力タイプ": "1",
            "最終出力値": "ff",
        })

    indent_xml(root)
    output_xml.parent.mkdir(parents=True, exist_ok=True)
    ET.ElementTree(root).write(output_xml, encoding="utf-8", xml_declaration=True)

    print(f"[OK] template : {template_xml}")
    print(f"[OK] define   : {definition_xlsx}")
    print(f"[OK] output   : {output_xml}")
    print(f"[OK] rows     : {len(rows)}")
    print(f"[OK] records  : {len(record_groups)}")
    return 0


def main() -> int:
    args = parse_args()

    template_xml = Path(args.i)
    definition_xlsx = Path(args.d)
    output_xml = Path(args.o)

    if not template_xml.exists():
        print(f"[ERROR] テンプレートXMLが存在しません: {template_xml}")
        return 1
    if not definition_xlsx.exists():
        print(f"[ERROR] 変換定義Excelが存在しません: {definition_xlsx}")
        return 1

    try:
        return build_xml(template_xml, definition_xlsx, output_xml)
    except Exception as e:
        print(f"[ERROR] {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
