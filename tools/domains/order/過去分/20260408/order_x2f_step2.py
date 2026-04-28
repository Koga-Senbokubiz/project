#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import sys
import re
from pathlib import Path
from typing import Dict, List, Tuple
import xml.etree.ElementTree as ET

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


OUTPUT_COLUMNS: List[Tuple[str, str]] = [
    ("judgement", "判定"),
    ("customer_id", "顧客ID"),
    ("customer_field_id", "顧客項目ID"),
    ("customer_tag", "顧客タグ名"),
    ("candidate_bms_field_id", "BMS項目ID"),
    ("candidate_bms_tag", "BMSタグ名"),
    ("normalized_customer_tag", "正規化タグ"),
    ("repeat_group", "繰返しグループ"),
    ("data_type", "データ型"),
    ("length_max", "最大桁数"),
    ("required_flag", "必須フラグ"),
    ("repeat_flag", "繰返しフラグ"),
    ("customer_path", "顧客パス"),
    ("parent_path", "親パス"),
    ("field_class", "項目区分"),
    ("confirmed_bms_field_id", "確定BMS項目ID"),
    ("match_method", "マッチ方法"),
    ("sample_value", "サンプル値"),
    ("remarks", "備考"),
]


# =========================================================
# 共通
# =========================================================
def normalize_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def normalize_key(value: str) -> str:
    s = normalize_text(value).lower()
    s = re.sub(r"[\s_\-./\\]+", "", s)
    s = re.sub(r"[()\[\]{}:：]", "", s)
    return s


def strip_xml_namespace(tag: str) -> str:
    if not tag:
        return ""
    if "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def strip_prefix(tag: str) -> str:
    tag = normalize_text(tag)
    if ":" in tag:
        return tag.split(":", 1)[1]
    return tag


def normalize_tag_for_match(tag: str) -> str:
    local = strip_xml_namespace(tag)
    local = strip_prefix(local)
    return normalize_key(local)


def infer_repeat_group(path_value: str) -> str:
    p = path_value.lower()

    if "lineitem" in p or "line_item" in p:
        return "lineItem"
    if "delivery" in p:
        return "delivery"
    if "summary" in p:
        return "summary"

    parts = [x for x in path_value.split("/") if x]
    if len(parts) <= 2:
        return "header"

    return ""


# =========================================================
# XML解析
# =========================================================
def extract_xml_fields(xml_path: Path) -> List[Dict[str, str]]:
    tree = ET.parse(xml_path)
    root = tree.getroot()

    results: List[Dict[str, str]] = []

    def walk(elem: ET.Element, path_parts: List[str]) -> None:
        tag = strip_xml_namespace(elem.tag)
        current_parts = path_parts + [tag]
        current_path = "/" + "/".join(current_parts)
        parent_path = "/" + "/".join(path_parts) if path_parts else ""

        text_value = normalize_text(elem.text)
        children = list(elem)

        if len(children) == 0:
            results.append({
                "customer_tag": tag,
                "customer_path": current_path,
                "parent_path": parent_path,
                "sample_value": text_value,
                "repeat_group": infer_repeat_group(current_path),
                "normalized_customer_tag": normalize_tag_for_match(tag),
            })

        for child in children:
            walk(child, current_parts)

    walk(root, [])

    # path単位で重複排除
    dedup: Dict[str, Dict[str, str]] = {}
    for row in results:
        path_key = row["customer_path"]
        if path_key not in dedup:
            dedup[path_key] = row
        else:
            if not dedup[path_key].get("sample_value") and row.get("sample_value"):
                dedup[path_key]["sample_value"] = row["sample_value"]

    return list(dedup.values())


# =========================================================
# 顧客辞書
# =========================================================
def load_customer_dictionary(dict_path: Path) -> pd.DataFrame:
    if not dict_path.exists():
        raise FileNotFoundError(f"顧客項目辞書が存在しません: {dict_path}")

    df = pd.read_excel(dict_path, header=0, dtype=str).fillna("")

    # 2行目の日本語見出し除外
    if len(df) > 0:
        first_row_values = [normalize_text(v) for v in df.iloc[0].tolist()]
        if "状態" in first_row_values or "顧客ID" in first_row_values:
            df = df.iloc[1:].reset_index(drop=True)

    return df


def build_customer_dict_indexes(dict_df: pd.DataFrame):
    by_path: Dict[str, dict] = {}
    by_tag: Dict[str, List[dict]] = {}
    by_normalized_tag: Dict[str, List[dict]] = {}

    for _, row in dict_df.iterrows():
        rec = row.to_dict()

        customer_path = normalize_text(rec.get("customer_path", ""))
        customer_tag = normalize_text(rec.get("customer_tag", ""))
        normalized_customer_tag = normalize_text(rec.get("normalized_customer_tag", ""))

        if customer_path:
            by_path[customer_path] = rec

        if customer_tag:
            by_tag.setdefault(customer_tag, []).append(rec)

        if normalized_customer_tag:
            by_normalized_tag.setdefault(normalized_customer_tag, []).append(rec)

    return by_path, by_tag, by_normalized_tag


def choose_best_customer_record(
    xml_field: dict,
    customer_indexes,
) -> Tuple[dict, str]:
    """
    returns:
      matched_record, match_basis
    """
    by_path, by_tag, by_normalized_tag = customer_indexes

    customer_path = normalize_text(xml_field.get("customer_path", ""))
    customer_tag = normalize_text(xml_field.get("customer_tag", ""))
    normalized_customer_tag = normalize_text(xml_field.get("normalized_customer_tag", ""))

    # 1. path一致最優先
    rec = by_path.get(customer_path)
    if rec:
        return rec, "パス一致"

    # 2. タグ一致
    tag_hits = by_tag.get(customer_tag, [])
    if len(tag_hits) == 1:
        return tag_hits[0], "タグ一致"

    # 3. 正規化タグ一致
    norm_hits = by_normalized_tag.get(normalized_customer_tag, [])
    if len(norm_hits) == 1:
        return norm_hits[0], "正規化タグ一致"

    # 4. タグ一致＋グループ補正
    if len(tag_hits) > 1:
        repeat_group = normalize_text(xml_field.get("repeat_group", ""))
        filtered = [
            r for r in tag_hits
            if normalize_text(r.get("repeat_group", "")) == repeat_group
        ]
        if len(filtered) == 1:
            return filtered[0], "タグ一致＋グループ補正"

    return {}, "未登録"


# =========================================================
# 判定
# =========================================================
def is_blank_attr(rec: dict, key: str) -> bool:
    return normalize_text(rec.get(key, "")) == ""


def determine_judgement(rec: dict, found: bool) -> str:
    """
    判定ルール
    - 未登録 : 顧客項目辞書に存在しない
    - 未補完 : 顧客項目辞書に存在するが、変換に必要な属性が不足
    - 変換可 : 顧客項目辞書に存在し、変換に必要な属性が揃っている
    """

    if not found:
        return "未登録"

    if (
        is_blank_attr(rec, "data_type")
        or is_blank_attr(rec, "length_max")
        or is_blank_attr(rec, "required_flag")
        or is_blank_attr(rec, "repeat_flag")
    ):
        return "未補完"

    return "変換可"

# =========================================================
# 差異表作成
# =========================================================
def create_step2_diff_table(
    input_xml: Path,
    customer_dictionary: Path,
    output_xlsx: Path,
) -> None:
    xml_fields = extract_xml_fields(input_xml)
    dict_df = load_customer_dictionary(customer_dictionary)
    customer_indexes = build_customer_dict_indexes(dict_df)

    output_rows: List[Dict[str, str]] = []

    for xml_field in xml_fields:
        matched_rec, match_basis = choose_best_customer_record(xml_field, customer_indexes)
        found = bool(matched_rec)

        if found:
            judgement = determine_judgement(matched_rec, found=True)

            row = {
                "judgement": judgement,
                "customer_id": normalize_text(matched_rec.get("customer_id", "")),
                "customer_field_id": normalize_text(matched_rec.get("customer_field_id", "")),
                "customer_tag": normalize_text(matched_rec.get("customer_tag", xml_field.get("customer_tag", ""))),
                "candidate_bms_field_id": normalize_text(matched_rec.get("candidate_bms_field_id", "")),
                "candidate_bms_tag": normalize_text(matched_rec.get("candidate_bms_tag", "")),
                "normalized_customer_tag": normalize_text(matched_rec.get("normalized_customer_tag", xml_field.get("normalized_customer_tag", ""))),
                "repeat_group": normalize_text(matched_rec.get("repeat_group", xml_field.get("repeat_group", ""))),
                "data_type": normalize_text(matched_rec.get("data_type", "")),
                "length_max": normalize_text(matched_rec.get("length_max", "")),
                "required_flag": normalize_text(matched_rec.get("required_flag", "")),
                "repeat_flag": normalize_text(matched_rec.get("repeat_flag", "")),
                "customer_path": normalize_text(matched_rec.get("customer_path", xml_field.get("customer_path", ""))),
                "parent_path": normalize_text(matched_rec.get("parent_path", xml_field.get("parent_path", ""))),
                "field_class": normalize_text(matched_rec.get("field_class", "")),
                "confirmed_bms_field_id": normalize_text(matched_rec.get("confirmed_bms_field_id", "")),
                "match_method": normalize_text(matched_rec.get("match_method", match_basis)),
                "sample_value": normalize_text(xml_field.get("sample_value", "")),
                "remarks": normalize_text(matched_rec.get("remarks", "")),
            }
        else:
            row = {
                "judgement": "未登録",
                "customer_id": "",
                "customer_field_id": "",
                "customer_tag": normalize_text(xml_field.get("customer_tag", "")),
                "candidate_bms_field_id": "",
                "candidate_bms_tag": "",
                "normalized_customer_tag": normalize_text(xml_field.get("normalized_customer_tag", "")),
                "repeat_group": normalize_text(xml_field.get("repeat_group", "")),
                "data_type": "",
                "length_max": "",
                "required_flag": "",
                "repeat_flag": "",
                "customer_path": normalize_text(xml_field.get("customer_path", "")),
                "parent_path": normalize_text(xml_field.get("parent_path", "")),
                "field_class": "",
                "confirmed_bms_field_id": "",
                "match_method": "未登録",
                "sample_value": normalize_text(xml_field.get("sample_value", "")),
                "remarks": "",
            }

        output_rows.append(row)

    out_df = pd.DataFrame(output_rows, columns=[en for en, _ in OUTPUT_COLUMNS])
    write_diff_xlsx(out_df, output_xlsx)


# =========================================================
# Excel出力
# =========================================================
def write_diff_xlsx(df: pd.DataFrame, output_xlsx: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "step2_diff"

    # 1行目 英語
    for col_idx, (en, _ja) in enumerate(OUTPUT_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=en)

    # 2行目 日本語
    for col_idx, (_en, ja) in enumerate(OUTPUT_COLUMNS, start=1):
        ws.cell(row=2, column=col_idx, value=ja)

    # 3行目以降 データ
    for row_idx, (_, row) in enumerate(df.iterrows(), start=3):
        for col_idx, (en, _ja) in enumerate(OUTPUT_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(en, ""))

    style_sheet(ws)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def style_sheet(ws) -> None:
    fill_en = PatternFill("solid", fgColor="D9EAF7")
    fill_ja = PatternFill("solid", fgColor="FFF2CC")
    fill_key = PatternFill("solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="000000")

    key_headers = {
        "judgement",
        "customer_tag",
        "candidate_bms_field_id",
        "candidate_bms_tag",
        "normalized_customer_tag",
        "repeat_group",
        "data_type",
        "length_max",
        "required_flag",
        "repeat_flag",
    }

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill_en
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = fill_ja
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value in key_headers:
            ws.cell(1, c).fill = fill_key

    for r in range(3, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 12,
        "B": 12,
        "C": 18,
        "D": 24,
        "E": 18,
        "F": 24,
        "G": 20,
        "H": 14,
        "I": 14,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 55,
        "N": 45,
        "O": 12,
        "P": 18,
        "Q": 20,
        "R": 20,
        "S": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


# =========================================================
# CLI
# =========================================================
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Order-X2F Step2 差異表作成")
    parser.add_argument("input_xml", help="order_x2f_step1.xml")
    parser.add_argument("customer_dictionary", help="order_bigboss_dictionary.xlsx")
    parser.add_argument("-o", "--output", default="order_x2f_step2.xlsx", help="出力xlsx")
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    input_xml = Path(args.input_xml)
    customer_dictionary = Path(args.customer_dictionary)
    output_xlsx = Path(args.output)

    if not input_xml.exists():
        print(f"[ERROR] 入力XMLが存在しません: {input_xml}", file=sys.stderr)
        return 1

    if not customer_dictionary.exists():
        print(f"[ERROR] 顧客項目辞書が存在しません: {customer_dictionary}", file=sys.stderr)
        return 1

    try:
        create_step2_diff_table(
            input_xml=input_xml,
            customer_dictionary=customer_dictionary,
            output_xlsx=output_xlsx,
        )
        print(f"[INFO] Step2 差異表を作成しました: {output_xlsx}")
        return 0
    except Exception as e:
        print(f"[ERROR] Step2 差異表作成に失敗しました: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())