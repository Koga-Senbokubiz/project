# -*- coding: utf-8 -*-
"""
order_x2f_step5.py

Step5: LogicXML作成
- Step1の正規化XML
- Step2の差異一覧xlsx
- 基本形テンプレートXML(TEMPLATE_FILE)
を入力にして、LogicXMLを生成する。

方針:
- Step2のB列に「〇」が入っている行を採用対象とする
- source_path -> bms_path の対応を LogicXML に落とす
- template_file はメタ情報保持と将来拡張用
- まずは採用行ベースで安定出力する
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import xml.etree.ElementTree as ET

from openpyxl import load_workbook


VERSION = "1.0.0"

DEFAULT_ROOT_TAG = "logic"
DEFAULT_MAPPING_TAG = "mapping"
DEFAULT_RULE_TAG = "rule"

SELECT_MARKS = {"〇", "○", "o", "O", "1", "Y", "y", "yes", "YES", "TRUE", "true"}

HEADER_ALIASES = {
    "selected": [
        "採用",
        "採用マッピング",
        "採用可否",
        "採用フラグ",
        "use",
        "selected",
    ],
    "source_path": [
        "顧客XPath",
        "顧客XMLパス",
        "顧客パス",
        "顧客項目パス",
        "source_path",
        "xml_path",
        "path",
        "XPath",
    ],
    "source_name": [
        "顧客項目名",
        "顧客項目",
        "source_name",
        "xml_name",
        "項目名",
    ],
    "bms_path": [
        "BMSXPath",
        "BMSパス",
        "流通BMSパス",
        "target_path",
        "bms_path",
    ],
    "bms_name": [
        "BMS項目名",
        "流通BMS項目名",
        "target_name",
        "bms_name",
    ],
    "note": [
        "備考",
        "メモ",
        "note",
    ],
}


def log(msg: str) -> None:
    print(msg, flush=True)


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_selected(value) -> bool:
    return normalize_text(value) in SELECT_MARKS


def strip_namespace(tag: str) -> str:
    if "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def split_path(path_text: str) -> List[str]:
    path_text = normalize_text(path_text)
    if not path_text:
        return []
    return [x for x in path_text.split("/") if x]


def detect_header_map(header_row: List[str]) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    normalized_headers = [normalize_text(x) for x in header_row]

    for logical_name, aliases in HEADER_ALIASES.items():
        for idx, header in enumerate(normalized_headers):
            if header in aliases:
                header_map[logical_name] = idx
                break

    return header_map


def find_path_col_by_scan(rows: List[List[str]], header_map: Dict[str, int], prefer_key: str) -> Optional[int]:
    if prefer_key in header_map:
        return header_map[prefer_key]

    if not rows:
        return None

    best_idx = None
    best_score = -1
    col_count = max(len(r) for r in rows)

    for col_idx in range(col_count):
        score = 0
        for row in rows[1:min(len(rows), 30)]:
            if col_idx >= len(row):
                continue
            val = normalize_text(row[col_idx])
            if "/" in val and len(val) >= 3:
                score += 1
        if score > best_score:
            best_score = score
            best_idx = col_idx

    if best_score <= 0:
        return None

    return best_idx


def load_step2_rows(step2_xlsx: Path) -> Tuple[List[Dict[str, str]], Dict[str, int]]:
    """
    Step2 xlsx 読み込み
    - selected列があればそれを採用判定列とする
    - なければ B列(1) を採用判定列とする
    - source_path と bms_path の両方がある行を対象とする
    """
    wb = load_workbook(step2_xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]

    raw_rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        raw_rows.append([normalize_text(c) for c in row])

    if not raw_rows:
        raise ValueError("Step2 xlsx にデータがありません。")

    header = raw_rows[0]
    data_rows = raw_rows[1:]
    header_map = detect_header_map(header)

    if "selected" not in header_map:
        header_map["selected"] = 1

    if "source_path" not in header_map:
        source_path_idx = find_path_col_by_scan(raw_rows, header_map, "source_path")
        if source_path_idx is not None:
            header_map["source_path"] = source_path_idx

    if "bms_path" not in header_map:
        bms_path_idx = find_path_col_by_scan(raw_rows, header_map, "bms_path")
        if bms_path_idx is not None:
            header_map["bms_path"] = bms_path_idx

    rows: List[Dict[str, str]] = []

    for excel_row_no, row in enumerate(data_rows, start=2):
        selected_val = row[header_map["selected"]] if header_map["selected"] < len(row) else ""
        if not is_selected(selected_val):
            continue

        source_path = row[header_map["source_path"]] if "source_path" in header_map and header_map["source_path"] < len(row) else ""
        bms_path = row[header_map["bms_path"]] if "bms_path" in header_map and header_map["bms_path"] < len(row) else ""

        if not source_path or not bms_path:
            continue

        item = {
            "excel_row_no": str(excel_row_no),
            "selected": selected_val,
            "source_path": source_path,
            "source_name": row[header_map["source_name"]] if "source_name" in header_map and header_map["source_name"] < len(row) else "",
            "bms_path": bms_path,
            "bms_name": row[header_map["bms_name"]] if "bms_name" in header_map and header_map["bms_name"] < len(row) else "",
            "note": row[header_map["note"]] if "note" in header_map and header_map["note"] < len(row) else "",
        }
        rows.append(item)

    return rows, header_map


def build_xml_leaf_map(xml_path: Path) -> Dict[str, Dict[str, str]]:
    """
    Step1 XML を解析し、leaf要素のパス辞書を返す
    """
    tree = ET.parse(xml_path)
    root = tree.getroot()

    leaf_map: Dict[str, Dict[str, str]] = {}

    def walk(elem: ET.Element, current_path: str) -> None:
        children = list(elem)
        tag_name = strip_namespace(elem.tag)
        next_path = f"{current_path}/{tag_name}" if current_path else f"/{tag_name}"

        if not children:
            leaf_map[next_path] = {
                "tag": tag_name,
                "text": normalize_text(elem.text),
                "parent": current_path if current_path else "/",
            }
            return

        for child in children:
            walk(child, next_path)

    walk(root, "")
    return leaf_map


def read_template_root_info(template_file: Path) -> Dict[str, str]:
    try:
        tree = ET.parse(template_file)
        root = tree.getroot()
        return {
            "root_tag": strip_namespace(root.tag),
            "template_name": template_file.name,
        }
    except Exception:
        return {
            "root_tag": "",
            "template_name": template_file.name,
        }


def indent_xml(elem: ET.Element, level: int = 0) -> None:
    i = "\n" + level * "  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "  "
        for child in elem:
            indent_xml(child, level + 1)
        if not child.tail or not child.tail.strip():
            child.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i


def build_logic_xml(
    input_xml: Path,
    step2_xlsx: Path,
    template_file: Path,
    selected_rows: List[Dict[str, str]],
    xml_leaf_map: Dict[str, Dict[str, str]],
) -> ET.ElementTree:
    template_info = read_template_root_info(template_file)

    root = ET.Element(DEFAULT_ROOT_TAG)
    root.set("version", VERSION)

    meta = ET.SubElement(root, "meta")
    meta.set("input_xml", str(input_xml))
    meta.set("mapping_xlsx", str(step2_xlsx))
    meta.set("template_file", str(template_file))
    meta.set("template_root_tag", template_info["root_tag"])
    meta.set("template_name", template_info["template_name"])

    summary = ET.SubElement(root, "summary")
    summary.set("selected_count", str(len(selected_rows)))
    summary.set("xml_leaf_count", str(len(xml_leaf_map)))

    mappings_elem = ET.SubElement(root, "mappings")

    used_pairs = set()
    mapping_count = 0

    for seq, row in enumerate(selected_rows, start=1):
        source_path = normalize_text(row["source_path"])
        bms_path = normalize_text(row["bms_path"])
        pair_key = (source_path, bms_path)

        if pair_key in used_pairs:
            continue

        mapping_elem = ET.SubElement(mappings_elem, DEFAULT_MAPPING_TAG)
        mapping_elem.set("seq", str(seq))
        mapping_elem.set("excel_row_no", normalize_text(row["excel_row_no"]))

        source_name = normalize_text(row["source_name"])
        if not source_name:
            source_parts = split_path(source_path)
            source_name = source_parts[-1] if source_parts else ""

        bms_name = normalize_text(row["bms_name"])
        if not bms_name:
            bms_parts = split_path(bms_path)
            bms_name = bms_parts[-1] if bms_parts else ""

        from_elem = ET.SubElement(mapping_elem, "from")
        from_elem.set("path", source_path)
        from_elem.set("name", source_name)

        source_info = xml_leaf_map.get(source_path)
        if source_info:
            from_elem.set("exists_in_step1_xml", "true")
            from_elem.set("sample_value", normalize_text(source_info.get("text", "")))
            from_elem.set("parent_path", normalize_text(source_info.get("parent", "")))
        else:
            from_elem.set("exists_in_step1_xml", "false")
            from_elem.set("sample_value", "")
            from_elem.set("parent_path", "")

        to_elem = ET.SubElement(mapping_elem, "to")
        to_elem.set("path", bms_path)
        to_elem.set("name", bms_name)

        # 初版では固定値の move ルール
        rule_elem = ET.SubElement(mapping_elem, DEFAULT_RULE_TAG)
        rule_elem.set("type", "move")
        rule_elem.set("mode", "direct")

        note = normalize_text(row["note"])
        if note:
            note_elem = ET.SubElement(mapping_elem, "note")
            note_elem.text = note

        used_pairs.add(pair_key)
        mapping_count += 1

    summary.set("mapping_count", str(mapping_count))

    warnings = ET.SubElement(root, "warnings")

    for row in selected_rows:
        source_path = normalize_text(row["source_path"])
        if source_path and source_path not in xml_leaf_map:
            warn = ET.SubElement(warnings, "warning")
            warn.set("type", "missing_source_path_in_step1_xml")
            warn.set("excel_row_no", normalize_text(row["excel_row_no"]))
            warn.text = source_path

    return ET.ElementTree(root)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Step5: LogicXML作成")
    parser.add_argument("--input-xml", required=True, help="Step1出力XML (order_x2f_step1.xml)")
    parser.add_argument("--mapping-xlsx", required=True, help="Step2出力xlsx (order_x2f_step2.xlsx)")
    parser.add_argument("--template-file", required=True, help="基本形テンプレートXML")
    parser.add_argument("--output-xml", required=True, help="Step5出力XML (order_x2f_step5_logic_xml.xml)")
    return parser.parse_args()


def validate_inputs(input_xml: Path, mapping_xlsx: Path, template_file: Path) -> None:
    if not input_xml.exists():
        raise FileNotFoundError(f"入力XMLが見つかりません: {input_xml}")
    if not mapping_xlsx.exists():
        raise FileNotFoundError(f"Step2 xlsx が見つかりません: {mapping_xlsx}")
    if not template_file.exists():
        raise FileNotFoundError(f"テンプレートXMLが見つかりません: {template_file}")


def main() -> int:
    args = parse_args()

    input_xml = Path(args.input_xml)
    mapping_xlsx = Path(args.mapping_xlsx)
    template_file = Path(args.template_file)
    output_xml = Path(args.output_xml)

    try:
        log(f"[INFO] order_x2f_step5.py VERSION={VERSION}")

        validate_inputs(input_xml, mapping_xlsx, template_file)

        log("[INFO] Step1 XML 解析開始")
        xml_leaf_map = build_xml_leaf_map(input_xml)
        log(f"[INFO] Step1 XML leaf数: {len(xml_leaf_map)}")

        log("[INFO] Step2 xlsx 読み込み開始")
        selected_rows, header_map = load_step2_rows(mapping_xlsx)
        log(f"[INFO] 採用行数: {len(selected_rows)}")
        log(f"[INFO] 検出ヘッダ: {header_map}")

        log("[INFO] Step5 LogicXML 生成開始")
        tree = build_logic_xml(
            input_xml=input_xml,
            step2_xlsx=mapping_xlsx,
            template_file=template_file,
            selected_rows=selected_rows,
            xml_leaf_map=xml_leaf_map,
        )

        output_xml.parent.mkdir(parents=True, exist_ok=True)
        indent_xml(tree.getroot())
        tree.write(output_xml, encoding="utf-8", xml_declaration=True)

        log(f"[INFO] 出力完了: {output_xml}")
        return 0

    except Exception as e:
        log(f"[ERROR] Step5失敗: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())