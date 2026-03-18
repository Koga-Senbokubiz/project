# -*- coding: utf-8 -*-
"""
order_x2f_step4.py
Ver. 2026-03-18 template-transform edition

目的:
- Step2 XLSX
- 変換先テンプレートレイアウトXML
をもとに、

「MappingEditor で読める to レイアウトXML」

を生成する。

方針:
- テンプレートXMLをそのまま土台に使う
- Step2 の「採用マッピング」に対応する bms_path を抽出
- テンプレート内の対応レイアウトを残し、それ以外を削除
- 祖先ノード（親レコード/親グループ）は残す
- 出力は EasyExchange のレイアウトXML形式のまま

想定引数:
  --mapping-xlsx
  --template-file
  --output-xml

互換用:
  --input-xml が渡されても無視可能
"""

from __future__ import annotations

import argparse
import copy
import os
import sys
import traceback
import xml.etree.ElementTree as ET
from collections import defaultdict
from typing import Dict, List, Set, Tuple

from openpyxl import load_workbook


# ============================================================
# 共通
# ============================================================

def log(msg: str) -> None:
    print(msg, flush=True)


def safe_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def local_name(name: str) -> str:
    s = safe_str(name)
    if "}" in s:
        s = s.split("}", 1)[1]
    if ":" in s:
        s = s.split(":", 1)[1]
    return s


def normalize_path(path: str) -> str:
    p = safe_str(path).replace("\\", "/")
    if not p:
        return ""

    while "//" in p:
        p = p.replace("//", "/")

    if not p.startswith("/"):
        p = "/" + p

    parts = [x for x in p.split("/") if x]
    parts = [local_name(x) for x in parts]
    return "/" + "/".join(parts)


def indent_xml(elem: ET.Element, level: int = 0) -> None:
    i = "\n" + level * "  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "  "
        for child in elem:
            indent_xml(child, level + 1)
        if not child.tail or not child.tail.strip():
            child.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i


# ============================================================
# Step2 XLSX 読み込み
# ============================================================

COLUMN_ALIASES = {
    "seq": [
        "seq", "no", "番号", "連番", "id", "No"
    ],
    "source_path": [
        "source_path", "src_path", "xml_path", "from_path",
        "変換元パス", "元パス", "変換元XPath", "XMLパス"
    ],
    "source_name": [
        "source_name", "src_name", "xml_name", "from_name",
        "変換元項目名", "元項目名", "XMLタグ"
    ],
    "bms_path": [
        "bms_path", "target_path", "to_path",
        "変換先パス", "BMSパス", "変換先XPath"
    ],
    "bms_name": [
        "bms_name", "target_name", "to_name",
        "変換先項目名", "BMS項目名"
    ],
    "adopt": [
        "adopt", "採用", "use", "selected", "採用マッピング", "match"
    ],
}


def normalize_header_name(name: str) -> str:
    return safe_str(name).replace(" ", "").replace("　", "").lower()


def detect_header_row(ws, scan_rows: int = 10) -> int:
    alias_flat = set()
    for aliases in COLUMN_ALIASES.values():
        for a in aliases:
            alias_flat.add(normalize_header_name(a))

    best_row = 1
    best_score = -1

    for r in range(1, min(scan_rows, ws.max_row) + 1):
        score = 0
        for c in range(1, ws.max_column + 1):
            v = normalize_header_name(ws.cell(r, c).value)
            if v in alias_flat:
                score += 1
        if score > best_score:
            best_score = score
            best_row = r

    return best_row


def build_header_map(ws, header_row: int) -> Dict[str, int]:
    raw_headers: Dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        raw_headers[c] = normalize_header_name(ws.cell(header_row, c).value)

    result: Dict[str, int] = {}
    for logical_name, aliases in COLUMN_ALIASES.items():
        alias_set = {normalize_header_name(x) for x in aliases}
        for c, header_name in raw_headers.items():
            if header_name in alias_set:
                result[logical_name] = c
                break
    return result


def choose_sheet(wb):
    for ws in wb.worksheets:
        if ws.sheet_state == "visible":
            return ws
    return wb.worksheets[0]


def is_adopted(value: str) -> bool:
    v = safe_str(value)
    if v == "":
        return False
    return v in {"○", "〇", "1", "true", "True", "TRUE", "yes", "YES", "採用", "候補あり"}


def load_mapping_rows(xlsx_file: str) -> List[Dict[str, str]]:
    wb = load_workbook(xlsx_file, data_only=True)
    ws = choose_sheet(wb)

    header_row = detect_header_row(ws)
    header_map = build_header_map(ws, header_row)

    if "bms_path" not in header_map:
        raise ValueError("Step2 XLSX に bms_path 相当の列が見つかりません。")

    rows: List[Dict[str, str]] = []

    for r in range(header_row + 1, ws.max_row + 1):
        bms_path = safe_str(ws.cell(r, header_map["bms_path"]).value)
        if not bms_path:
            continue

        row: Dict[str, str] = {}
        row["seq"] = safe_str(ws.cell(r, header_map.get("seq", 0)).value) if "seq" in header_map else str(r)
        row["source_path"] = (
            normalize_path(ws.cell(r, header_map["source_path"]).value)
            if "source_path" in header_map
            else ""
        )
        row["source_name"] = (
            safe_str(ws.cell(r, header_map["source_name"]).value)
            if "source_name" in header_map
            else ""
        )
        row["bms_path"] = normalize_path(bms_path)
        row["bms_name"] = (
            safe_str(ws.cell(r, header_map["bms_name"]).value)
            if "bms_name" in header_map
            else local_name(bms_path)
        )
        row["adopt"] = (
            safe_str(ws.cell(r, header_map["adopt"]).value)
            if "adopt" in header_map
            else ""
        )
        row["excel_row_no"] = str(r)

        rows.append(row)

    return rows


# ============================================================
# テンプレート解析
# ============================================================

class TemplateModel:
    def __init__(self, tree: ET.ElementTree) -> None:
        self.tree = tree
        self.root = tree.getroot()

        self.layouts: Dict[str, ET.Element] = {}
        self.children: Dict[str, List[str]] = defaultdict(list)

        self.records: Dict[str, ET.Element] = {}
        self.record_xml: Dict[str, ET.Element] = {}
        self.record_aux: Dict[str, ET.Element] = {}

        self.groups: Dict[str, ET.Element] = {}
        self.items: Dict[str, ET.Element] = {}
        self.item_xml: Dict[str, ET.Element] = {}
        self.item_aux: Dict[str, ET.Element] = {}

        self.layout_path_map: Dict[str, str] = {}
        self.path_to_layout_ids: Dict[str, List[str]] = defaultdict(list)

        self._load()

    def _load(self) -> None:
        for elem in self.root.findall("レイアウト"):
            lid = elem.attrib["ID"]
            self.layouts[lid] = elem

        for lid, elem in self.layouts.items():
            parent_id = elem.attrib.get("ParentsID", "")
            self.children[parent_id].append(lid)

        for elem in self.root.findall("レコード"):
            self.records[elem.attrib["ID"]] = elem
        for elem in self.root.findall("レコードXML"):
            self.record_xml[elem.attrib["ID"]] = elem
        for elem in self.root.findall("レコード補助情報"):
            self.record_aux[elem.attrib["ID"]] = elem

        for elem in self.root.findall("項目グループ"):
            self.groups[elem.attrib["ID"]] = elem
        for elem in self.root.findall("項目"):
            self.items[elem.attrib["ID"]] = elem
        for elem in self.root.findall("項目XML"):
            self.item_xml[elem.attrib["ID"]] = elem
        for elem in self.root.findall("項目補助情報"):
            self.item_aux[elem.attrib["ID"]] = elem

        for lid in self.layouts.keys():
            p = self.build_layout_path(lid)
            self.layout_path_map[lid] = p
            self.path_to_layout_ids[p].append(lid)

    def build_layout_path(self, layout_id: str) -> str:
        parts: List[str] = []
        current = layout_id

        while current and current in self.layouts:
            elem = self.layouts[current]
            typ = elem.attrib.get("Type", "")
            name = elem.attrib.get("Name", "")

            # Type=1 はデータストアなのでパスに入れない
            if typ != "1":
                parts.append(local_name(name))

            parent_id = elem.attrib.get("ParentsID", "")
            if parent_id == "0":
                break
            current = parent_id

        parts.reverse()
        return normalize_path("/" + "/".join(parts))

    def find_matching_layout_ids(self, bms_path: str) -> List[str]:
        """
        Step2 の bms_path に対応するテンプレート layout ID を返す。
        - exact match
        - bms_path が template_path で終わる
        - 最長一致優先
        """
        tgt = normalize_path(bms_path)

        exact = self.path_to_layout_ids.get(tgt, [])
        if exact:
            return exact

        candidates: List[Tuple[int, str, str]] = []
        for tpl_path, layout_ids in self.path_to_layout_ids.items():
            if not tpl_path:
                continue
            if tgt.endswith(tpl_path):
                candidates.append((len(tpl_path), tpl_path, layout_ids[0]))

        if not candidates:
            return []

        candidates.sort(reverse=True, key=lambda x: x[0])
        best_len = candidates[0][0]
        best_paths = [x[1] for x in candidates if x[0] == best_len]

        result: List[str] = []
        for p in best_paths:
            result.extend(self.path_to_layout_ids[p])
        return result


# ============================================================
# keep判定
# ============================================================

def collect_adopted_bms_paths(mapping_rows: List[Dict[str, str]]) -> List[str]:
    adopted_paths: List[str] = []
    adopted_exists = any(is_adopted(r.get("adopt", "")) for r in mapping_rows)

    for row in mapping_rows:
        bms_path = normalize_path(row.get("bms_path", ""))
        if not bms_path:
            continue

        if adopted_exists:
            if not is_adopted(row.get("adopt", "")):
                continue

        adopted_paths.append(bms_path)

    return adopted_paths


def collect_keep_layout_ids(model: TemplateModel, adopted_bms_paths: List[str]) -> Tuple[Set[str], List[str]]:
    keep_layout_ids: Set[str] = set()
    unmatched_paths: List[str] = []

    for tgt in adopted_bms_paths:
        matched_ids = model.find_matching_layout_ids(tgt)
        if not matched_ids:
            unmatched_paths.append(tgt)
            continue

        for lid in matched_ids:
            current = lid
            while current and current in model.layouts:
                keep_layout_ids.add(current)
                parent_id = model.layouts[current].attrib.get("ParentsID", "")
                if not parent_id or parent_id == "0":
                    break
                current = parent_id

    # データストア root(Type=1) は常に残す
    for lid, elem in model.layouts.items():
        if elem.attrib.get("Type") == "1":
            keep_layout_ids.add(lid)

    return keep_layout_ids, unmatched_paths


# ============================================================
# テンプレート削減
# ============================================================

def prune_template_tree(tree: ET.ElementTree, keep_layout_ids: Set[str]) -> ET.ElementTree:
    new_tree = copy.deepcopy(tree)
    root = new_tree.getroot()

    layouts = {e.attrib["ID"]: e for e in root.findall("レイアウト")}
    records = {e.attrib["ID"]: e for e in root.findall("レコード")}
    record_xml = {e.attrib["ID"]: e for e in root.findall("レコードXML")}
    record_aux = {e.attrib["ID"]: e for e in root.findall("レコード補助情報")}
    groups = {e.attrib["ID"]: e for e in root.findall("項目グループ")}
    items = {e.attrib["ID"]: e for e in root.findall("項目")}
    item_xml = {e.attrib["ID"]: e for e in root.findall("項目XML")}
    item_aux = {e.attrib["ID"]: e for e in root.findall("項目補助情報")}

    keep_record_ids: Set[str] = set()
    keep_group_ids: Set[str] = set()
    keep_item_ids: Set[str] = set()

    for lid in keep_layout_ids:
        layout = layouts.get(lid)
        if layout is None:
            continue

        typ = layout.attrib.get("Type", "")
        pid = layout.attrib.get("PropertyID", "")

        if typ == "3" and pid:
            keep_record_ids.add(pid)
        elif typ == "4" and pid:
            keep_group_ids.add(pid)
        elif typ == "5" and pid:
            keep_item_ids.add(pid)

    def remove_unkept(elem_map: Dict[str, ET.Element], keep_ids: Set[str]) -> None:
        for eid, elem in list(elem_map.items()):
            if eid not in keep_ids:
                root.remove(elem)

    # レイアウト削除
    for lid, elem in list(layouts.items()):
        if lid not in keep_layout_ids:
            root.remove(elem)

    # 関連定義削除
    remove_unkept(records, keep_record_ids)
    remove_unkept(record_xml, keep_record_ids)
    remove_unkept(record_aux, keep_record_ids)

    remove_unkept(groups, keep_group_ids)
    remove_unkept(items, keep_item_ids)
    remove_unkept(item_xml, keep_item_ids)
    remove_unkept(item_aux, keep_item_ids)

    return new_tree


# ============================================================
# 保存
# ============================================================

def save_xml(tree: ET.ElementTree, output_xml: str) -> None:
    out_dir = os.path.dirname(output_xml)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    indent_xml(tree.getroot())
    tree.write(output_xml, encoding="utf-8", xml_declaration=True)


# ============================================================
# 引数
# ============================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Order-X2F Step4")
    parser.add_argument("--mapping-xlsx", required=True, help="Step2 mapping xlsx")
    parser.add_argument("--template-file", required=True, help="Target template layout xml")
    parser.add_argument("--output-xml", required=True, help="Step4 output layout xml")
    parser.add_argument("--input-xml", required=False, help="Compatibility option (unused)")
    return parser.parse_args()


def validate_args(args: argparse.Namespace) -> None:
    if not os.path.isfile(args.mapping_xlsx):
        raise FileNotFoundError(f"MAPPING_XLSX が見つかりません: {args.mapping_xlsx}")
    if not os.path.isfile(args.template_file):
        raise FileNotFoundError(f"TEMPLATE_FILE が見つかりません: {args.template_file}")


# ============================================================
# main
# ============================================================

def main() -> int:
    try:
        args = parse_args()
        validate_args(args)

        log("==================================================")
        log("Order-X2F Step4 start")
        log("==================================================")
        log(f"MAPPING_XLSX  = {args.mapping_xlsx}")
        log(f"TEMPLATE_FILE = {args.template_file}")
        log(f"OUTPUT_XML    = {args.output_xml}")
        if args.input_xml:
            log(f"INPUT_XML     = {args.input_xml} (unused)")
        log("")

        log("[1/4] Step2 XLSX 読込...")
        mapping_rows = load_mapping_rows(args.mapping_xlsx)
        log(f"  mapping row count = {len(mapping_rows)}")

        log("[2/4] 採用 bms_path 抽出...")
        adopted_bms_paths = collect_adopted_bms_paths(mapping_rows)
        log(f"  adopted bms path count = {len(adopted_bms_paths)}")

        log("[3/4] テンプレート読込...")
        template_tree = ET.parse(args.template_file)
        model = TemplateModel(template_tree)
        log(f"  template layout count = {len(model.layouts)}")

        log("[4/4] keep layout 判定 → 削減 → 保存...")
        keep_layout_ids, unmatched_paths = collect_keep_layout_ids(model, adopted_bms_paths)
        log(f"  keep layout count = {len(keep_layout_ids)}")
        log(f"  unmatched adopted path count = {len(unmatched_paths)}")

        if unmatched_paths:
            log("  unmatched adopted paths (先頭20件):")
            for p in unmatched_paths[:20]:
                log(f"    {p}")

        output_tree = prune_template_tree(template_tree, keep_layout_ids)
        save_xml(output_tree, args.output_xml)

        log("")
        log("==================================================")
        log("Order-X2F Step4 finished")
        log(f"OUTPUT_XML = {args.output_xml}")
        log("==================================================")
        return 0

    except Exception as e:
        log("")
        log("[ERROR] Step4 failed")
        log(str(e))
        log("")
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())