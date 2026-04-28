#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
order_xml_layout.py

XML から 02.FromLayout / 03.ToLayout 用の横展開レイアウトを生成する。

【仕様】
- from / to 共通
- A列ラベルは動的生成
- A列は 02.FromLayout と同じ思想で
    - 太字
    - 中央揃え
    - グレー塗り
    - 枠線あり
- データ列はグループ別に色分け
- 必須はデフォルト「〇」
- 繰返しは「同一親配下に同名タグが複数あるノード」を自動判定
- 繰返し配下は、レイアウト定義としては「最初の1件だけ」を採用する
  （2件目以降を横に増やさない）
- --max-group-level に応じて lv0～lvN を増やす

【A列構成】
グループID_lv0
グループID_lv1
...
グループID_lvN
繰返しグループID
繰返しID
項目ID
項目名
属性
桁数
開始位置
必須
サンプル1件目
"""

from __future__ import annotations

import argparse
import re
import xml.etree.ElementTree as ET
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# =========================
# 定数
# =========================
MAX_DEFAULT_GROUP_LEVEL = 5

# 色
FILL_LABEL = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")
FILL_COMMON = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
FILL_BLUE = PatternFill(fill_type="solid", start_color="DDEBF7", end_color="DDEBF7")
FILL_ORANGE = PatternFill(fill_type="solid", start_color="FCE4D6", end_color="FCE4D6")
FILL_GREEN = PatternFill(fill_type="solid", start_color="E2F0D9", end_color="E2F0D9")
FILL_PURPLE = PatternFill(fill_type="solid", start_color="E4DFEC", end_color="E4DFEC")
FILL_PINK = PatternFill(fill_type="solid", start_color="FCE4EC", end_color="FCE4EC")
FILL_CYAN = PatternFill(fill_type="solid", start_color="DDEBF7", end_color="DDEBF7")

GROUP_FILL_CYCLE = [FILL_BLUE, FILL_ORANGE, FILL_GREEN, FILL_PURPLE, FILL_PINK, FILL_CYAN]

THIN = Side(style="thin", color="A6A6A6")
ALL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT_LABEL = Font(bold=True)
ALIGN_LABEL = Alignment(vertical="center", horizontal="center", wrap_text=True)
ALIGN_DATA = Alignment(vertical="center", horizontal="center", wrap_text=True)
ALIGN_SAMPLE = Alignment(vertical="center", horizontal="left", wrap_text=True)


# =========================
# 行位置管理
# =========================
@dataclass
class RowMap:
    row_group_start: int
    row_repeat_group_id: int
    row_repeat_id: int
    row_field_id: int
    row_field_name: int
    row_data_type: int
    row_length: int
    row_start_pos: int
    row_required: int
    row_sample: int

    @property
    def max_header_row(self) -> int:
        return self.row_required

    def row_group(self, lv: int) -> int:
        return self.row_group_start + lv


def build_row_map(max_group_level: int) -> RowMap:
    row_group_start = 1
    row_repeat_group_id = row_group_start + (max_group_level + 1)
    row_repeat_id = row_repeat_group_id + 1
    row_field_id = row_repeat_id + 1
    row_field_name = row_field_id + 1
    row_data_type = row_field_name + 1
    row_length = row_data_type + 1
    row_start_pos = row_length + 1
    row_required = row_start_pos + 1
    row_sample = row_required + 1

    return RowMap(
        row_group_start=row_group_start,
        row_repeat_group_id=row_repeat_group_id,
        row_repeat_id=row_repeat_id,
        row_field_id=row_field_id,
        row_field_name=row_field_name,
        row_data_type=row_data_type,
        row_length=row_length,
        row_start_pos=row_start_pos,
        row_required=row_required,
        row_sample=row_sample,
    )


# =========================
# データクラス
# =========================
@dataclass
class LayoutField:
    group_levels: List[str]
    repeat_group_id: str
    repeat_id: str
    field_id: str
    field_name: str
    data_type: str
    length: int
    start_pos: str
    required: str
    sample_value: str


# =========================
# XML処理
# =========================
def local_name(tag: str) -> str:
    if "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def normalize_id(text: str) -> str:
    text = local_name(text).strip()
    text = re.sub(r"\s+", "_", text)
    return text


def detect_repeat_paths(root: ET.Element) -> set[Tuple[str, ...]]:
    """
    同一親配下に同名タグが複数あるノードのパスを繰返しパスとして検出する。
    """
    repeat_paths: set[Tuple[str, ...]] = set()

    def walk(elem: ET.Element, path: Tuple[str, ...]) -> None:
        children = [c for c in list(elem) if isinstance(c.tag, str)]
        counter: Dict[str, int] = defaultdict(int)

        for child in children:
            counter[local_name(child.tag)] += 1

        processed_repeat_name = set()

        for child in children:
            child_name = local_name(child.tag)
            child_path = path + (child_name,)

            if counter[child_name] > 1:
                repeat_paths.add(child_path)
                # 同一繰返しノードは最初の1件だけ潜る
                if child_name in processed_repeat_name:
                    continue
                processed_repeat_name.add(child_name)

            walk(child, child_path)

    walk(root, (local_name(root.tag),))
    return repeat_paths


def nearest_repeat_group_id(path_without_field: Tuple[str, ...], repeat_paths: set[Tuple[str, ...]]) -> str:
    for i in range(len(path_without_field), 0, -1):
        sub = path_without_field[:i]
        if sub in repeat_paths:
            return normalize_id(sub[-1])
    return ""


def collect_leaf_fields(root: ET.Element, max_group_level: int) -> List[LayoutField]:
    """
    葉要素を収集する。
    繰返しノード配下は、レイアウト定義として最初の1件だけ採用する。
    """
    repeat_paths = detect_repeat_paths(root)
    repeat_id_map: Dict[str, int] = {}
    fields: List[LayoutField] = []

    def walk(elem: ET.Element, path: Tuple[str, ...]) -> None:
        children = [c for c in list(elem) if isinstance(c.tag, str)]

        if not children:
            field_tag = local_name(elem.tag)
            field_id = normalize_id(field_tag)
            field_name = field_tag
            sample_value = (elem.text or "").strip()
            length = len(sample_value)

            groups = [normalize_id(p) for p in path[:-1]]

            slot_count = max_group_level + 1
            if len(groups) > slot_count:
                groups = groups[:slot_count]
            else:
                groups = groups + [""] * (slot_count - len(groups))

            repeat_group_id = nearest_repeat_group_id(path[:-1], repeat_paths)

            if repeat_group_id:
                if repeat_group_id not in repeat_id_map:
                    repeat_id_map[repeat_group_id] = len(repeat_id_map) + 1
                repeat_id = str(repeat_id_map[repeat_group_id])
            else:
                repeat_id = ""

            fields.append(
                LayoutField(
                    group_levels=groups,
                    repeat_group_id=repeat_group_id,
                    repeat_id=repeat_id,
                    field_id=field_id,
                    field_name=field_name,
                    data_type="String",
                    length=length,
                    start_pos="",
                    required="〇",
                    sample_value=sample_value,
                )
            )
            return

        counter: Dict[str, int] = defaultdict(int)
        for child in children:
            counter[local_name(child.tag)] += 1

        processed_repeat_name = set()

        for child in children:
            child_name = local_name(child.tag)
            child_path = path + (child_name,)

            if counter[child_name] > 1:
                if child_name in processed_repeat_name:
                    continue
                processed_repeat_name.add(child_name)

            walk(child, child_path)

    walk(root, (local_name(root.tag),))
    return fields


# =========================
# Excel出力
# =========================
def write_a_column_labels(ws, max_group_level: int, rowmap: RowMap) -> None:
    for lv in range(max_group_level + 1):
        ws.cell(rowmap.row_group(lv), 1).value = f"グループID_lv{lv}"

    ws.cell(rowmap.row_repeat_group_id, 1).value = "繰返しグループID"
    ws.cell(rowmap.row_repeat_id, 1).value = "繰返しID"
    ws.cell(rowmap.row_field_id, 1).value = "項目ID"
    ws.cell(rowmap.row_field_name, 1).value = "項目名"
    ws.cell(rowmap.row_data_type, 1).value = "属性"
    ws.cell(rowmap.row_length, 1).value = "桁数"
    ws.cell(rowmap.row_start_pos, 1).value = "開始位置"
    ws.cell(rowmap.row_required, 1).value = "必須"


def clear_sheet(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.alignment = Alignment()
            cell.font = Font(bold=False)


def effective_group_name(field: LayoutField) -> str:
    """
    色分け用の代表グループを返す。
    優先順:
    - lv1～lvN の最も深い非空グループ
    - lv0
    """
    non_empty = [g for g in field.group_levels if g]
    if not non_empty:
        return "DEFAULT"
    if len(non_empty) >= 2:
        return non_empty[-1]
    return non_empty[0]


def build_group_fill_map(fields: List[LayoutField]) -> Dict[str, PatternFill]:
    fill_map: Dict[str, PatternFill] = {}
    palette_index = 0

    for field in fields:
        group_name = effective_group_name(field)
        if group_name in fill_map:
            continue

        if group_name.upper() == "COMMON":
            fill_map[group_name] = FILL_COMMON
        else:
            fill_map[group_name] = GROUP_FILL_CYCLE[palette_index % len(GROUP_FILL_CYCLE)]
            palette_index += 1

    return fill_map


def apply_styles(ws, fields: List[LayoutField], max_group_level: int, rowmap: RowMap) -> None:
    max_col = 1 + len(fields)
    max_row = rowmap.max_header_row

    # A列
    for r in range(1, max_row + 1):
        cell = ws.cell(r, 1)
        cell.fill = FILL_LABEL
        cell.border = ALL_BORDER
        cell.alignment = ALIGN_LABEL
        cell.font = FONT_LABEL

    # データ列
    fill_map = build_group_fill_map(fields)

    for idx, field in enumerate(fields, start=2):
        group_name = effective_group_name(field)
        fill = fill_map[group_name]

        for r in range(1, max_row + 1):
            cell = ws.cell(r, idx)
            cell.fill = fill
            cell.border = ALL_BORDER
            cell.alignment = ALIGN_DATA

    # サンプル行
    sample_label = ws.cell(rowmap.row_sample, 1)
    sample_label.value = 1
    sample_label.border = ALL_BORDER
    sample_label.alignment = Alignment(vertical="center", horizontal="right")
    sample_label.font = Font(bold=False)

    for c in range(2, max_col + 1):
        cell = ws.cell(rowmap.row_sample, c)
        cell.border = ALL_BORDER
        cell.alignment = ALIGN_SAMPLE

    # 行高さ
    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = 24
    ws.row_dimensions[rowmap.row_field_name].height = 34

    # 列幅
    ws.column_dimensions["A"].width = 18
    for c in range(2, max_col + 1):
        ws.column_dimensions[ws.cell(1, c).column_letter].width = 18

    # 枠固定
    ws.freeze_panes = "B1"


def write_layout_sheet(ws, fields: List[LayoutField], max_group_level: int) -> None:
    rowmap = build_row_map(max_group_level)

    clear_sheet(ws)
    write_a_column_labels(ws, max_group_level, rowmap)

    col = 2
    for f in fields:
        for lv, group_id in enumerate(f.group_levels):
            ws.cell(rowmap.row_group(lv), col).value = group_id

        ws.cell(rowmap.row_repeat_group_id, col).value = f.repeat_group_id
        ws.cell(rowmap.row_repeat_id, col).value = f.repeat_id
        ws.cell(rowmap.row_field_id, col).value = f.field_id
        ws.cell(rowmap.row_field_name, col).value = f.field_name
        ws.cell(rowmap.row_data_type, col).value = f.data_type
        ws.cell(rowmap.row_length, col).value = f.length
        ws.cell(rowmap.row_start_pos, col).value = f.start_pos
        ws.cell(rowmap.row_required, col).value = f.required
        ws.cell(rowmap.row_sample, col).value = f.sample_value

        col += 1

    apply_styles(ws, fields, max_group_level, rowmap)


# =========================
# メイン
# =========================
def main() -> None:
    parser = argparse.ArgumentParser(description="XMLからFrom/To共通レイアウトシートを生成する")
    parser.add_argument("--xml", required=True, help="入力XML")
    parser.add_argument("--mode", choices=["from", "to"], default="to", help="from/to")
    parser.add_argument("--book", default="", help="既存ブックに書き込む場合に指定")
    parser.add_argument("--sheet", default="", help="出力シート名。省略時は mode に応じて自動決定")
    parser.add_argument("--out", default="", help="出力xlsx。book省略時は xml名ベースで保存")
    parser.add_argument("--max-group-level", type=int, default=MAX_DEFAULT_GROUP_LEVEL, help="lv0～lvN の最大N")
    args = parser.parse_args()

    xml_path = Path(args.xml)
    if not xml_path.exists():
        raise FileNotFoundError(f"XMLが見つかりません: {xml_path}")

    root = ET.parse(xml_path).getroot()
    fields = collect_leaf_fields(root, args.max_group_level)
    if not fields:
        raise ValueError("XMLから葉要素を抽出できませんでした。")

    sheet_name = args.sheet if args.sheet else ("02.FromLayout" if args.mode == "from" else "03.ToLayout")

    if args.book:
        book_path = Path(args.book)
        if not book_path.exists():
            raise FileNotFoundError(f"ブックが見つかりません: {book_path}")
        wb = load_workbook(book_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        write_layout_sheet(ws, fields, args.max_group_level)
        out_path = Path(args.out) if args.out else book_path.with_name(f"{book_path.stem}_{args.mode}_layout{book_path.suffix}")
        wb.save(out_path)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        write_layout_sheet(ws, fields, args.max_group_level)
        out_path = Path(args.out) if args.out else xml_path.with_name(f"{xml_path.stem}_{args.mode}_layout.xlsx")
        wb.save(out_path)

    print("完了")
    print(f"入力XML      : {xml_path}")
    print(f"mode         : {args.mode}")
    print(f"出力シート   : {sheet_name}")
    print(f"max level    : {args.max_group_level}")
    print(f"項目数       : {len(fields)}")
    print(f"出力ファイル : {out_path}")


if __name__ == "__main__":
    main()