#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import xml.etree.ElementTree as ET

from openpyxl import load_workbook


DEFAULT_MAX_GROUP_LEVEL = 5


@dataclass
class LayoutColumn:
    col: int
    group_levels: List[str]
    repeat_group_id: str
    repeat_id: str
    field_id: str
    field_name: str
    data_type: str
    length: Optional[int]

    @property
    def full_path(self) -> Tuple[str, ...]:
        groups = [g for g in self.group_levels if g]
        return tuple(groups + [self.field_id])

    @property
    def repeat_group_path(self) -> Tuple[str, ...]:
        groups = [g for g in self.group_levels if g]
        if not self.repeat_group_id:
            return tuple()
        if groups and groups[-1] == self.repeat_group_id:
            return tuple(groups)
        for i in range(len(groups) - 1, -1, -1):
            if groups[i] == self.repeat_group_id:
                return tuple(groups[: i + 1])
        return tuple(groups + [self.repeat_group_id])


def normalize_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def local_name(tag: str) -> str:
    if "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def find_label_rows(ws, max_scan_row: int = 50) -> Dict[str, List[int]]:
    labels: Dict[str, List[int]] = {}
    for r in range(1, min(ws.max_row, max_scan_row) + 1):
        label = normalize_value(ws.cell(r, 1).value)
        if label:
            labels.setdefault(label, []).append(r)
    return labels


def resolve_layout_rows(ws, max_group_level: int) -> Dict[str, int | List[int]]:
    labels = find_label_rows(ws)

    def require_one(label: str) -> int:
        rows = labels.get(label, [])
        if not rows:
            raise ValueError(f"A列にラベル [{label}] が見つかりません。")
        return rows[0]

    group_rows: List[int] = []
    for i in range(max_group_level + 1):
        label = f"グループID_lv{i}"
        rows = labels.get(label, [])
        if rows:
            group_rows.append(rows[0])

    if not group_rows:
        raise ValueError("グループID_lv0 ～ グループID_lvN の行が見つかりません。")

    repeat_group_row = require_one("繰返しグループID") if labels.get("繰返しグループID") else -1
    repeat_id_row = require_one("繰返しID") if labels.get("繰返しID") else -1
    field_id_row = require_one("項目ID")
    field_name_row = require_one("項目名")
    data_type_row = require_one("属性")
    length_row = require_one("桁数")
    required_row = require_one("必須")

    return {
        "group_rows": group_rows,
        "repeat_group_row": repeat_group_row,
        "repeat_id_row": repeat_id_row,
        "field_id_row": field_id_row,
        "field_name_row": field_name_row,
        "data_type_row": data_type_row,
        "length_row": length_row,
        "required_row": required_row,
        "sample_start_row": required_row + 1,
    }


def parse_layout_columns(ws, max_group_level: int) -> tuple[List[LayoutColumn], int]:
    rows = resolve_layout_rows(ws, max_group_level)
    group_rows: List[int] = rows["group_rows"]
    repeat_group_row: int = rows["repeat_group_row"]
    repeat_id_row: int = rows["repeat_id_row"]
    field_id_row: int = rows["field_id_row"]
    field_name_row: int = rows["field_name_row"]
    data_type_row: int = rows["data_type_row"]
    length_row: int = rows["length_row"]
    sample_start_row: int = rows["sample_start_row"]

    columns: List[LayoutColumn] = []
    for col in range(2, ws.max_column + 1):
        field_id = normalize_value(ws.cell(field_id_row, col).value)
        if not field_id:
            continue
        field_name = normalize_value(ws.cell(field_name_row, col).value)
        data_type = normalize_value(ws.cell(data_type_row, col).value)
        length_raw = ws.cell(length_row, col).value
        length = None
        try:
            if length_raw not in (None, ""):
                length = int(length_raw)
        except ValueError:
            length = None
        group_levels = [normalize_value(ws.cell(r, col).value) for r in group_rows]
        repeat_group_id = normalize_value(ws.cell(repeat_group_row, col).value) if repeat_group_row > 0 else ""
        repeat_id = normalize_value(ws.cell(repeat_id_row, col).value) if repeat_id_row > 0 else ""
        columns.append(LayoutColumn(col, group_levels, repeat_group_id, repeat_id, field_id, field_name, data_type, length))

    return columns, sample_start_row


def clear_existing_sample_area(ws, start_row: int) -> None:
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None


def build_tree_index(root: ET.Element) -> Dict[Tuple[str, ...], List[ET.Element]]:
    index: Dict[Tuple[str, ...], List[ET.Element]] = defaultdict(list)

    def walk(elem: ET.Element, path: Tuple[str, ...]) -> None:
        index[path].append(elem)
        for child in list(elem):
            if not isinstance(child.tag, str):
                continue
            child_name = local_name(child.tag)
            walk(child, path + (child_name,))

    root_name = local_name(root.tag)
    walk(root, (root_name,))
    return index


def first_text_by_path(index: Dict[Tuple[str, ...], List[ET.Element]], path: Tuple[str, ...]) -> str:
    elems = index.get(path, [])
    if not elems:
        return ""
    return (elems[0].text or "").strip()


def relative_tail(full_path: Tuple[str, ...], prefix: Tuple[str, ...]) -> Tuple[str, ...]:
    if len(full_path) < len(prefix):
        return tuple()
    if full_path[:len(prefix)] != prefix:
        return tuple()
    return full_path[len(prefix):]


def descendant_text(instance: ET.Element, relative_path: Tuple[str, ...]) -> str:
    if not relative_path:
        return (instance.text or "").strip()
    current = instance
    for name in relative_path:
        found = None
        for child in list(current):
            if isinstance(child.tag, str) and local_name(child.tag) == name:
                found = child
                break
        if found is None:
            return ""
        current = found
    return (current.text or "").strip()


def build_repeat_instances(index: Dict[Tuple[str, ...], List[ET.Element]], columns: List[LayoutColumn]) -> Dict[Tuple[str, ...], List[ET.Element]]:
    repeat_paths = {c.repeat_group_path for c in columns if c.repeat_group_path}
    repeat_instances: Dict[Tuple[str, ...], List[ET.Element]] = {}
    for rp in repeat_paths:
        repeat_instances[rp] = index.get(rp, [])
    return repeat_instances


def expand_xml_data(ws, root: ET.Element, columns: List[LayoutColumn], start_row: int) -> int:
    index = build_tree_index(root)
    repeat_instances = build_repeat_instances(index, columns)
    max_repeat_count = max((len(v) for v in repeat_instances.values()), default=1)
    row_count = max_repeat_count if max_repeat_count > 0 else 1

    for row_offset in range(row_count):
        excel_row = start_row + row_offset
        ws.cell(excel_row, 1).value = row_offset + 1
        for coldef in columns:
            value = ""
            if coldef.repeat_group_path:
                instances = repeat_instances.get(coldef.repeat_group_path, [])
                if row_offset < len(instances):
                    instance = instances[row_offset]
                    rel = relative_tail(coldef.full_path, coldef.repeat_group_path)
                    value = descendant_text(instance, rel)
            else:
                value = first_text_by_path(index, coldef.full_path)
            ws.cell(excel_row, coldef.col).value = value
    return row_count


def main() -> None:
    parser = argparse.ArgumentParser(description="XMLレイアウトシートに実データXMLを展開する")
    parser.add_argument("--book", required=True, help="データ変換定義書.xlsx")
    parser.add_argument("--xml", required=True, help="実データXML")
    parser.add_argument("--sheet", default="02.FromLayout", help="対象シート名")
    parser.add_argument("--max-group-level", type=int, default=DEFAULT_MAX_GROUP_LEVEL, help="グループ階層最大深さ（例: 5 -> lv0～lv5）")
    parser.add_argument("--out", default="", help="出力先xlsx。省略時は元ブックに _xml_expanded を付ける")
    args = parser.parse_args()

    book_path = Path(args.book)
    xml_path = Path(args.xml)
    if not book_path.exists():
        raise FileNotFoundError(f"定義書が見つかりません: {book_path}")
    if not xml_path.exists():
        raise FileNotFoundError(f"XMLが見つかりません: {xml_path}")

    wb = load_workbook(book_path)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {args.sheet}")
    ws = wb[args.sheet]

    columns, sample_start_row = parse_layout_columns(ws, args.max_group_level)
    if not columns:
        raise ValueError("レイアウト項目を読めませんでした。シート構成を確認してください。")

    root = ET.parse(xml_path).getroot()
    clear_existing_sample_area(ws, sample_start_row)
    row_count = expand_xml_data(ws, root, columns, sample_start_row)

    if args.out:
        out_path = Path(args.out)
    else:
        out_path = book_path.with_name(f"{book_path.stem}_xml_expanded{book_path.suffix}")

    wb.save(out_path)
    print("完了")
    print(f"入力ブック : {book_path}")
    print(f"入力XML   : {xml_path}")
    print(f"対象シート: {args.sheet}")
    print(f"読込項目数: {len(columns)}")
    print(f"展開件数  : {row_count}")
    print(f"出力      : {out_path}")


if __name__ == "__main__":
    main()