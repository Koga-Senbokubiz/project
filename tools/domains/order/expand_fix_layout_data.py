#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
データ変換定義書の 02.FromLayout に、実データ(DAT)を横展開するツール（改訂版）。

改訂内容:
- 項目ID / 項目名の重複行がない新レイアウトに対応
- A列ラベルを見て行位置を自動判定
- 旧レイアウト（重複あり）にもできるだけ対応
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


DEFAULT_DATA_START_ROW = 14
DEFAULT_MAX_GROUP_LEVEL = 5


def normalize_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def detect_line_separator(raw: bytes) -> bytes:
    if b"\r\n" in raw:
        return b"\r\n"
    if b"\n" in raw:
        return b"\n"
    if b"\r" in raw:
        return b"\r"
    return b""


def read_records(dat_path: Path, encoding: str = "cp932") -> List[str]:
    raw = dat_path.read_bytes()
    sep = detect_line_separator(raw)
    if sep:
        parts = raw.split(sep)
    else:
        parts = [raw]

    records: List[str] = []
    for part in parts:
        if not part:
            continue
        records.append(part.decode(encoding, errors="replace"))
    return records


def find_label_rows(ws, max_scan_row: int = 50) -> Dict[str, List[int]]:
    labels: Dict[str, List[int]] = {}
    for r in range(1, min(ws.max_row, max_scan_row) + 1):
        label = normalize_value(ws.cell(r, 1).value)
        if not label:
            continue
        labels.setdefault(label, []).append(r)
    return labels


def resolve_layout_rows(ws, max_group_level: int) -> Dict[str, int | List[int]]:
    labels = find_label_rows(ws)

    def require_one(label: str) -> int:
        rows = labels.get(label, [])
        if not rows:
            raise ValueError(f"A列にラベル [{label}] が見つかりません。")
        return rows[0]

    def require_last(label: str) -> int:
        rows = labels.get(label, [])
        if not rows:
            raise ValueError(f"A列にラベル [{label}] が見つかりません。")
        return rows[-1]

    group_rows: List[int] = []
    for i in range(max_group_level + 1):
        label = f"グループID_lv{i}"
        rows = labels.get(label, [])
        if rows:
            group_rows.append(rows[0])

    if not group_rows:
        raise ValueError("グループID_lv0 ～ グループID_lvN の行が見つかりません。")

    repeat_group_rows = labels.get("繰返しグループID", [])
    repeat_id_rows = labels.get("繰返しID", [])

    field_id_row = require_last("項目ID")
    field_name_row = require_last("項目名")
    data_type_row = require_one("属性")
    length_row = require_one("桁数")
    start_pos_row = require_one("開始位置")

    return {
        "group_rows": group_rows,
        "repeat_group_row": repeat_group_rows[0] if repeat_group_rows else -1,
        "repeat_id_row": repeat_id_rows[0] if repeat_id_rows else -1,
        "field_id_row": field_id_row,
        "field_name_row": field_name_row,
        "data_type_row": data_type_row,
        "length_row": length_row,
        "start_pos_row": start_pos_row,
    }


def clear_existing_sample_area(ws, start_row: int) -> None:
    max_col = ws.max_column
    max_row = ws.max_row
    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).value = None


def parse_field_definitions(ws, max_group_level: int) -> List[Dict]:
    rows = resolve_layout_rows(ws, max_group_level)

    group_rows: List[int] = rows["group_rows"]  # type: ignore[assignment]
    repeat_group_row: int = rows["repeat_group_row"]  # type: ignore[assignment]
    repeat_id_row: int = rows["repeat_id_row"]  # type: ignore[assignment]
    field_id_row: int = rows["field_id_row"]  # type: ignore[assignment]
    field_name_row: int = rows["field_name_row"]  # type: ignore[assignment]
    data_type_row: int = rows["data_type_row"]  # type: ignore[assignment]
    length_row: int = rows["length_row"]  # type: ignore[assignment]
    start_pos_row: int = rows["start_pos_row"]  # type: ignore[assignment]

    fields: List[Dict] = []

    for col in range(2, ws.max_column + 1):
        field_id = normalize_value(ws.cell(field_id_row, col).value)
        if not field_id:
            continue

        field_name = normalize_value(ws.cell(field_name_row, col).value)
        data_type = normalize_value(ws.cell(data_type_row, col).value)
        length_val = ws.cell(length_row, col).value
        start_pos_val = ws.cell(start_pos_row, col).value

        if start_pos_val in (None, "") or length_val in (None, ""):
            continue

        try:
            start_pos = int(start_pos_val)
            length = int(length_val)
        except ValueError:
            continue

        group_levels = [normalize_value(ws.cell(r, col).value) for r in group_rows]
        repeat_group_id = normalize_value(ws.cell(repeat_group_row, col).value) if repeat_group_row > 0 else ""
        repeat_id = normalize_value(ws.cell(repeat_id_row, col).value) if repeat_id_row > 0 else ""

        fields.append(
            {
                "col": col,
                "field_id": field_id,
                "field_name": field_name,
                "data_type": data_type,
                "length": length,
                "start_pos": start_pos,
                "group_levels": group_levels,
                "repeat_group_id": repeat_group_id,
                "repeat_id": repeat_id,
            }
        )

    return fields


def infer_target_record_type(group_levels: List[str]) -> Optional[str]:
    for group in group_levels:
        if not group:
            continue
        if group == "COMMON":
            return None
        m = re.match(r"^([A-Z])(?:_|$)", group)
        if m:
            return m.group(1)
    return None


def field_applies_to_record(field_def: Dict, rec_type: str) -> bool:
    target = infer_target_record_type(field_def["group_levels"])
    if target is None:
        return True
    return target == rec_type


def slice_value(record: str, start_pos: int, length: int) -> str:
    start_idx = start_pos - 1
    end_idx = start_idx + length
    return record[start_idx:end_idx]


def expand_samples(ws, records: List[str], fields: List[Dict], start_row: int) -> int:
    row = start_row
    seq = 1

    for record in records:
        if not record:
            continue

        rec_type = record[:1]
        ws.cell(row, 1).value = seq

        for field_def in fields:
            col = field_def["col"]
            if field_applies_to_record(field_def, rec_type):
                value = slice_value(record, field_def["start_pos"], field_def["length"])
                ws.cell(row, col).value = value
            else:
                ws.cell(row, col).value = None

        row += 1
        seq += 1

    return seq - 1


def main() -> None:
    parser = argparse.ArgumentParser(description="02.FromLayout に実データを横展開する")
    parser.add_argument("--book", required=True, help="データ変換定義書.xlsx")
    parser.add_argument("--dat", required=True, help="固定長DATファイル")
    parser.add_argument("--sheet", default="02.FromLayout", help="対象シート名")
    parser.add_argument("--start-row", type=int, default=DEFAULT_DATA_START_ROW, help="実データ展開開始行")
    parser.add_argument("--encoding", default="cp932", help="DATファイルの文字コード")
    parser.add_argument(
        "--max-group-level",
        type=int,
        default=DEFAULT_MAX_GROUP_LEVEL,
        help="グループ階層の最大深さ（例: 5 -> lv0～lv5 を読む）",
    )
    parser.add_argument("--out", default="", help="出力先xlsx。省略時は元ブックに _expanded を付けて保存")
    args = parser.parse_args()

    book_path = Path(args.book)
    dat_path = Path(args.dat)

    if not book_path.exists():
        raise FileNotFoundError(f"定義書が見つかりません: {book_path}")
    if not dat_path.exists():
        raise FileNotFoundError(f"DATファイルが見つかりません: {dat_path}")

    wb = load_workbook(book_path)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {args.sheet}")

    ws = wb[args.sheet]

    fields = parse_field_definitions(ws, args.max_group_level)
    if not fields:
        raise ValueError("項目定義を読めませんでした。02.FromLayout のレイアウトを確認してください。")

    records = read_records(dat_path, encoding=args.encoding)
    if not records:
        raise ValueError("DATファイルにレコードが見つかりませんでした。")

    clear_existing_sample_area(ws, args.start_row)
    rec_count = expand_samples(ws, records, fields, args.start_row)

    if args.out:
        out_path = Path(args.out)
    else:
        out_path = book_path.with_name(f"{book_path.stem}_expanded{book_path.suffix}")

    wb.save(out_path)

    print("完了")
    print(f"入力ブック: {book_path}")
    print(f"DATファイル : {dat_path}")
    print(f"対象シート : {args.sheet}")
    print(f"読込項目数 : {len(fields)}")
    print(f"展開件数   : {rec_count}")
    print(f"出力ファイル: {out_path}")


if __name__ == "__main__":
    main()
