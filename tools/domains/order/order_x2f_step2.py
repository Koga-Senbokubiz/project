#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Step2
顧客XML項目と流通BMS標準XMLとの差異を抽出し、
XML vs BMS マッピング表（Excel）を作成する。
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CURRENT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = CURRENT_DIR.parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from tools.common.build_xml_bms_mapping_rows import build_mapping_rows  # noqa: E402


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build Step2 XML vs BMS mapping Excel"
    )
    parser.add_argument("input_xml", help="Step1 output XML file")
    parser.add_argument("output_xlsx", help="Step2 output XLSX file")
    parser.add_argument("template_file", help="standard XML file")
    return parser.parse_args()


def ensure_parent_dir(file_path: str) -> None:
    parent = os.path.dirname(os.path.abspath(file_path))
    if parent:
        os.makedirs(parent, exist_ok=True)


def auto_fit_width(ws) -> None:
    max_widths = {}

    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            width = len(value) + 2
            col = cell.column
            if col not in max_widths or width > max_widths[col]:
                max_widths[col] = width

    for col, width in max_widths.items():
        adjusted = min(max(width, 10), 60)
        ws.column_dimensions[get_column_letter(col)].width = adjusted


def write_mapping_excel(output_xlsx: str, rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "XML_BMS_Mapping"

    # B列を「採用マッピング」に変更
    headers = [
        "No",               # A
        "採用マッピング",    # B
        "XMLパス",          # C
        "XMLタグ",          # D
        "サンプル値",        # E
        "出現回数",         # F
        "BMS項目名",        # G
        "BMSタグ",          # H
        "BMSパス",          # I
        "必須",             # J
        "桁数",             # K
        "データ型",          # L
        "状態",             # M
        "突合方法",         # N
        "備考",             # O
    ]

    ws.append(headers)

    # ヘッダ書式（背景色はヘッダのみ維持。不要ならこれも外せます）
    header_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # データ行
    for idx, row in enumerate(rows, start=1):
        # M列（状態）が「候補あり」の場合だけ B列に〇
        adopted_mark = "〇" if row.get("status", "") == "候補あり" else ""

        ws.append(
            [
                idx,                              # A No
                adopted_mark,                     # B 採用マッピング
                row.get("xml_path", ""),          # C XMLパス
                row.get("xml_tag", ""),           # D XMLタグ
                row.get("sample_value", ""),      # E サンプル値
                row.get("occurs", ""),            # F 出現回数
                row.get("bms_item_name", ""),     # G BMS項目名
                row.get("bms_tag", ""),           # H BMSタグ
                row.get("bms_path", ""),          # I BMSパス
                row.get("required", ""),          # J 必須
                row.get("length", ""),            # K 桁数
                row.get("data_type", ""),         # L データ型
                row.get("status", ""),            # M 状態
                row.get("match_type", ""),        # N 突合方法
                row.get("notes", ""),             # O 備考
            ]
        )

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # 背景色設定は廃止

    # 配置
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].alignment = center   # A No
        row[1].alignment = center   # B 採用マッピング
        row[5].alignment = center   # F 出現回数
        row[9].alignment = center   # J 必須
        row[10].alignment = center  # K 桁数
        row[11].alignment = center  # L データ型
        row[12].alignment = center  # M 状態
        row[13].alignment = center  # N 突合方法

    auto_fit_width(ws)
    ensure_parent_dir(output_xlsx)
    wb.save(output_xlsx)


def main() -> None:
    args = parse_args()

    input_xml = os.path.abspath(args.input_xml)
    output_xlsx = os.path.abspath(args.output_xlsx)
    template_file = os.path.abspath(args.template_file)

    print("==========================================")
    print("Order-X2F Step2 start")
    print("==========================================")
    print(f"input_xml    : {input_xml}")
    print(f"output_xlsx  : {output_xlsx}")
    print(f"template_file: {template_file}")

    if not os.path.exists(input_xml):
        raise FileNotFoundError(f"input xml not found: {input_xml}")

    if not os.path.exists(template_file):
        raise FileNotFoundError(f"template file not found: {template_file}")

    rows = build_mapping_rows(
        xml_file=input_xml,
        template_file=template_file,
    )

    write_mapping_excel(output_xlsx, rows)

    print("------------------------------------------")
    print(f"rows         : {len(rows)}")
    print(f"output saved : {output_xlsx}")
    print("Order-X2F Step2 finished")
    print("------------------------------------------")


if __name__ == "__main__":
    main()